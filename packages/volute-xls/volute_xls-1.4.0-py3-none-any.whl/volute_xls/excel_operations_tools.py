#!/usr/bin/env python3
"""
Excel Operations MCP Tools

This module provides three MCP tools for Excel operations:
1. excel_cell_operations: Set values, format cells, copy/paste ranges
2. excel_row_column_operations: Insert/delete/resize/hide rows and columns  
3. excel_sheet_operations: Add/delete/rename/move/copy worksheets

These tools use a DSL (Domain Specific Language) format based on the existing
volute-xls parser and writer infrastructure.
"""

import logging
import json
from typing import Dict, Any, List, Optional, Union
from pathlib import Path
import re
import sys

# Setup logging
logger = logging.getLogger(__name__)

# Add the current directory to Python path for imports
root_dir_path = Path(__file__).parent.absolute()
sys.path.append(str(root_dir_path))

try:
    from parser.excel_parser import parse_markdown_formulas, write_formulas_to_excel_complex_agent
    from writer.excel_writer import ExcelWriter
except ImportError as e:
    logger.error(f"Failed to import required modules: {e}")
    raise


def excel_cell_operations(excel_path: str, cell_operations: str) -> str:
    """
    Perform cell operations on an Excel file using DSL format.
    
    This tool handles:
    • Set cell values (text, numbers, formulas)
    • Clear cell contents  
    • Copy/paste cell ranges
    • Format cells (fonts, colors, borders, alignment, etc.)
    • Data validation and comments
    • Conditional formatting
    
    Args:
        excel_path: Path to the Excel file (.xlsx, .xlsm, .xls)
        cell_operations: DSL string specifying cell operations
        
    Simple DSL Format:
        A1 = "text value"
        B1 = 123
        C1 = =SUM(A1:B1)
        A1 font bold
        B1 background yellow
        
    Or Full DSL Format:
        sheet_name: SheetName | A1, "=SUM(B1:B10)", b=true | B1, 100
        
    Returns:
        JSON string with updated cell information
    """
    logger.info(f"Starting cell operations on {excel_path}")
    
    try:
        # Validate Excel file exists
        excel_file = Path(excel_path)
        if not excel_file.exists():
            return json.dumps({"error": f"Excel file not found: {excel_path}"})
            
        # Convert simple DSL to markdown format if needed
        if not cell_operations.strip().startswith('sheet_name:'):
            # Convert simple DSL format to markdown format
            sheet_name = _get_default_sheet_name(str(excel_file))
            markdown_operations = _convert_simple_dsl_to_markdown(cell_operations, sheet_name)
        else:
            markdown_operations = cell_operations
            
        # Parse the DSL operations
        logger.debug(f"Parsing cell operations DSL: {markdown_operations[:200]}...")
        parsed_operations = parse_markdown_formulas(markdown_operations)
        
        if not parsed_operations:
            return json.dumps({"error": "Failed to parse cell operations DSL"})
        
        # Execute the operations using the writer
        logger.info("Executing cell operations")
        updated_cells = write_formulas_to_excel_complex_agent(str(excel_file), parsed_operations)
        
        if not updated_cells:
            return json.dumps({"error": "No cells were updated"})
            
        result = {
            "success": True,
            "excel_file": str(excel_file),
            "operation": "cell_operations",
            "updated_cells": updated_cells,
            "cells_modified": len(updated_cells)
        }
        
        logger.info(f"Cell operations completed successfully. Modified {len(updated_cells)} cells.")
        return json.dumps(result, indent=2)
        
    except Exception as e:
        error_msg = f"Error during cell operations: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return json.dumps({"error": error_msg})


def excel_row_column_operations(excel_path: str, row_column_operations: str) -> str:
    """
    Perform row and column operations on an Excel file using DSL format.
    
    This tool handles:
    • Insert/delete rows and columns
    • Resize rows and columns (height/width)
    • Hide/unhide rows and columns
    • Move rows and columns
    • Format entire rows/columns
    
    Args:
        excel_path: Path to the Excel file (.xlsx, .xlsm, .xls)
        row_column_operations: DSL string specifying row/column operations
        
    DSL Format:
        sheet_name: SheetName | operation_type: parameters | next_operation: parameters
        
    Operations:
        # Row Operations
        - insert_row: row_number=5, count=2, copy_format_from=4
        - delete_row: row_number=3, count=1  
        - resize_row: row_number=2, height=25.5
        - hide_row: row_number=1:5 (range supported)
        - unhide_row: row_number=1:5
        - move_row: from_row=3, to_row=7, count=2
        - format_row: row_number=1, properties="b=true,sz=14,fill=#CCCCCC"
        
        # Column Operations  
        - insert_column: column=C, count=3, copy_format_from=B
        - delete_column: column=D, count=2
        - resize_column: column=A, width=15.5
        - hide_column: column=B:F (range supported)
        - unhide_column: column=B:F  
        - move_column: from_column=C, to_column=G, count=2
        - format_column: column=A, properties="ha=center,b=true"
        
    Examples:
        # Insert 2 rows at position 5, copying format from row 4
        "sheet_name: Data | insert_row: row_number=5, count=2, copy_format_from=4"
        
        # Delete column C and resize column A
        "sheet_name: Sheet1 | delete_column: column=C, count=1 | resize_column: column=A, width=20"
        
        # Hide rows 10-15 and format row 1 as header
        "sheet_name: Report | hide_row: row_number=10:15 | format_row: row_number=1, properties=\"b=true,sz=14,fill=#DDDDDD\""
        
        # Multiple operations
        "sheet_name: Data | insert_column: column=B, count=1 | resize_column: column=A, width=25 | format_column: column=B, properties=\"ha=center\""
        
    Returns:
        JSON string with operation results
    """
    logger.info(f"Starting row/column operations on {excel_path}")
    
    try:
        # Validate Excel file exists
        excel_file = Path(excel_path)
        if not excel_file.exists():
            return json.dumps({"error": f"Excel file not found: {excel_path}"})
        
        # Parse operations
        operations = _parse_row_column_operations(row_column_operations)
        if not operations:
            return json.dumps({"error": "Failed to parse row/column operations DSL"})
        
        # Execute operations using xlwings
        results = _execute_row_column_operations(str(excel_file), operations)
        
        result = {
            "success": True,
            "excel_file": str(excel_file), 
            "operation": "row_column_operations",
            "results": results,
            "operations_count": len(results)
        }
        
        logger.info(f"Row/column operations completed successfully. {len(results)} operations executed.")
        return json.dumps(result, indent=2)
        
    except Exception as e:
        error_msg = f"Error during row/column operations: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return json.dumps({"error": error_msg})


def excel_sheet_operations(excel_path: str, sheet_operations: str) -> str:
    """
    Perform sheet operations on an Excel file using DSL format.
    
    This tool handles:
    • Add/delete worksheets
    • Rename worksheets  
    • Move/copy worksheets
    • Hide/unhide worksheets
    • Set worksheet tab colors
    • Duplicate sheets with data
    
    Args:
        excel_path: Path to the Excel file (.xlsx, .xlsm, .xls)
        sheet_operations: DSL string specifying sheet operations
        
    DSL Format:
        operation_type: parameters | next_operation: parameters
        
    Operations:
        # Sheet Management
        - add_sheet: name="New Sheet", position=2, template_sheet="Sheet1"
        - delete_sheet: name="Sheet2"
        - rename_sheet: old_name="Sheet1", new_name="Data Analysis"
        - copy_sheet: source="Template", name="Copy of Template", position=3
        - move_sheet: name="Sheet3", position=1
        
        # Visibility and Styling
        - hide_sheet: name="Hidden Data"
        - unhide_sheet: name="Hidden Data"
        - set_tab_color: name="Sheet1", color="#FF0000"
        - activate_sheet: name="Summary"
        
        # Bulk Operations
        - duplicate_sheet: source="Template", names=["Q1 Data","Q2 Data","Q3 Data"], positions=[2,3,4]
        - protect_sheet: name="Final Report", password="secret123"
        - unprotect_sheet: name="Final Report", password="secret123"
        
    Examples:
        # Add new sheet after position 1
        "add_sheet: name=\"Analysis\", position=2"
        
        # Rename and set tab color
        "rename_sheet: old_name=\"Sheet1\", new_name=\"Sales Data\" | set_tab_color: name=\"Sales Data\", color=\"#00FF00\""
        
        # Copy sheet and hide original
        "copy_sheet: source=\"Template\", name=\"Q1 Report\", position=2 | hide_sheet: name=\"Template\""
        
        # Multiple sheet creation
        "duplicate_sheet: source=\"Monthly Template\", names=[\"January\",\"February\",\"March\"], positions=[2,3,4]"
        
        # Delete multiple sheets
        "delete_sheet: name=\"Sheet2\" | delete_sheet: name=\"Sheet3\""
        
    Returns:
        JSON string with operation results
    """
    logger.info(f"Starting sheet operations on {excel_path}")
    
    try:
        # Validate Excel file exists
        excel_file = Path(excel_path)
        if not excel_file.exists():
            return json.dumps({"error": f"Excel file not found: {excel_path}"})
        
        # Parse operations
        operations = _parse_sheet_operations(sheet_operations)
        if not operations:
            return json.dumps({"error": "Failed to parse sheet operations DSL"})
        
        # Execute operations using xlwings
        results = _execute_sheet_operations(str(excel_file), operations)
        
        result = {
            "success": True,
            "excel_file": str(excel_file),
            "operation": "sheet_operations", 
            "results": results,
            "operations_count": len(results)
        }
        
        logger.info(f"Sheet operations completed successfully. {len(results)} operations executed.")
        return json.dumps(result, indent=2)
        
    except Exception as e:
        error_msg = f"Error during sheet operations: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return json.dumps({"error": error_msg})


# Helper functions for parsing and executing operations

def _parse_row_column_operations(operations_dsl: str) -> List[Dict[str, Any]]:
    """Parse row/column operations DSL into structured operations list."""
    logger.debug("Parsing row/column operations DSL")
    
    operations = []
    
    try:
        # Split by sheet_name: to separate different sheets
        sheet_sections = [s.strip() for s in operations_dsl.split('sheet_name:') if s.strip()]
        
        for section in sheet_sections:
            if not section:
                continue
                
            # Split into sheet name and operations
            parts = [p.strip() for p in section.split('|', 1)]
            if not parts:
                continue
                
            sheet_name = parts[0].strip()
            if not sheet_name:
                continue
                
            if len(parts) == 1:  # No operations for this sheet
                continue
                
            # Process operation entries
            operation_entries = [e.strip() for e in parts[1].split('|') if e.strip()]
            
            for entry in operation_entries:
                if ':' not in entry:
                    continue
                    
                operation_type, params_str = entry.split(':', 1)
                operation_type = operation_type.strip()
                params_str = params_str.strip()
                
                # Parse parameters
                params = {}
                if params_str:
                    # Simple parameter parsing: key=value, key2=value2
                    for param in params_str.split(','):
                        if '=' in param:
                            key, value = param.split('=', 1)
                            key = key.strip()
                            value = value.strip().strip('"\'')
                            
                            # Type conversion
                            if value.isdigit():
                                value = int(value)
                            elif value.replace('.', '').isdigit():
                                value = float(value)
                            elif value.lower() in ['true', 'false']:
                                value = value.lower() == 'true'
                            elif '[' in value and ']' in value:
                                # Handle array parameters
                                value = value.strip('[]').split(',')
                                value = [v.strip().strip('\"\'') for v in value]
                            
                            params[key] = value
                
                operations.append({
                    'sheet_name': sheet_name,
                    'operation': operation_type,
                    'parameters': params
                })
        
        logger.debug(f"Parsed {len(operations)} row/column operations")
        return operations
        
    except Exception as e:
        logger.error(f"Error parsing row/column operations: {e}")
        return []


def _parse_sheet_operations(operations_dsl: str) -> List[Dict[str, Any]]:
    """Parse sheet operations DSL into structured operations list."""
    logger.debug("Parsing sheet operations DSL")
    
    operations = []
    
    try:
        # Split by | to get individual operations
        operation_entries = [e.strip() for e in operations_dsl.split('|') if e.strip()]
        
        for entry in operation_entries:
            if ':' not in entry:
                continue
                
            operation_type, params_str = entry.split(':', 1)
            operation_type = operation_type.strip()
            params_str = params_str.strip()
            
            # Parse parameters
            params = {}
            if params_str:
                # Simple parameter parsing: key=value, key2=value2
                for param in params_str.split(','):
                    if '=' in param:
                        key, value = param.split('=', 1)
                        key = key.strip()
                        value = value.strip().strip('"\'')
                        
                        # Type conversion
                        if value.isdigit():
                            value = int(value)
                        elif value.replace('.', '').isdigit():
                            value = float(value)
                        elif value.lower() in ['true', 'false']:
                            value = value.lower() == 'true'
                        elif value.startswith('[') and value.endswith(']'):
                            # Handle array parameters
                            value = value[1:-1].split(',')
                            value = [v.strip().strip('\"\'') for v in value]
                        
                        params[key] = value
            
            operations.append({
                'operation': operation_type,
                'parameters': params
            })
        
        logger.debug(f"Parsed {len(operations)} sheet operations")
        return operations
        
    except Exception as e:
        logger.error(f"Error parsing sheet operations: {e}")
        return []


def _execute_row_column_operations(excel_path: str, operations: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Execute row/column operations using xlwings."""
    logger.info(f"Executing {len(operations)} row/column operations")
    
    results = []
    
    try:
        import xlwings as xw
        
        # Open the workbook
        wb = xw.Book(excel_path)
        
        for operation in operations:
            try:
                sheet_name = operation['sheet_name']
                op_type = operation['operation']
                params = operation['parameters']
                
                # Get the worksheet
                if sheet_name not in [sheet.name for sheet in wb.sheets]:
                    results.append({
                        'operation': op_type,
                        'sheet': sheet_name,
                        'success': False,
                        'error': f'Sheet "{sheet_name}" not found'
                    })
                    continue
                    
                ws = wb.sheets[sheet_name]
                
                # Execute operation based on type
                if op_type == 'insert_row':
                    _execute_insert_row(ws, params)
                elif op_type == 'delete_row':
                    _execute_delete_row(ws, params)
                elif op_type == 'resize_row':
                    _execute_resize_row(ws, params)
                elif op_type == 'hide_row':
                    _execute_hide_row(ws, params)
                elif op_type == 'unhide_row':
                    _execute_unhide_row(ws, params)
                elif op_type == 'move_row':
                    _execute_move_row(ws, params)
                elif op_type == 'format_row':
                    _execute_format_row(ws, params)
                elif op_type == 'insert_column':
                    _execute_insert_column(ws, params)
                elif op_type == 'delete_column':
                    _execute_delete_column(ws, params)
                elif op_type == 'resize_column':
                    _execute_resize_column(ws, params)
                elif op_type == 'hide_column':
                    _execute_hide_column(ws, params)
                elif op_type == 'unhide_column':
                    _execute_unhide_column(ws, params)
                elif op_type == 'move_column':
                    _execute_move_column(ws, params)
                elif op_type == 'format_column':
                    _execute_format_column(ws, params)
                else:
                    results.append({
                        'operation': op_type,
                        'sheet': sheet_name,
                        'success': False,
                        'error': f'Unknown operation: {op_type}'
                    })
                    continue
                
                results.append({
                    'operation': op_type,
                    'sheet': sheet_name,
                    'parameters': params,
                    'success': True
                })
                
            except Exception as e:
                results.append({
                    'operation': operation.get('operation', 'unknown'),
                    'sheet': operation.get('sheet_name', 'unknown'),
                    'success': False,
                    'error': str(e)
                })
        
        # Save workbook
        wb.save()
        wb.close()
        
    except Exception as e:
        logger.error(f"Error executing row/column operations: {e}")
        results.append({
            'operation': 'general',
            'success': False,
            'error': str(e)
        })
    
    return results


def _execute_sheet_operations(excel_path: str, operations: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Execute sheet operations using xlwings."""
    logger.info(f"Executing {len(operations)} sheet operations")
    
    results = []
    
    try:
        import xlwings as xw
        
        # Open the workbook
        wb = xw.Book(excel_path)
        
        for operation in operations:
            try:
                op_type = operation['operation']
                params = operation['parameters']
                
                # Execute operation based on type
                if op_type == 'add_sheet':
                    _execute_add_sheet(wb, params)
                elif op_type == 'delete_sheet':
                    _execute_delete_sheet(wb, params)
                elif op_type == 'rename_sheet':
                    _execute_rename_sheet(wb, params)
                elif op_type == 'copy_sheet':
                    _execute_copy_sheet(wb, params)
                elif op_type == 'move_sheet':
                    _execute_move_sheet(wb, params)
                elif op_type == 'hide_sheet':
                    _execute_hide_sheet(wb, params)
                elif op_type == 'unhide_sheet':
                    _execute_unhide_sheet(wb, params)
                elif op_type == 'set_tab_color':
                    _execute_set_tab_color(wb, params)
                elif op_type == 'activate_sheet':
                    _execute_activate_sheet(wb, params)
                elif op_type == 'duplicate_sheet':
                    _execute_duplicate_sheet(wb, params)
                elif op_type == 'protect_sheet':
                    _execute_protect_sheet(wb, params)
                elif op_type == 'unprotect_sheet':
                    _execute_unprotect_sheet(wb, params)
                else:
                    results.append({
                        'operation': op_type,
                        'success': False,
                        'error': f'Unknown operation: {op_type}'
                    })
                    continue
                
                results.append({
                    'operation': op_type,
                    'parameters': params,
                    'success': True
                })
                
            except Exception as e:
                results.append({
                    'operation': operation.get('operation', 'unknown'),
                    'success': False,
                    'error': str(e)
                })
        
        # Save workbook
        wb.save()
        wb.close()
        
    except Exception as e:
        logger.error(f"Error executing sheet operations: {e}")
        results.append({
            'operation': 'general',
            'success': False,
            'error': str(e)
        })
    
    return results


# Row operation implementations
def _execute_insert_row(ws, params):
    """Insert rows at specified position."""
    row_number = params.get('row_number', 1)
    count = params.get('count', 1)
    copy_format_from = params.get('copy_format_from')
    
    # Insert rows
    for i in range(count):
        ws.api.Rows(row_number).Insert()
    
    # Copy format if specified
    if copy_format_from:
        source_range = ws.api.Rows(copy_format_from)
        target_range = ws.api.Rows(f"{row_number}:{row_number + count - 1}")
        source_range.Copy()
        target_range.PasteSpecial(-4122)  # xlPasteFormats


def _execute_delete_row(ws, params):
    """Delete rows at specified position."""
    row_number = params.get('row_number', 1)
    count = params.get('count', 1)
    
    # Delete rows
    ws.api.Rows(f"{row_number}:{row_number + count - 1}").Delete()


def _execute_resize_row(ws, params):
    """Resize row height."""
    row_number = params.get('row_number', 1)
    height = params.get('height', 15)
    
    ws.api.Rows(row_number).RowHeight = height


def _execute_hide_row(ws, params):
    """Hide rows."""
    row_range = params.get('row_number', '1')
    ws.api.Rows(row_range).Hidden = True


def _execute_unhide_row(ws, params):
    """Unhide rows.""" 
    row_range = params.get('row_number', '1')
    ws.api.Rows(row_range).Hidden = False


def _execute_move_row(ws, params):
    """Move rows (copy and delete)."""
    from_row = params.get('from_row', 1)
    to_row = params.get('to_row', 1)
    count = params.get('count', 1)
    
    # Copy rows
    source_range = ws.api.Rows(f"{from_row}:{from_row + count - 1}")
    source_range.Copy()
    
    # Insert rows at destination
    for i in range(count):
        ws.api.Rows(to_row).Insert()
    
    # Paste to destination
    target_range = ws.api.Rows(f"{to_row}:{to_row + count - 1}")
    target_range.PasteSpecial(-4104)  # xlPasteAll
    
    # Delete original rows (adjust for inserted rows)
    if from_row > to_row:
        from_row += count
    ws.api.Rows(f"{from_row}:{from_row + count - 1}").Delete()


def _execute_format_row(ws, params):
    """Format entire row."""
    row_number = params.get('row_number', 1)
    properties = params.get('properties', '')
    
    # Apply formatting using existing cell operations DSL
    dsl = f"sheet_name: {ws.name} | 1:{row_number}, [no_change], {properties}"
    parsed = parse_markdown_formulas(dsl)
    if parsed:
        writer = ExcelWriter()
        writer.write_to_existing(parsed, ws.parent.fullname)


# Column operation implementations
def _execute_insert_column(ws, params):
    """Insert columns at specified position."""
    column = params.get('column', 'A')
    count = params.get('count', 1)
    copy_format_from = params.get('copy_format_from')
    
    # Insert columns
    for i in range(count):
        ws.api.Columns(column).Insert()
    
    # Copy format if specified  
    if copy_format_from:
        source_range = ws.api.Columns(copy_format_from)
        target_range = ws.api.Columns(column)
        source_range.Copy()
        target_range.PasteSpecial(-4122)  # xlPasteFormats


def _execute_delete_column(ws, params):
    """Delete columns at specified position."""
    column = params.get('column', 'A')
    count = params.get('count', 1)
    
    # Calculate column range for deletion
    start_col = column
    end_col = chr(ord(start_col) + count - 1)
    
    ws.api.Columns(f"{start_col}:{end_col}").Delete()


def _execute_resize_column(ws, params):
    """Resize column width."""
    column = params.get('column', 'A')
    width = params.get('width', 8.5)
    
    ws.api.Columns(column).ColumnWidth = width


def _execute_hide_column(ws, params):
    """Hide columns."""
    column_range = params.get('column', 'A')
    ws.api.Columns(column_range).Hidden = True


def _execute_unhide_column(ws, params):
    """Unhide columns."""
    column_range = params.get('column', 'A') 
    ws.api.Columns(column_range).Hidden = False


def _execute_move_column(ws, params):
    """Move columns (copy and delete)."""
    from_column = params.get('from_column', 'A')
    to_column = params.get('to_column', 'B') 
    count = params.get('count', 1)
    
    # Calculate column ranges
    from_end_col = chr(ord(from_column) + count - 1)
    
    # Copy columns
    source_range = ws.api.Columns(f"{from_column}:{from_end_col}")
    source_range.Copy()
    
    # Insert columns at destination
    for i in range(count):
        ws.api.Columns(to_column).Insert()
    
    # Paste to destination
    target_range = ws.api.Columns(f"{to_column}:{chr(ord(to_column) + count - 1)}")
    target_range.PasteSpecial(-4104)  # xlPasteAll
    
    # Delete original columns
    ws.api.Columns(f"{from_column}:{from_end_col}").Delete()


def _execute_format_column(ws, params):
    """Format entire column."""
    column = params.get('column', 'A')
    properties = params.get('properties', '')
    
    # Apply formatting using existing cell operations DSL
    dsl = f"sheet_name: {ws.name} | {column}:{column}, [no_change], {properties}"
    parsed = parse_markdown_formulas(dsl)
    if parsed:
        writer = ExcelWriter()
        writer.write_to_existing(parsed, ws.parent.fullname)


# Sheet operation implementations  
def _execute_add_sheet(wb, params):
    """Add new worksheet."""
    name = params.get('name', 'New Sheet')
    position = params.get('position', -1)
    template_sheet = params.get('template_sheet')
    
    if template_sheet and template_sheet in [sheet.name for sheet in wb.sheets]:
        # Copy from template
        template = wb.sheets[template_sheet]
        new_sheet = template.copy(name=name)
    else:
        # Create new sheet
        new_sheet = wb.sheets.add(name=name)
    
    # Move to position if specified
    if position > 0 and position <= len(wb.sheets):
        new_sheet.api.Move(Before=wb.sheets[position - 1].api)


def _execute_delete_sheet(wb, params):
    """Delete worksheet."""
    name = params.get('name', '')
    if name and name in [sheet.name for sheet in wb.sheets]:
        wb.sheets[name].delete()


def _execute_rename_sheet(wb, params):
    """Rename worksheet."""
    old_name = params.get('old_name', '')
    new_name = params.get('new_name', '')
    
    if old_name and new_name and old_name in [sheet.name for sheet in wb.sheets]:
        wb.sheets[old_name].name = new_name


def _execute_copy_sheet(wb, params):
    """Copy worksheet."""
    source = params.get('source', '')
    name = params.get('name', 'Copy of Sheet')
    position = params.get('position', -1)
    
    if source and source in [sheet.name for sheet in wb.sheets]:
        new_sheet = wb.sheets[source].copy(name=name)
        
        # Move to position if specified
        if position > 0 and position <= len(wb.sheets):
            new_sheet.api.Move(Before=wb.sheets[position - 1].api)


def _execute_move_sheet(wb, params):
    """Move worksheet to new position."""
    name = params.get('name', '')
    position = params.get('position', 1)
    
    if name and name in [sheet.name for sheet in wb.sheets]:
        sheet = wb.sheets[name]
        if position > 0 and position <= len(wb.sheets):
            sheet.api.Move(Before=wb.sheets[position - 1].api)


def _execute_hide_sheet(wb, params):
    """Hide worksheet."""
    name = params.get('name', '')
    if name and name in [sheet.name for sheet in wb.sheets]:
        wb.sheets[name].api.Visible = False


def _execute_unhide_sheet(wb, params):
    """Unhide worksheet."""
    name = params.get('name', '')
    if name and name in [sheet.name for sheet in wb.sheets]:
        wb.sheets[name].api.Visible = True


def _execute_set_tab_color(wb, params):
    """Set worksheet tab color."""
    name = params.get('name', '')
    color = params.get('color', '#FFFFFF')
    
    if name and name in [sheet.name for sheet in wb.sheets]:
        # Convert hex color to RGB
        if color.startswith('#'):
            rgb_tuple = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
            rgb_int = rgb_tuple[0] + (rgb_tuple[1] << 8) + (rgb_tuple[2] << 16)
            wb.sheets[name].api.Tab.Color = rgb_int


def _execute_activate_sheet(wb, params):
    """Activate worksheet."""
    name = params.get('name', '')
    if name and name in [sheet.name for sheet in wb.sheets]:
        wb.sheets[name].activate()


def _execute_duplicate_sheet(wb, params):
    """Duplicate sheet multiple times."""
    source = params.get('source', '')
    names = params.get('names', [])
    positions = params.get('positions', [])
    
    if source and source in [sheet.name for sheet in wb.sheets] and names:
        for i, name in enumerate(names):
            new_sheet = wb.sheets[source].copy(name=name)
            
            # Move to position if specified
            if i < len(positions) and positions[i] > 0:
                position = positions[i]
                if position <= len(wb.sheets):
                    new_sheet.api.Move(Before=wb.sheets[position - 1].api)


def _execute_protect_sheet(wb, params):
    """Protect worksheet."""
    name = params.get('name', '')
    password = params.get('password', '')
    
    if name and name in [sheet.name for sheet in wb.sheets]:
        wb.sheets[name].api.Protect(Password=password)


def _execute_unprotect_sheet(wb, params):
    """Unprotect worksheet."""
    name = params.get('name', '')
    password = params.get('password', '')
    
    if name and name in [sheet.name for sheet in wb.sheets]:
        wb.sheets[name].api.Unprotect(Password=password)


def _get_default_sheet_name(excel_path: str) -> str:
    """
    Get the default sheet name from an Excel file.
    
    Args:
        excel_path: Path to the Excel file
        
    Returns:
        Name of the first sheet or 'Sheet1' as fallback
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_name = wb.sheetnames[0] if wb.sheetnames else 'Sheet1'
        wb.close()
        return sheet_name
    except Exception as e:
        logger.warning(f"Could not get sheet name from {excel_path}: {e}")
        return 'Sheet1'


def _convert_simple_dsl_to_markdown(simple_dsl: str, sheet_name: str) -> str:
    """
    Convert simple DSL format to markdown format expected by parser.
    
    Args:
        simple_dsl: Simple DSL like 'A1 = "Hello"\nB1 = 123'
        sheet_name: Sheet name to use
        
    Returns:
        Markdown format like 'sheet_name: Sheet1 | A1, "Hello" | B1, 123'
    """
    logger.debug(f"Converting simple DSL to markdown format for sheet: {sheet_name}")
    
    try:
        lines = [line.strip() for line in simple_dsl.split('\n') if line.strip()]
        cell_entries = []
        
        for line in lines:
            # Handle cell value assignments: A1 = "value" or A1 = 123 or A1 = =SUM(B1:C1)
            if '=' in line and not line.strip().startswith('#'):
                parts = line.split('=', 1)
                if len(parts) == 2:
                    cell_ref = parts[0].strip()
                    value = parts[1].strip()
                    
                    # Format the value properly
                    if value.startswith('='):
                        # Formula - wrap in quotes
                        formatted_value = f'"{value}"'
                    elif value.startswith('"') and value.endswith('"'):
                        # Already quoted string - keep as is
                        formatted_value = value
                    elif value.replace('.', '').replace('-', '').isdigit():
                        # Numeric value - no quotes needed
                        formatted_value = value
                    else:
                        # String value - add quotes if not present
                        cleaned_value = value.strip('"\'')
                        formatted_value = f'"{cleaned_value}"'
                    
                    cell_entries.append(f"{cell_ref}, {formatted_value}")
            
            # Handle formatting commands: A1 font bold, B1 background yellow
            elif ' font ' in line or ' background ' in line or ' color ' in line:
                # Parse formatting commands
                parts = line.split(' ', 1)
                if len(parts) == 2:
                    cell_ref = parts[0].strip()
                    format_command = parts[1].strip()
                    
                    # Convert common formatting commands to DSL properties
                    properties = []
                    if 'font bold' in format_command:
                        properties.append('b=true')
                    if 'background yellow' in format_command:
                        properties.append('fill="#FFFF00"')
                    if 'background red' in format_command:
                        properties.append('fill="#FF0000"')
                    if 'font color red' in format_command:
                        properties.append('font="#FF0000"')
                    
                    if properties:
                        cell_entries.append(f"{cell_ref}, [no_change], {', '.join(properties)}")
        
        if cell_entries:
            markdown_format = f"sheet_name: {sheet_name} | " + " | ".join(cell_entries)
            logger.debug(f"Converted to markdown: {markdown_format[:100]}...")
            return markdown_format
        else:
            return f"sheet_name: {sheet_name}"
            
    except Exception as e:
        logger.error(f"Error converting simple DSL to markdown: {e}")
        return f"sheet_name: {sheet_name}"


# Missing function implementation
def parse_conditional_formatting(cf_content: str) -> Dict[str, Any]:
    """
    Parse conditional formatting content into a structured format.
    
    Args:
        cf_content: Conditional formatting string like "type:cell_value,operator:greater,value1:100,background_color:#FF0000"
        
    Returns:
        Dictionary with conditional formatting parameters
    """
    cf_data = {}
    
    try:
        # Split by comma and parse key:value pairs
        parts = cf_content.split(',')
        for part in parts:
            if ':' in part:
                key, value = part.split(':', 1)
                key = key.strip()
                value = value.strip()
                
                # Type conversions
                if key in ['value1', 'value2']:
                    try:
                        value = float(value)
                    except ValueError:
                        pass  # Keep as string
                elif key in ['type', 'operator', 'background_color', 'font_color']:
                    pass  # Keep as string
                
                cf_data[key] = value
        
        logger.debug(f"Parsed conditional formatting: {cf_data}")
        return cf_data
        
    except Exception as e:
        logger.error(f"Error parsing conditional formatting: {e}")
        return {}


if __name__ == "__main__":
    # Example usage for testing
    sample_operations = {
        "cell": 'sheet_name: Sheet1 | A1, [=SUM(B1:B10)], b=true, sz="14", fill="#FFFF00" | B1, [100], ha="center"',
        "row_column": 'sheet_name: Data | insert_row: row_number=5, count=2 | resize_column: column=A, width=20',
        "sheet": 'add_sheet: name="Analysis", position=2 | set_tab_color: name="Analysis", color="#00FF00"'
    }
    
    print("Excel Operations Tools DSL Examples:")
    print("=" * 50)
    for op_type, dsl in sample_operations.items():
        print(f"\n{op_type.upper()} OPERATIONS DSL:")
        print(dsl)
        print("-" * 30)
