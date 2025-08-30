#!/usr/bin/env python3
"""
Excel File Creation Tools for MCP

This module provides MCP tools for creating new Excel files at specific file paths
with initial data and formatting.
"""

import logging
import json
import os
from typing import Dict, Any, List, Optional
from pathlib import Path

# Setup logging
logger = logging.getLogger(__name__)

def create_excel_file(
    excel_path: str,
    initial_data: Optional[str] = None,
    sheets: Optional[List[str]] = None,
    overwrite: bool = False
) -> str:
    """
    Create a new Excel file at the specified path with optional initial data.
    
    Args:
        excel_path: Path where the Excel file should be created
        initial_data: Optional natural language commands to populate the file
        sheets: Optional list of sheet names to create (default: ['Sheet1'])
        overwrite: Whether to overwrite existing files (default: False)
        
    Returns:
        JSON string with creation result
        
    Examples:
        # Create empty file
        create_excel_file("C:/data/report.xlsx")
        
        # Create with initial data
        create_excel_file("C:/data/report.xlsx", '''
            A1 = "Company Report"
            A1 bold
            A1 fill #0000FF
            A1 font #FFFFFF
            A3 = "Sales Data:"
            B4 = "Q1"
            C4 = 125000
        ''')
        
        # Create with multiple sheets
        create_excel_file("C:/data/workbook.xlsx", sheets=["Summary", "Data", "Charts"])
    """
    logger.info(f"Creating Excel file at {excel_path}")
    
    try:
        # Validate and prepare path
        excel_file = Path(excel_path)
        excel_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Check if file exists and overwrite policy
        if excel_file.exists() and not overwrite:
            return json.dumps({
                "error": f"File already exists: {excel_path}. Use overwrite=True to replace it."
            })
        
        # Import required modules
        import openpyxl
        from .excel_operations_optimized import excel_edit
        
        # Create basic Excel file
        workbook = openpyxl.Workbook()
        
        # Set up sheets
        if sheets:
            # Remove default sheet if we're creating custom sheets
            if workbook.active.title == 'Sheet':
                workbook.remove(workbook.active)
            
            # Create requested sheets
            for i, sheet_name in enumerate(sheets):
                if i == 0:
                    # First sheet
                    ws = workbook.create_sheet(title=sheet_name)
                    workbook.active = ws
                else:
                    workbook.create_sheet(title=sheet_name)
        else:
            # Use default sheet name
            workbook.active.title = 'Sheet1'
        
        # Save the basic file
        workbook.save(excel_path)
        workbook.close()
        
        logger.info(f"Created basic Excel file at {excel_path}")
        
        result = {
            "success": True,
            "excel_file": str(excel_file.resolve()),
            "operation": "create_excel_file",
            "sheets_created": sheets or ['Sheet1'],
            "file_size": excel_file.stat().st_size
        }
        
        # If initial data provided, populate it
        if initial_data and initial_data.strip():
            logger.info("Populating initial data...")
            
            try:
                edit_result = excel_edit(str(excel_file), initial_data.strip())
                
                if isinstance(edit_result, dict) and edit_result.get('success'):
                    result['initial_data'] = {
                        "success": True,
                        "cells_populated": edit_result.get('results', [])
                    }
                else:
                    result['initial_data'] = {
                        "success": False,
                        "error": str(edit_result)
                    }
            except Exception as e:
                logger.error(f"Error populating initial data: {e}")
                result['initial_data'] = {
                    "success": False,
                    "error": f"Failed to populate initial data: {str(e)}"
                }
        
        # Update file size after data population
        result['final_file_size'] = excel_file.stat().st_size
        
        logger.info(f"Excel file creation completed successfully at {excel_path}")
        return json.dumps(result, indent=2)
        
    except Exception as e:
        error_msg = f"Error creating Excel file: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return json.dumps({"error": error_msg})


def create_excel_from_template(
    excel_path: str,
    template_data: Dict[str, Any],
    overwrite: bool = False
) -> str:
    """
    Create a new Excel file from a structured template.
    
    Args:
        excel_path: Path where the Excel file should be created
        template_data: Dictionary defining the template structure
        overwrite: Whether to overwrite existing files
        
    Template Format:
        {
            "sheets": {
                "Summary": {
                    "data": [
                        {"cell": "A1", "value": "Report Title", "bold": True, "fill": "#0000FF"},
                        {"cell": "A3", "value": "Date:", "bold": True},
                        {"cell": "B3", "formula": "=TODAY()"}
                    ]
                },
                "Data": {
                    "headers": ["Name", "Value", "Status"],
                    "data": [
                        ["Item 1", 100, "Active"],
                        ["Item 2", 200, "Pending"]
                    ]
                }
            }
        }
        
    Returns:
        JSON string with creation result
    """
    logger.info(f"Creating Excel file from template at {excel_path}")
    
    try:
        # Validate and prepare path
        excel_file = Path(excel_path)
        excel_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Check if file exists and overwrite policy
        if excel_file.exists() and not overwrite:
            return json.dumps({
                "error": f"File already exists: {excel_path}. Use overwrite=True to replace it."
            })
        
        # Import required modules
        from .writer.excel_writer import ExcelWriter
        
        # Convert template to Excel writer format
        excel_data = {}
        
        for sheet_name, sheet_config in template_data.get("sheets", {}).items():
            excel_data[sheet_name] = []
            
            # Handle direct cell data
            if "data" in sheet_config:
                for cell_config in sheet_config["data"]:
                    excel_data[sheet_name].append(cell_config)
            
            # Handle tabular data with headers
            if "headers" in sheet_config and "data" in sheet_config:
                headers = sheet_config["headers"]
                data_rows = sheet_config["data"]
                
                # Add headers (row 1)
                for col_idx, header in enumerate(headers):
                    excel_data[sheet_name].append({
                        "cell": f"{chr(65 + col_idx)}1",
                        "formula": header,
                        "bold": True
                    })
                
                # Add data rows
                for row_idx, data_row in enumerate(data_rows, 2):  # Start from row 2
                    for col_idx, cell_value in enumerate(data_row):
                        excel_data[sheet_name].append({
                            "cell": f"{chr(65 + col_idx)}{row_idx}",
                            "formula": str(cell_value)
                        })
        
        # Create the Excel file
        writer = ExcelWriter()
        success, updated_cells = writer.create_new_excel(
            data=excel_data,
            output_filepath=str(excel_file)
        )
        
        if success:
            result = {
                "success": True,
                "excel_file": str(excel_file.resolve()),
                "operation": "create_excel_from_template",
                "sheets_created": list(template_data.get("sheets", {}).keys()),
                "updated_cells": updated_cells,
                "file_size": excel_file.stat().st_size
            }
            
            logger.info(f"Excel file created successfully from template at {excel_path}")
            return json.dumps(result, indent=2)
        else:
            return json.dumps({
                "error": "Failed to create Excel file from template"
            })
        
    except Exception as e:
        error_msg = f"Error creating Excel file from template: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return json.dumps({"error": error_msg})


# Export functions
__all__ = ['create_excel_file', 'create_excel_from_template']
