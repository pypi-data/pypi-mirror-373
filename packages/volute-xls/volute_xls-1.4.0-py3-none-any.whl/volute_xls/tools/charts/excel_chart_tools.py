"""
Excel Chart Tools for MCP

This module provides comprehensive MCP tools for creating and modifying Excel charts.
Designed for AI agents to easily manipulate chart properties through structured commands.
"""

import logging
import json
import os
import re
from typing import Dict, Any, List, Optional
from pathlib import Path

# Setup logging
logger = logging.getLogger(__name__)

# Import the writer for Excel operations
try:
    from ...writer.excel_writer import ExcelWriter
except ImportError:
    try:
        # Try alternative import path
        import sys
        from pathlib import Path
        # Add parent directory to path
        parent_dir = Path(__file__).parent.parent.parent
        sys.path.insert(0, str(parent_dir))
        from writer.excel_writer import ExcelWriter
    except ImportError:
        logger.error("Could not import ExcelWriter")
        ExcelWriter = None


def create_excel_chart(
    excel_path: str,
    chart_name: str,
    chart_type: str,
    x_axis_range: str,
    data_series: str,
    sheet_name: str = None
) -> str:
    """
    Create a new Excel chart with data series and basic settings.
    
    This tool creates charts with essential properties. Use chart_style_excel_chart 
    and chart_data_excel_chart for advanced formatting and data modifications.
    
    Args:
        excel_path: Path to the Excel file
        chart_name: Unique name for the chart
        chart_type: Type of chart (line, column_clustered, bar_clustered, pie, xy_scatter, etc.)
        x_axis_range: Cell range for X-axis data (e.g., "A1:A10")
        data_series: Data series in format "series_1=B1:B10,series_2=C1:C10" or "series_1=B1:B10"
        sheet_name: Sheet name (optional, uses first sheet if not specified)
        
    Returns:
        JSON string with creation result
        
    Examples:
        # Simple line chart
        create_excel_chart(
            excel_path="C:/data/sales.xlsx",
            chart_name="monthly_sales",
            chart_type="line",
            x_axis_range="A2:A13",
            data_series="series_1=B2:B13",
            sheet_name="Dashboard"
        )
        
        # Multi-series column chart
        create_excel_chart(
            excel_path="C:/reports/comparison.xlsx",
            chart_name="quarterly_comparison",
            chart_type="column_clustered",
            x_axis_range="A1:A4",
            data_series="series_1=B1:B4,series_2=C1:C4,series_3=D1:D4"
        )
        
    Supported Chart Types:
        line, line_markers, column_clustered, bar_clustered, pie, xy_scatter,
        area, area_stacked, 3d_column, 3d_pie, bubble, doughnut, radar
    """
    try:
        if not ExcelWriter:
            return json.dumps({"error": "Excel writer not available"})
            
        # Validate inputs
        if not all([excel_path, chart_name, chart_type, x_axis_range, data_series]):
            return json.dumps({"error": "Missing required parameters"})
            
        excel_file = Path(excel_path)
        if not excel_file.exists():
            return json.dumps({"error": f"Excel file not found: {excel_path}"})
            
        # Parse data series
        series_data = _parse_data_series(data_series)
        if not series_data:
            return json.dumps({"error": "Invalid data_series format"})
            
        # Build chart data structure
        chart_data = {
            'chart': chart_name,
            'type': chart_type,
            'x_axis': x_axis_range,
            'height': 300,  # Default height
            'left': 50,     # Default position
            'title': chart_name.replace('_', ' ').title(),  # Default title
            'legend': 'true'
        }
        
        # Add series data
        chart_data.update(series_data)
        
        # Get default sheet name if not provided
        if not sheet_name:
            sheet_name = _get_first_sheet_name(excel_path)
            
        # Write chart to Excel using ExcelWriter
        writer = ExcelWriter()
        
        data_structure = {
            sheet_name: [chart_data]
        }
        
        success, results = writer.write_to_existing(data_structure, excel_path)
        
        if success and results:
            return json.dumps({
                "success": True,
                "chart_name": chart_name,
                "chart_type": chart_type,
                "sheet_name": sheet_name,
                "message": f"Chart '{chart_name}' created successfully",
                "results": results
            }, indent=2)
        else:
            return json.dumps({"error": "Failed to create chart"})
            
    except Exception as e:
        logger.error(f"Error creating chart: {e}", exc_info=True)
        return json.dumps({"error": f"Chart creation failed: {str(e)}"})


def style_excel_chart(
    excel_path: str,
    chart_name: str,
    formatting_properties: str,
    sheet_name: str = None
) -> str:
    """
    Apply advanced styling to an existing Excel chart.
    
    This tool handles chart appearance including titles, axes, gridlines, colors,
    plot areas, and chart areas. Use specific property syntax for precise control.
    
    Args:
        excel_path: Path to the Excel file
        chart_name: Name of the chart to style
        formatting_properties: Chart formatting properties in structured syntax
        sheet_name: Sheet name (optional)
        
    Returns:
        JSON string with styling result
        
    Formatting Properties Syntax:
        title="Chart Title"|title_pos="above"|title_font=[Arial,14,true,false,#000000]|
        legend="true"|legend_pos="bottom"|
        x_title="X Axis"|x_title_font=[Arial,10,false,false,#666666]|x_grid="true"|x_grid_color="#E0E0E0"|
        y_title="Y Axis"|y_title_font=[Arial,10,false,false,#666666]|y_grid="true"|y_grid_color="#E0E0E0"|
        s1_color="#FF0000"|s2_color="#00FF00"|s3_color="#0000FF"|
        data_labels="true"|data_label_position="outside_end"|data_label_color="#FF0000"|
        data_label_font=[Arial,10,true,false,#000000]|s1_data_labels="true"|s2_data_labels="false"|
        plot_fill="#FAFAFA"|plot_border="true"|plot_border_color="#CCCCCC"|
        chart_fill="#FFFFFF"|chart_border="true"|chart_border_color="#333333"
        
    Examples:
        # Style chart with title and colors
        style_excel_chart(
            excel_path="C:/data/sales.xlsx",
            chart_name="monthly_sales",
            formatting_properties="title=Monthly Sales Report|title_font=[Arial,16,true,false,#2F4F4F]|s1_color=#FF6B6B|legend=true|legend_pos=bottom"
        )
        
        # Advanced axis and grid styling
        style_excel_chart(
            excel_path="C:/reports/analysis.xlsx",
            chart_name="trend_analysis",
            formatting_properties="x_title=Time Period|y_title=Sales ($)|x_grid=true|y_grid=true|x_grid_color=#E8E8E8|y_grid_color=#E8E8E8|plot_fill=#F8F8F8"
        )
        
        # Data labels - Global settings
        style_excel_chart(
            excel_path="C:/data/sales.xlsx",
            chart_name="monthly_sales",
            formatting_properties="data_labels=true|data_label_position=outside_end|data_label_color=#0066CC|data_label_font=[Arial,10,true,false,#0066CC]"
        )
        
        # Data labels - Series-specific control
        style_excel_chart(
            excel_path="C:/reports/comparison.xlsx",
            chart_name="multi_series_chart",
            formatting_properties="s1_data_labels=true|s2_data_labels=false|s3_data_labels=true|data_label_position=center|data_label_color=#FF6600"
        )
    """
    try:
        if not ExcelWriter:
            return json.dumps({"error": "Excel writer not available"})
            
        # Validate inputs
        if not all([excel_path, chart_name, formatting_properties]):
            return json.dumps({"error": "Missing required parameters"})
            
        excel_file = Path(excel_path)
        if not excel_file.exists():
            return json.dumps({"error": f"Excel file not found: {excel_path}"})
            
        # Parse formatting properties
        chart_properties = _parse_chart_properties(formatting_properties)
        if not chart_properties:
            return json.dumps({"error": "Invalid formatting_properties format"})
            
        # Build chart update structure
        chart_data = {
            'chart': chart_name
        }
        chart_data.update(chart_properties)
        
        # Get default sheet name if not provided
        if not sheet_name:
            sheet_name = _get_first_sheet_name(excel_path)
            
        # Apply styling using ExcelWriter
        writer = ExcelWriter()
        
        data_structure = {
            sheet_name: [chart_data]
        }
        
        success, results = writer.write_to_existing(data_structure, excel_path)
        
        if success and results:
            return json.dumps({
                "success": True,
                "chart_name": chart_name,
                "sheet_name": sheet_name,
                "message": f"Chart '{chart_name}' styled successfully",
                "properties_applied": list(chart_properties.keys()),
                "results": results
            }, indent=2)
        else:
            return json.dumps({"error": "Failed to style chart"})
            
    except Exception as e:
        logger.error(f"Error styling chart: {e}", exc_info=True)
        return json.dumps({"error": f"Chart styling failed: {str(e)}"})


def update_chart_data(
    excel_path: str,
    chart_name: str,
    x_axis_range: str = None,
    data_series: str = None,
    series_names: str = None,
    sheet_name: str = None
) -> str:
    """
    Update data ranges and series names for an existing Excel chart.
    
    This tool modifies the data source and series labels of charts without
    affecting the visual styling. Use for dynamic data updates.
    
    Args:
        excel_path: Path to the Excel file
        chart_name: Name of the chart to update
        x_axis_range: New X-axis data range (optional, e.g., "A1:A12")
        data_series: New data series (optional, e.g., "series_1=B1:B12,series_2=C1:C12")
        series_names: Series name mappings (optional, e.g., "series_1=B1,series_2=C1")
        sheet_name: Sheet name (optional)
        
    Returns:
        JSON string with update result
        
    Examples:
        # Update data ranges
        update_chart_data(
            excel_path="C:/data/sales.xlsx",
            chart_name="monthly_sales",
            x_axis_range="A1:A24",
            data_series="series_1=B1:B24,series_2=C1:C24"
        )
        
        # Update series names only
        update_chart_data(
            excel_path="C:/reports/comparison.xlsx",
            chart_name="quarterly_comparison",
            series_names="series_1=B1,series_2=C1"
        )
        
        # Add new data series
        update_chart_data(
            excel_path="C:/analysis/trends.xlsx",
            chart_name="trend_chart",
            data_series="series_1=B2:B13,series_2=C2:C13,series_3=D2:D13",
            series_names="series_1=B1,series_2=C1,series_3=D1"
        )
    """
    try:
        if not ExcelWriter:
            return json.dumps({"error": "Excel writer not available"})
            
        # Validate inputs
        if not all([excel_path, chart_name]):
            return json.dumps({"error": "Missing required parameters"})
            
        if not any([x_axis_range, data_series, series_names]):
            return json.dumps({"error": "At least one data parameter must be provided"})
            
        excel_file = Path(excel_path)
        if not excel_file.exists():
            return json.dumps({"error": f"Excel file not found: {excel_path}"})
            
        # Build chart update structure
        chart_data = {
            'chart': chart_name
        }
        
        # Add X-axis range if provided
        if x_axis_range:
            chart_data['x_axis'] = x_axis_range
            
        # Parse and add data series if provided
        if data_series:
            series_data = _parse_data_series(data_series)
            if series_data:
                chart_data.update(series_data)
            else:
                return json.dumps({"error": "Invalid data_series format"})
                
        # Parse and add series names if provided
        if series_names:
            name_data = _parse_series_names(series_names)
            if name_data:
                chart_data.update(name_data)
            else:
                return json.dumps({"error": "Invalid series_names format"})
                
        # Get default sheet name if not provided
        if not sheet_name:
            sheet_name = _get_first_sheet_name(excel_path)
            
        # Update chart data using ExcelWriter
        writer = ExcelWriter()
        
        data_structure = {
            sheet_name: [chart_data]
        }
        
        success, results = writer.write_to_existing(data_structure, excel_path)
        
        if success and results:
            return json.dumps({
                "success": True,
                "chart_name": chart_name,
                "sheet_name": sheet_name,
                "message": f"Chart '{chart_name}' data updated successfully",
                "updates_applied": {
                    "x_axis_updated": bool(x_axis_range),
                    "data_series_updated": bool(data_series),
                    "series_names_updated": bool(series_names)
                },
                "results": results
            }, indent=2)
        else:
            return json.dumps({"error": "Failed to update chart data"})
            
    except Exception as e:
        logger.error(f"Error updating chart data: {e}", exc_info=True)
        return json.dumps({"error": f"Chart data update failed: {str(e)}"})


# Helper functions for parsing input formats

def _parse_data_series(data_series: str) -> Dict[str, str]:
    """
    Parse data series string into chart data format.
    
    Args:
        data_series: String like "series_1=B1:B10,series_2=C1:C10"
        
    Returns:
        Dictionary with series data
    """
    try:
        series_data = {}
        
        # Split by comma to get individual series
        series_parts = [part.strip() for part in data_series.split(',')]
        
        for part in series_parts:
            if '=' not in part:
                continue
                
            series_key, range_value = part.split('=', 1)
            series_key = series_key.strip()
            range_value = range_value.strip()
            
            # Validate series key format
            if not series_key.startswith('series_'):
                continue
                
            # Validate range format
            if not _is_valid_range(range_value):
                continue
                
            series_data[series_key] = range_value
            
        return series_data
        
    except Exception as e:
        logger.warning(f"Error parsing data series: {e}")
        return {}


def _parse_series_names(series_names: str) -> Dict[str, str]:
    """
    Parse series names string into chart name format.
    
    Args:
        series_names: String like "series_1=B1,series_2=C1"
        
    Returns:
        Dictionary with series name mappings
    """
    try:
        name_data = {}
        
        # Split by comma to get individual name mappings
        name_parts = [part.strip() for part in series_names.split(',')]
        
        for part in name_parts:
            if '=' not in part:
                continue
                
            series_key, name_cell = part.split('=', 1)
            series_key = series_key.strip()
            name_cell = name_cell.strip()
            
            # Validate series key format and convert to name format
            if series_key.startswith('series_'):
                name_key = f"{series_key}_name"
                
                # Validate cell reference format
                if _is_valid_cell_ref(name_cell):
                    name_data[name_key] = name_cell
                    
        return name_data
        
    except Exception as e:
        logger.warning(f"Error parsing series names: {e}")
        return {}


def _parse_chart_properties(formatting_properties: str) -> Dict[str, Any]:
    """
    Parse formatting properties string into chart properties.
    
    Args:
        formatting_properties: String with pipe-separated properties
        
    Returns:
        Dictionary with chart properties
    """
    try:
        properties = {}
        
        # Split by pipe to get individual properties
        prop_parts = [part.strip() for part in formatting_properties.split('|')]
        
        for part in prop_parts:
            if '=' not in part:
                continue
                
            key, value = part.split('=', 1)
            key = key.strip()
            value = value.strip()
            
            # Remove quotes if present
            if (value.startswith('"') and value.endswith('"')) or \
               (value.startswith("'") and value.endswith("'")):
                value = value[1:-1]
                
            # Handle boolean values
            if value.lower() in ['true', 'false']:
                value = value.lower() == 'true'
            # Handle numeric values
            elif value.isdigit():
                value = int(value)
            # Handle font arrays (keep as string for processing by writer)
            elif value.startswith('[') and value.endswith(']'):
                # Keep font arrays as strings for writer to parse
                pass
                
            properties[key] = value
            
        return properties
        
    except Exception as e:
        logger.warning(f"Error parsing chart properties: {e}")
        return {}


def _is_valid_range(range_str: str) -> bool:
    """
    Validate Excel range format like A1:A10 or B1:C10.
    """
    pattern = r'^[A-Z]+\d+:[A-Z]+\d+$'
    return bool(re.match(pattern, range_str.upper()))


def _is_valid_cell_ref(cell_ref: str) -> bool:
    """
    Validate Excel cell reference like A1 or B10.
    """
    pattern = r'^[A-Z]+\d+$'
    return bool(re.match(pattern, cell_ref.upper()))


def _get_first_sheet_name(excel_path: str) -> str:
    """
    Get the first sheet name from an Excel file.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_name = wb.sheetnames[0] if wb.sheetnames else 'Sheet1'
        wb.close()
        return sheet_name
    except Exception as e:
        logger.warning(f"Could not get sheet name from {excel_path}: {e}")
        return 'Sheet1'


# Export functions for MCP registration
__all__ = [
    'create_excel_chart',
    'style_excel_chart', 
    'update_chart_data'
]

