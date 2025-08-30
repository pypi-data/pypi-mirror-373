#!/usr/bin/env python3
"""
Optimized Excel Operations MCP Tools with AI-Friendly Syntax

This module provides the most intuitive DSL syntax for AI agents to generate Excel operations.
The syntax is designed to be as natural as possible while maintaining powerful functionality.
"""

import logging
import json
import re
from typing import Dict, Any, List, Optional, Union, Tuple
from pathlib import Path
import sys

# Setup logging
logger = logging.getLogger(__name__)

# Add the current directory to Python path for imports
root_dir_path = Path(__file__).parent.absolute()
sys.path.append(str(root_dir_path))

try:
    from parser.excel_parser import parse_markdown_formulas, write_formulas_to_excel_complex_agent
    from writer.excel_writer import ExcelWriter
except ImportError as e:
    logger.error(f"Failed to import required modules: {e}")
    raise


def excel_edit(excel_path: str, commands: str, sheet: str = None) -> str:
    """
    Ultra-simple Excel editing with natural language-like commands.
    
    This is the most AI-friendly tool - agents can generate commands intuitively.
    
    Args:
        excel_path: Path to Excel file
        commands: Natural commands separated by newlines
        sheet: Optional sheet name (auto-detected if not provided)
        
    Simple Commands Format:
        # Cell operations (most common)
        A1 = "Hello World"
        B1 = 42
        C1 = =SUM(A1:B1)
        
        # Formatting
        A1 bold
        B1 background yellow
        C1 font red
        
        # Row/Column operations
        insert row 5
        delete row 3
        column A width 20
        hide column B
        
        # Sheet operations
        add sheet "Data"
        rename sheet "Sheet1" to "Summary"
        delete sheet "Temp"
        
    Returns:
        JSON result of operations
        
    Examples:
        # Simple data entry
        commands = '''
        A1 = "Product"
        B1 = "Price"
        A2 = "Laptop"
        B2 = 999.99
        C2 = =B2*1.1
        '''
        
        # With formatting
        commands = '''
        A1 = "Total Sales"
        A1 bold
        A1 background blue
        A1 font white
        '''
    """
    logger.info(f"Starting Excel edit operations on {excel_path}")
    
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            return json.dumps({"error": f"Excel file not found: {excel_path}"})
        
        # Get the default sheet name properly 
        if not sheet:
            sheet = _get_default_sheet_name(str(excel_file))
        
        # Parse commands into operations
        operations = _parse_natural_commands(commands, sheet)
        
        if not operations:
            return json.dumps({"error": "No valid operations found"})
        
        # Execute operations with safe serialization
        results = []
        error_occurred = False
        
        # Execute cell operations if any
        if operations.get('cells'):
            cell_result = _execute_cell_operations(str(excel_file), operations['cells'])
            # Clean result for JSON serialization
            safe_result = _make_json_safe(cell_result)
            results.append({"type": "cells", "result": safe_result})
            if safe_result.get('error'):
                error_occurred = True
        
        # Execute row/column operations if any
        if operations.get('rows_columns'):
            rc_result = _execute_row_column_operations_optimized(str(excel_file), operations['rows_columns'])
            # Clean result for JSON serialization
            safe_result = _make_json_safe(rc_result)
            results.append({"type": "rows_columns", "result": safe_result})
            if safe_result.get('error'):
                error_occurred = True
        
        # Execute sheet operations if any
        if operations.get('sheets'):
            sheet_result = _execute_sheet_operations_optimized(str(excel_file), operations['sheets'])
            # Clean result for JSON serialization
            safe_result = _make_json_safe(sheet_result)
            results.append({"type": "sheets", "result": safe_result})
            if safe_result.get('error'):
                error_occurred = True
        
        response_data = {
            "success": not error_occurred,
            "excel_file": str(excel_file),
            "operations_executed": len([r for r in results if r]),
            "results": results
        }
        
        return json.dumps(response_data, indent=2)
        
    except Exception as e:
        error_msg = f"Error during Excel edit operations: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return json.dumps({"error": error_msg})


def excel_quick_cells(excel_path: str, cell_data: str, sheet: str = None) -> str:
    """
    Fastest way for agents to set cell values and formulas.
    
    Args:
        excel_path: Path to Excel file
        cell_data: Simple cell assignments
        sheet: Sheet name (optional)
        
    Format (one per line):
        A1 = "Text"
        B1 = 123
        C1 = =SUM(A1:B1)
        
    This is the simplest possible interface for AI agents.
    """
    if not sheet:
        sheet = _get_default_sheet_name(excel_path)
    
    commands = f"sheet: {sheet}\n{cell_data}"
    return excel_edit(excel_path, commands)


def excel_quick_format(excel_path: str, format_commands: str, sheet: str = None) -> str:
    """
    Quick formatting for agents.
    
    Args:
        excel_path: Path to Excel file  
        format_commands: Simple format commands
        sheet: Sheet name (optional)
        
    Format:
        A1 bold
        B1 background yellow
        C1 font red
    """
    if not sheet:
        sheet = _get_default_sheet_name(excel_path)
    
    commands = f"sheet: {sheet}\n{format_commands}"
    return excel_edit(excel_path, commands)


def _parse_natural_commands(commands: str, default_sheet: str) -> Dict[str, Any]:
    """
    Parse natural language-like commands into structured operations.
    
    This parser is designed to be very forgiving and intuitive for AI agents.
    """
    logger.info(f"Starting to parse natural commands for sheet: {default_sheet}")
    logger.debug(f"Commands to parse:\n{commands}")
    
    operations = {
        'cells': {},
        'rows_columns': [],
        'sheets': []
    }
    
    current_sheet = default_sheet
    lines = [line.strip() for line in commands.split('\n') if line.strip()]
    logger.info(f"Found {len(lines)} non-empty lines to parse")
    
    for i, line in enumerate(lines):
        line = line.strip()
        logger.debug(f"Line {i+1}: '{line}'")
        
        if not line or line.startswith('#'):
            logger.debug(f"Line {i+1}: Skipping empty or comment line")
            continue
            
        # Parse sheet declaration
        if line.lower().startswith('sheet:'):
            current_sheet = line.split(':', 1)[1].strip().strip('"\'')
            logger.info(f"Line {i+1}: Changed sheet to '{current_sheet}'")
            continue
            
        # Parse cell assignment: A1 = "value" or A1 = 123 or A1 = =FORMULA()
        if '=' in line and _is_cell_assignment(line):
            logger.debug(f"Line {i+1}: Processing as cell assignment")
            cell_ref, value = _parse_cell_assignment(line)
            if cell_ref:
                if current_sheet not in operations['cells']:
                    operations['cells'][current_sheet] = {}
                operations['cells'][current_sheet][cell_ref] = value
                logger.info(f"Line {i+1}: Added cell assignment {cell_ref} = '{value}' to sheet '{current_sheet}'")
            else:
                logger.warning(f"Line {i+1}: Failed to parse cell assignment: {line}")
            continue
        
        # Parse cell formatting: A1 bold, B1 background yellow, etc.
        if _is_cell_formatting(line):
            logger.debug(f"Line {i+1}: Processing as cell formatting")
            cell_ref, format_props = _parse_cell_formatting(line)
            if cell_ref and format_props:
                if current_sheet not in operations['cells']:
                    operations['cells'][current_sheet] = {}
                
                # If cell already exists, merge formatting
                if cell_ref in operations['cells'][current_sheet]:
                    if isinstance(operations['cells'][current_sheet][cell_ref], dict):
                        operations['cells'][current_sheet][cell_ref].update(format_props)
                        logger.info(f"Line {i+1}: Merged formatting for {cell_ref}: {format_props}")
                    else:
                        # Convert value to dict format
                        old_value = operations['cells'][current_sheet][cell_ref]
                        operations['cells'][current_sheet][cell_ref] = {
                            "formula": old_value,
                            **format_props
                        }
                        logger.info(f"Line {i+1}: Converted {cell_ref} to dict format with value '{old_value}' and formatting: {format_props}")
                else:
                    operations['cells'][current_sheet][cell_ref] = {
                        "formula": "[no_change]",
                        **format_props
                    }
                    logger.info(f"Line {i+1}: Added formatting-only for {cell_ref}: {format_props}")
            else:
                logger.warning(f"Line {i+1}: Failed to parse cell formatting: {line}")
            continue
        
        # Parse row/column operations
        if _is_row_column_operation(line):
            logger.debug(f"Line {i+1}: Processing as row/column operation")
            op = _parse_row_column_operation(line, current_sheet)
            if op:
                operations['rows_columns'].append(op)
                logger.info(f"Line {i+1}: Added row/column operation: {op}")
            else:
                logger.warning(f"Line {i+1}: Failed to parse row/column operation: {line}")
            continue
            
        # Parse sheet operations
        if _is_sheet_operation(line):
            logger.debug(f"Line {i+1}: Processing as sheet operation")
            op = _parse_sheet_operation(line)
            if op:
                operations['sheets'].append(op)
                logger.info(f"Line {i+1}: Added sheet operation: {op}")
            else:
                logger.warning(f"Line {i+1}: Failed to parse sheet operation: {line}")
            continue
        
        logger.warning(f"Line {i+1}: Unrecognized command: '{line}'")
    
    logger.info(f"Parsing complete. Operations summary:")
    logger.info(f"  - Cells operations: {len(operations['cells'])} sheets")
    for sheet_name, cells in operations['cells'].items():
        logger.info(f"    - Sheet '{sheet_name}': {len(cells)} cells")
        for cell_ref, cell_data in cells.items():
            logger.info(f"      - {cell_ref}: {cell_data}")
    logger.info(f"  - Row/column operations: {len(operations['rows_columns'])}")
    logger.info(f"  - Sheet operations: {len(operations['sheets'])}")
    
    return operations


def _is_cell_assignment(line: str) -> bool:
    """Check if line is a cell assignment like A1 = "value" """
    # Look for pattern like: CELL_REF = VALUE
    pattern = r'^[A-Z]+\d+\s*=\s*.+'
    return bool(re.match(pattern, line, re.IGNORECASE))


def _parse_cell_assignment(line: str) -> Tuple[str, Any]:
    """Parse A1 = "value" into (cell_ref, value)"""
    try:
        parts = line.split('=', 1)
        if len(parts) != 2:
            return None, None
            
        cell_ref = parts[0].strip().upper()
        value = parts[1].strip()
        
        # Validate cell reference
        if not re.match(r'^[A-Z]+\d+$', cell_ref):
            return None, None
        
        # Parse value
        if value.startswith('='):
            # Formula
            return cell_ref, value
        elif value.startswith('"') and value.endswith('"'):
            # Quoted string
            return cell_ref, value[1:-1]
        elif value.startswith("'") and value.endswith("'"):
            # Single quoted string  
            return cell_ref, value[1:-1]
        else:
            # Try to parse as number, otherwise treat as string
            try:
                if '.' in value:
                    return cell_ref, float(value)
                else:
                    return cell_ref, int(value)
            except ValueError:
                return cell_ref, value
                
    except Exception as e:
        logger.warning(f"Error parsing cell assignment '{line}': {e}")
        return None, None


def _is_cell_formatting(line: str) -> bool:
    """Check if line is cell formatting like A1 bold"""
    # Look for pattern: CELL_REF FORMAT_COMMAND
    format_keywords = [
        'bold', 'italic', 'underline',
        'background', 'font', 'color',
        'center', 'left', 'right',
        'border', 'size'
    ]
    
    parts = line.lower().split()
    if len(parts) < 2:
        return False
        
    # Check if first part looks like cell reference
    if not re.match(r'^[a-z]+\d+$', parts[0]):
        return False
        
    # Check if contains formatting keywords
    return any(keyword in line.lower() for keyword in format_keywords)


def _parse_cell_formatting(line: str) -> Tuple[str, Dict[str, Any]]:
    """Parse A1 bold background yellow into (cell_ref, format_props)"""
    try:
        parts = line.split()
        if len(parts) < 2:
            return None, None
            
        cell_ref = parts[0].upper()
        format_text = ' '.join(parts[1:]).lower()
        
        # Validate cell reference
        if not re.match(r'^[A-Z]+\d+$', cell_ref):
            return None, None
        
        format_props = {}
        
        # Parse formatting commands
        if 'bold' in format_text:
            format_props['b'] = True
        if 'italic' in format_text:
            format_props['it'] = True
        if 'underline' in format_text:
            format_props['u'] = True
        
        # Background colors
        if 'background yellow' in format_text or 'yellow background' in format_text:
            format_props['fill'] = '#FFFF00'
        elif 'background red' in format_text or 'red background' in format_text:
            format_props['fill'] = '#FF0000'
        elif 'background green' in format_text or 'green background' in format_text:
            format_props['fill'] = '#00FF00'
        elif 'background blue' in format_text or 'blue background' in format_text:
            format_props['fill'] = '#0000FF'
        elif 'background white' in format_text or 'white background' in format_text:
            format_props['fill'] = '#FFFFFF'
        elif 'background gray' in format_text or 'gray background' in format_text:
            format_props['fill'] = '#CCCCCC'
        
        # Font colors (check specific patterns first to avoid conflicts)
        if 'font white' in format_text or 'white font' in format_text:
            format_props['font'] = '#FFFFFF'
        elif 'font red' in format_text or 'red font' in format_text:
            format_props['font'] = '#FF0000'
        elif 'font blue' in format_text or 'blue font' in format_text:
            format_props['font'] = '#0000FF'
        elif 'font green' in format_text or 'green font' in format_text:
            format_props['font'] = '#00FF00'
        elif 'font black' in format_text or 'black font' in format_text:
            format_props['font'] = '#000000'
        
        # Alignment
        if 'center' in format_text:
            format_props['ha'] = 'center'
        elif 'left' in format_text:
            format_props['ha'] = 'left'  
        elif 'right' in format_text:
            format_props['ha'] = 'right'
        
        # Font size
        size_match = re.search(r'size (\d+)', format_text)
        if size_match:
            format_props['sz'] = size_match.group(1)
        
        return cell_ref, format_props if format_props else None
        
    except Exception as e:
        logger.warning(f"Error parsing cell formatting '{line}': {e}")
        return None, None


def _is_row_column_operation(line: str) -> bool:
    """Check if line is row/column operation"""
    keywords = [
        'insert row', 'delete row', 'hide row', 'unhide row',
        'insert column', 'delete column', 'hide column', 'unhide column',
        'column', 'row', 'width', 'height'
    ]
    return any(keyword in line.lower() for keyword in keywords)


def _parse_row_column_operation(line: str, sheet: str) -> Dict[str, Any]:
    """Parse row/column operations"""
    line_lower = line.lower().strip()
    
    try:
        # Insert row operations
        if line_lower.startswith('insert row'):
            match = re.search(r'insert row (\d+)', line_lower)
            if match:
                return {
                    'sheet': sheet,
                    'operation': 'insert_row',
                    'row_number': int(match.group(1)),
                    'count': 1
                }
        
        # Delete row operations  
        elif line_lower.startswith('delete row'):
            match = re.search(r'delete row (\d+)', line_lower)
            if match:
                return {
                    'sheet': sheet,
                    'operation': 'delete_row',
                    'row_number': int(match.group(1)),
                    'count': 1
                }
        
        # Hide row operations
        elif line_lower.startswith('hide row'):
            match = re.search(r'hide row (\d+)', line_lower)
            if match:
                return {
                    'sheet': sheet,
                    'operation': 'hide_row',
                    'row_number': match.group(1)
                }
        
        # Column width operations
        elif 'column' in line_lower and 'width' in line_lower:
            match = re.search(r'column ([a-z]+) width (\d+(?:\.\d+)?)', line_lower)
            if match:
                return {
                    'sheet': sheet,
                    'operation': 'resize_column',
                    'column': match.group(1).upper(),
                    'width': float(match.group(2))
                }
        
        # Insert column operations
        elif line_lower.startswith('insert column'):
            match = re.search(r'insert column ([a-z]+)', line_lower)
            if match:
                return {
                    'sheet': sheet,
                    'operation': 'insert_column',
                    'column': match.group(1).upper(),
                    'count': 1
                }
        
        # Delete column operations
        elif line_lower.startswith('delete column'):
            match = re.search(r'delete column ([a-z]+)', line_lower)
            if match:
                return {
                    'sheet': sheet,
                    'operation': 'delete_column',
                    'column': match.group(1).upper(),
                    'count': 1
                }
        
        # Hide column operations
        elif line_lower.startswith('hide column'):
            match = re.search(r'hide column ([a-z]+)', line_lower)
            if match:
                return {
                    'sheet': sheet,
                    'operation': 'hide_column',
                    'column': match.group(1).upper()
                }
        
        return None
        
    except Exception as e:
        logger.warning(f"Error parsing row/column operation '{line}': {e}")
        return None


def _is_sheet_operation(line: str) -> bool:
    """Check if line is sheet operation"""
    keywords = [
        'add sheet', 'delete sheet', 'rename sheet',
        'hide sheet', 'unhide sheet', 'copy sheet'
    ]
    return any(keyword in line.lower() for keyword in keywords)


def _parse_sheet_operation(line: str) -> Dict[str, Any]:
    """Parse sheet operations"""
    line_lower = line.lower().strip()
    
    try:
        # Add sheet operations
        if line_lower.startswith('add sheet'):
            match = re.search(r'add sheet ["\']([^"\']+)["\']', line)
            if match:
                return {
                    'operation': 'add_sheet',
                    'name': match.group(1)
                }
            # Handle without quotes
            match = re.search(r'add sheet (\w+)', line_lower)
            if match:
                return {
                    'operation': 'add_sheet',
                    'name': match.group(1)
                }
        
        # Delete sheet operations
        elif line_lower.startswith('delete sheet'):
            match = re.search(r'delete sheet ["\']([^"\']+)["\']', line)
            if match:
                return {
                    'operation': 'delete_sheet',
                    'name': match.group(1)
                }
            match = re.search(r'delete sheet (\w+)', line_lower)
            if match:
                return {
                    'operation': 'delete_sheet',
                    'name': match.group(1)
                }
        
        # Rename sheet operations
        elif 'rename sheet' in line_lower and ' to ' in line_lower:
            match = re.search(r'rename sheet ["\']([^"\']+)["\'] to ["\']([^"\']+)["\']', line)
            if match:
                return {
                    'operation': 'rename_sheet',
                    'old_name': match.group(1),
                    'new_name': match.group(2)
                }
            # Handle without quotes
            match = re.search(r'rename sheet (\w+) to (\w+)', line_lower)
            if match:
                return {
                    'operation': 'rename_sheet',
                    'old_name': match.group(1),
                    'new_name': match.group(2)
                }
        
        return None
        
    except Exception as e:
        logger.warning(f"Error parsing sheet operation '{line}': {e}")
        return None


def _execute_cell_operations(excel_path: str, cells_data: Dict[str, Dict]) -> Dict[str, Any]:
    """Execute cell operations using new simple parser"""
    logger.info(f"Executing cell operations for {len(cells_data)} sheets using new simple parser")
    
    try:
        # Use the new simple parser that bypasses markdown conversion
        from .simple_excel_parser import parse_and_execute_natural_commands
        
        logger.info("Using new simple parser for direct Excel operations")
        updated_cells = parse_and_execute_natural_commands(excel_path, cells_data)
        
        logger.info(f"Simple parser execution completed. Updated cells: {updated_cells}")
        
        return {
            "success": True,
            "updated_sheets": list(cells_data.keys()),
            "cells_modified": len(updated_cells) if updated_cells else 0,
            "details": updated_cells or [],
            "parser_used": "new_simple_parser"
        }
        
    except Exception as e:
        error_msg = f"Cell operations failed: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {"error": error_msg}


def _convert_cells_to_markdown(cells_data: Dict[str, Dict]) -> str:
    """Convert cells data to markdown format for existing parser"""
    logger.info(f"Converting cells data to markdown format")
    logger.debug(f"Input cells_data: {cells_data}")
    
    markdown_parts = []
    
    for sheet_name, cells in cells_data.items():
        logger.info(f"Processing sheet '{sheet_name}' with {len(cells)} cells")
        cell_entries = []
        
        for cell_ref, cell_data in cells.items():
            logger.debug(f"Processing cell {cell_ref}: {cell_data} (type: {type(cell_data)})")
            
            try:
                if isinstance(cell_data, dict):
                    # Complex format with properties
                    formula = cell_data.get('formula', '[no_change]')
                    logger.debug(f"Cell {cell_ref} formula: '{formula}'")
                    
                    # Only process if there's an actual formula/value (not just formatting)
                    if formula and formula != '[no_change]':
                        if not formula.startswith('='):
                            # Text or number value
                            formatted_formula = f'"{formula}"'
                        else:
                            # Excel formula
                            formatted_formula = f'"{formula}"'
                        
                        logger.debug(f"Cell {cell_ref} formatted formula: '{formatted_formula}'")
                    else:
                        # This is formatting-only, skip value but keep formatting
                        formatted_formula = '""'  # Empty value but apply formatting
                        logger.debug(f"Cell {cell_ref} is formatting-only")
                    
                    # Build properties string
                    props = []
                    for key, value in cell_data.items():
                        if key != 'formula':
                            if isinstance(value, bool):
                                props.append(f'{key}={str(value).lower()}')
                            elif isinstance(value, str) and value.startswith('#'):
                                props.append(f'{key}="{value}"')
                            else:
                                props.append(f'{key}="{value}"')
                    
                    if props:
                        entry = f'{cell_ref}, {formatted_formula}, {", ".join(props)}'
                        logger.debug(f"Cell {cell_ref} entry with props: '{entry}'")
                    else:
                        entry = f'{cell_ref}, {formatted_formula}'
                        logger.debug(f"Cell {cell_ref} entry no props: '{entry}'")
                    
                    cell_entries.append(entry)
                    
                else:
                    # Simple value
                    if isinstance(cell_data, str):
                        if cell_data.startswith('='):
                            formatted_value = f'"{cell_data}"'
                        else:
                            formatted_value = f'"{cell_data}"'
                    else:
                        formatted_value = f'"{str(cell_data)}"'
                    
                    entry = f'{cell_ref}, {formatted_value}'
                    logger.debug(f"Cell {cell_ref} simple entry: '{entry}'")
                    cell_entries.append(entry)
                    
            except Exception as e:
                logger.error(f"Error processing cell {cell_ref}: {e}")
                continue
        
        if cell_entries:
            sheet_entry = f'sheet_name: {sheet_name} | ' + ' | '.join(cell_entries)
            logger.info(f"Sheet '{sheet_name}' markdown: '{sheet_entry[:200]}...'")
            markdown_parts.append(sheet_entry)
        else:
            logger.warning(f"No valid cell entries for sheet '{sheet_name}'")
    
    result = ' | '.join(markdown_parts) if len(markdown_parts) > 1 else (markdown_parts[0] if markdown_parts else '')
    logger.info(f"Final markdown result length: {len(result)}")
    logger.debug(f"Final markdown result: '{result}'")
    
    return result


def _execute_row_column_operations_optimized(excel_path: str, operations: List[Dict]) -> Dict[str, Any]:
    """Execute row/column operations using new simple parser"""
    if not operations:
        return {"success": True, "operations_count": 0}
    
    try:
        from .simple_excel_parser import execute_row_column_operations
        return execute_row_column_operations(excel_path, operations)
        
    except Exception as e:
        logger.error(f"Row/column operations failed: {e}", exc_info=True)
        return {"error": f"Row/column operations failed: {str(e)}"}


def _execute_sheet_operations_optimized(excel_path: str, operations: List[Dict]) -> Dict[str, Any]:
    """Execute sheet operations using new simple parser"""
    if not operations:
        return {"success": True, "operations_count": 0}
    
    try:
        from .simple_excel_parser import execute_sheet_operations
        return execute_sheet_operations(excel_path, operations)
        
    except Exception as e:
        logger.error(f"Sheet operations failed: {e}", exc_info=True)
        return {"error": f"Sheet operations failed: {str(e)}"}


def _get_default_sheet_name(excel_path: str) -> str:
    """Get default sheet name from Excel file"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_name = wb.sheetnames[0] if wb.sheetnames else 'Sheet1'
        wb.close()
        return sheet_name
    except Exception as e:
        logger.warning(f"Could not get sheet name from {excel_path}: {e}")
        return 'Sheet1'


def _make_json_safe(data: Any) -> Any:
    """
    Recursively convert data to be JSON serializable by removing problematic objects.
    """
    if data is None:
        return None
    elif isinstance(data, (str, int, float, bool)):
        return data
    elif isinstance(data, dict):
        safe_dict = {}
        for key, value in data.items():
            # Skip FieldInfo objects and other non-serializable objects
            if hasattr(value, '__class__') and 'FieldInfo' in str(type(value)):
                continue
            safe_dict[str(key)] = _make_json_safe(value)
        return safe_dict
    elif isinstance(data, (list, tuple)):
        return [_make_json_safe(item) for item in data]
    elif hasattr(data, '__dict__'):
        # Convert objects to dict, but skip complex objects
        if 'FieldInfo' in str(type(data)):
            return str(data)
        try:
            return _make_json_safe(data.__dict__)
        except:
            return str(data)
    else:
        # Convert to string for anything else
        return str(data)


# Export the optimized functions
__all__ = [
    'excel_edit',
    'excel_quick_cells', 
    'excel_quick_format'
]
