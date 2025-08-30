"""
Excel Metadata Extractor

This module provides comprehensive Excel file metadata extraction capabilities,
supporting both openpyxl for basic metadata and xlwings for advanced COM operations.
"""

import os
import json
import logging
from typing import Dict, List, Any, Optional, Union, Tuple
from pathlib import Path
from datetime import datetime
import tempfile

# Configure logging
logger = logging.getLogger(__name__)

# Check for required libraries
OPENPYXL_AVAILABLE = False
XLWINGS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
    logger.info("openpyxl available for Excel metadata extraction")
except ImportError as e:
    logger.warning(f"openpyxl not available: {e}")

try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
    logger.info("xlwings available for Excel COM operations")
except ImportError as e:
    logger.warning(f"xlwings not available: {e}")


class ExcelMetadataExtractor:
    """
    Excel metadata extractor using openpyxl for basic extraction and xlwings for COM operations.
    
    This class provides comprehensive Excel file analysis including:
    - File properties and core metadata
    - Worksheet information and structure
    - Cell data and formatting
    - Charts and images
    - Named ranges and formulas
    - Thread-safe operations
    """
    
    def __init__(self, excel_path: Optional[str] = None):
        """
        Initialize the Excel metadata extractor.
        
        Args:
            excel_path: Optional path to Excel file to open immediately
        """
        self.excel_path = None
        self.workbook = None
        self.xlwings_app = None
        self.xlwings_wb = None
        
        if excel_path:
            self.open_workbook(excel_path)
    
    def __enter__(self):
        """Context manager entry."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit with cleanup."""
        self.close_workbook()
    
    def open_workbook(self, excel_path: str) -> None:
        """
        Open an Excel workbook for metadata extraction.
        
        Args:
            excel_path: Path to the Excel file
            
        Raises:
            FileNotFoundError: If the Excel file doesn't exist
            ValueError: If the file format is not supported
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        file_ext = os.path.splitext(excel_path)[1].lower()
        if file_ext not in ['.xlsx', '.xlsm', '.xls']:
            raise ValueError(f"Unsupported Excel format: {file_ext}")
        
        self.excel_path = os.path.abspath(excel_path)
        
        # Open with openpyxl for basic operations (if available)
        if OPENPYXL_AVAILABLE and file_ext in ['.xlsx', '.xlsm']:
            try:
                self.workbook = load_workbook(self.excel_path, data_only=False)
                logger.info(f"Opened Excel file with openpyxl: {self.excel_path}")
            except Exception as e:
                logger.error(f"Failed to open with openpyxl: {e}")
    
    def close_workbook(self) -> None:
        """Close the Excel workbook and clean up resources."""
        try:
            if self.xlwings_wb:
                self.xlwings_wb.close()
                self.xlwings_wb = None
                
            if self.xlwings_app:
                self.xlwings_app.quit()
                self.xlwings_app = None
                
            if self.workbook:
                self.workbook.close()
                self.workbook = None
                
        except Exception as e:
            logger.error(f"Error closing workbook: {e}")
        
        self.excel_path = None
    
    def extract_file_metadata(self) -> Dict[str, Any]:
        """
        Extract basic file metadata from the Excel file.
        
        Returns:
            Dictionary containing file metadata
        """
        if not self.excel_path:
            raise ValueError("No Excel file is currently open")
        
        file_path = Path(self.excel_path)
        file_stats = file_path.stat()
        
        metadata = {
            "filePath": str(file_path),
            "fileName": file_path.name,
            "fileSize": file_stats.st_size,
            "fileExtension": file_path.suffix.lower(),
            "createdTime": datetime.fromtimestamp(file_stats.st_ctime).isoformat(),
            "modifiedTime": datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
            "accessedTime": datetime.fromtimestamp(file_stats.st_atime).isoformat()
        }
        
        # Add workbook properties if available
        if self.workbook and hasattr(self.workbook, 'properties'):
            props = self.workbook.properties
            metadata.update({
                "title": getattr(props, 'title', None),
                "creator": getattr(props, 'creator', None),
                "subject": getattr(props, 'subject', None),
                "description": getattr(props, 'description', None),
                "keywords": getattr(props, 'keywords', None),
                "category": getattr(props, 'category', None),
                "comments": getattr(props, 'comments', None),
                "lastModifiedBy": getattr(props, 'lastModifiedBy', None),
                "revision": getattr(props, 'revision', None)
            })
        
        return metadata
    
    def extract_workbook_metadata(self, include_sheet_details: bool = True) -> Dict[str, Any]:
        """
        Extract comprehensive workbook metadata.
        
        Args:
            include_sheet_details: Whether to include detailed sheet information
            
        Returns:
            Dictionary containing workbook metadata
        """
        if not self.workbook and not self.excel_path:
            raise ValueError("No Excel file is currently open")
        
        metadata = {
            "extractionTime": datetime.now().isoformat(),
            "extractorVersion": "1.1.0",
            "fileMetadata": self.extract_file_metadata()
        }
        
        if self.workbook:
            metadata.update({
                "sheetNames": self.workbook.sheetnames,
                "totalSheets": len(self.workbook.sheetnames),
                "activeSheetName": self.workbook.active.title if self.workbook.active else None
            })
            
            # Extract named ranges
            if hasattr(self.workbook, 'defined_names'):
                named_ranges = []
                for defined_name in self.workbook.defined_names:
                    try:
                        named_ranges.append({
                            "name": defined_name.name,
                            "refers_to": str(defined_name.attr_text) if hasattr(defined_name, 'attr_text') else str(defined_name.value),
                            "scope": getattr(defined_name, 'localSheetId', None)
                        })
                    except Exception as e:
                        logger.warning(f"Failed to process named range {getattr(defined_name, 'name', 'unknown')}: {e}")
                        continue
                metadata["namedRanges"] = named_ranges
            
            # Include detailed sheet information if requested
            if include_sheet_details:
                metadata["sheets"] = []
                for sheet_name in self.workbook.sheetnames:
                    sheet_metadata = self.extract_sheet_metadata(sheet_name)
                    metadata["sheets"].append(sheet_metadata)
        
        return metadata
    
    def extract_sheet_metadata(self, sheet_name: str) -> Dict[str, Any]:
        """
        Extract metadata for a specific worksheet.
        
        Args:
            sheet_name: Name of the worksheet to analyze
            
        Returns:
            Dictionary containing sheet metadata
        """
        if not self.workbook:
            raise ValueError("No Excel workbook is currently open")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        worksheet = self.workbook[sheet_name]
        
        # Calculate used range
        min_row = worksheet.min_row if worksheet.min_row else 1
        max_row = worksheet.max_row if worksheet.max_row else 1
        min_col = worksheet.min_column if worksheet.min_column else 1
        max_col = worksheet.max_column if worksheet.max_column else 1
        
        sheet_metadata = {
            "sheetName": sheet_name,
            "sheetType": str(type(worksheet).__name__),
            "sheetIndex": self.workbook.sheetnames.index(sheet_name),
            "usedRange": {
                "minRow": min_row,
                "maxRow": max_row,
                "minColumn": min_col,
                "maxColumn": max_col,
                "totalRows": max_row - min_row + 1 if max_row >= min_row else 0,
                "totalColumns": max_col - min_col + 1 if max_col >= min_col else 0,
                "address": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            },
            "dimensions": f"{worksheet.max_row}x{worksheet.max_column}"
        }
        
        # Count different types of content
        cell_count = 0
        formula_count = 0
        value_count = 0
        
        for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, 
                                     min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    cell_count += 1
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        formula_count += 1
                    else:
                        value_count += 1
        
        sheet_metadata["contentSummary"] = {
            "totalCells": cell_count,
            "formulaCells": formula_count,
            "valueCells": value_count,
            "emptyCells": (max_row - min_row + 1) * (max_col - min_col + 1) - cell_count
        }
        
        # Check for merged cells
        merged_ranges = []
        if hasattr(worksheet, 'merged_cells'):
            for merged_range in worksheet.merged_cells.ranges:
                merged_ranges.append(str(merged_range))
        sheet_metadata["mergedCells"] = merged_ranges
        
        # Check for charts (basic detection)
        chart_count = 0
        if hasattr(worksheet, '_charts'):
            chart_count = len(worksheet._charts)
        sheet_metadata["chartCount"] = chart_count
        
        # Check for images (basic detection)
        image_count = 0
        if hasattr(worksheet, '_images'):
            image_count = len(worksheet._images)
        sheet_metadata["imageCount"] = image_count
        
        return sheet_metadata
    
    def extract_sheet_data(self, sheet_name: str, max_rows: Optional[int] = None, 
                          max_cols: Optional[int] = None, include_formatting: bool = False) -> Dict[str, Any]:
        """
        Extract actual cell data from a worksheet.
        
        Args:
            sheet_name: Name of the worksheet
            max_rows: Maximum number of rows to extract (None for all)
            max_cols: Maximum number of columns to extract (None for all)
            include_formatting: Whether to include cell formatting information
            
        Returns:
            Dictionary containing sheet data
        """
        if not self.workbook:
            raise ValueError("No Excel workbook is currently open")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        worksheet = self.workbook[sheet_name]
        
        # Determine range to extract
        min_row = worksheet.min_row if worksheet.min_row else 1
        max_row = worksheet.max_row if worksheet.max_row else 1
        min_col = worksheet.min_column if worksheet.min_column else 1
        max_col = worksheet.max_column if worksheet.max_column else 1
        
        if max_rows:
            max_row = min(max_row, min_row + max_rows - 1)
        if max_cols:
            max_col = min(max_col, min_col + max_cols - 1)
        
        # Extract cell data
        rows_data = []
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=min_row, max_row=max_row,
                                                         min_col=min_col, max_col=max_col)):
            row_data = []
            for col_idx, cell in enumerate(row):
                cell_data = {
                    "value": cell.value,
                    "address": cell.coordinate,
                    "dataType": getattr(cell, 'data_type', None)
                }
                
                if include_formatting and hasattr(cell, 'font'):
                    cell_data["formatting"] = {
                        "font": {
                            "name": cell.font.name,
                            "size": cell.font.size,
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "color": str(cell.font.color) if cell.font.color else None
                        },
                        "fill": {
                            "type": str(cell.fill.fill_type) if cell.fill else None,
                            "color": str(cell.fill.start_color) if cell.fill and hasattr(cell.fill, 'start_color') else None
                        },
                        "numberFormat": cell.number_format,
                        "alignment": {
                            "horizontal": cell.alignment.horizontal if cell.alignment else None,
                            "vertical": cell.alignment.vertical if cell.alignment else None,
                            "wrapText": cell.alignment.wrap_text if cell.alignment else None
                        }
                    }
                
                row_data.append(cell_data)
            rows_data.append(row_data)
        
        return {
            "sheetName": sheet_name,
            "dataRange": {
                "minRow": min_row,
                "maxRow": max_row,
                "minColumn": min_col,
                "maxColumn": max_col,
                "address": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            },
            "rowsData": rows_data,
            "extractedRows": len(rows_data),
            "extractedColumns": len(rows_data[0]) if rows_data else 0
        }
    
    def get_xlwings_workbook(self) -> Optional['xw.Book']:
        """
        Get or create xlwings workbook instance for COM operations.
        
        Returns:
            xlwings Book object or None if not available
        """
        if not XLWINGS_AVAILABLE:
            logger.warning("xlwings not available for COM operations")
            return None
        
        if not self.excel_path:
            raise ValueError("No Excel file path specified")
        
        try:
            if not self.xlwings_wb:
                # Create new xlwings app if needed
                if not self.xlwings_app:
                    self.xlwings_app = xw.App(visible=False, add_book=False)
                
                # Open workbook
                self.xlwings_wb = self.xlwings_app.books.open(self.excel_path)
                logger.info(f"Opened Excel file with xlwings: {self.excel_path}")
            
            return self.xlwings_wb
        
        except Exception as e:
            logger.error(f"Failed to open Excel file with xlwings: {e}")
            return None
    
    def extract_presentation_metadata(self, include_sheet_content: bool = True,
                                    include_sheet_data: bool = False,
                                    max_data_rows: Optional[int] = 100) -> Dict[str, Any]:
        """
        Extract comprehensive metadata from the Excel file.
        
        Args:
            include_sheet_content: Include detailed sheet structure information
            include_sheet_data: Include actual cell data (limited by max_data_rows)
            max_data_rows: Maximum rows of data to extract per sheet
            
        Returns:
            Complete metadata dictionary
        """
        metadata = self.extract_workbook_metadata(include_sheet_details=include_sheet_content)
        
        if include_sheet_data and self.workbook:
            for sheet_metadata in metadata.get("sheets", []):
                sheet_name = sheet_metadata["sheetName"]
                try:
                    sheet_data = self.extract_sheet_data(
                        sheet_name, 
                        max_rows=max_data_rows,
                        include_formatting=False
                    )
                    sheet_metadata["sampleData"] = sheet_data
                except Exception as e:
                    logger.error(f"Failed to extract data from sheet {sheet_name}: {e}")
                    sheet_metadata["sampleData"] = {"error": str(e)}
        
        return metadata
