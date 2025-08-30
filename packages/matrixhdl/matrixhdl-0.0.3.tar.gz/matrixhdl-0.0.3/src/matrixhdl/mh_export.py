##/////////////////////////////////////////////////////////////////////////////////////
##//
##//Copyright 2025  Matrix HDL
##//
##//Licensed under the Apache License, Version 2.0 (the "License");
##//you may not use this file except in compliance with the License.
##//You may obtain a copy of the License at
##//
##//    http://www.apache.org/licenses/LICENSE-2.0
##//
##//Unless required by applicable law or agreed to in writing, software
##//distributed under the License is distributed on an "AS IS" BASIS,
##//WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
##//See the License for the specific language governing permissions and
##//limitations under the License.
##//
##/////////////////////////////////////////////////////////////////////////////////////

###################################################################################
##import
###################################################################################
import openpyxl    ##openpyxl.load_workbook()
import os          ##os.path.exists()/os.makedirs()/os.getcwd()
import shutil      ##shutil.rmtree()
###################################################################################
##global definition
###################################################################################
default_split_char        = '.'
default_debug_symbol      = 'DEBUG'
default_xlsx_file_name    = 'verilog.xlsx'
default_relative_dir_name = 'rtl'
default_flist_f_prefix    = '../%s/' % default_relative_dir_name
default_flist_f_header    = ''
###################################################################################
##class:SysArgv
###################################################################################
class SysArgv:
    ''' To analyze sys.argv and generate parameters. '''
    from sys import argv 
    xlsx_file         = default_xlsx_file_name    
    relative_dir_name = default_relative_dir_name
    root_path         = "./"
    inner_path        = './'+relative_dir_name+'/'
    dir_path          = root_path + inner_path
    f_name            = relative_dir_name+'.f' 
    f_prefix          = default_flist_f_prefix 
    f_header          = default_flist_f_header    
    output_single     = False    
    
    def __init__(self):
        del self.argv[0]
        given_para_title = ['-o','-single']
        given_para_value = [self.get_extract_name(i) for i in given_para_title]
        oname = given_para_value[0]
        self.output_single = given_para_value[1]
        input_xlsx_file = [i for i in self.argv if i not in given_para_title+given_para_value]
        if oname:
            self.relative_dir_name = oname
            self.inner_path = './%s/'%(self.relative_dir_name)
            self.f_name     = oname.rstrip('/\\')+'.f'
            self.f_prefix   = ''
            self.f_header   = ''          
        if input_xlsx_file:
            input_xlsx_file = input_xlsx_file[0]
            (file_path,file_name) = os.path.split(input_xlsx_file)
            self.xlsx_file  = input_xlsx_file
            self.root_path  = file_path+'/' if file_path else self.root_path
        self.dir_path = self.root_path + self.inner_path
        
    def get_extract_name(self,char):
        if char in self.argv:
            if self.argv.index(char)<((len(self.argv))-1) and not self.argv[self.argv.index(char)+1].startswith('-'):
                value = self.argv[self.argv.index(char)+1]
            else:
                value = True
        else:
            value = False
        return value
            
    def do_make_path(self,path):
        if os.path.exists(path):
            shutil.rmtree(path)
        os.makedirs(path)
        
    def do_check_path(self,path):
        if not os.path.exists(path):
            os.makedirs(path)     

###################################################################################
##class: SheetRtl
###################################################################################
class SheetRtl:
    '''RTL of one sheet'''
    def initial(self):
        self.rtl       = False
        self.define    = False
        self.file_name = ''
        self.file_path = ''
        self.top_name  = ''
        self.top_para  = ''
        self.word_title        = ''
        self.word_port         = []
        self.word_wire         = []
        self.word_parameter    = {}
        self.word_body         = {}        
        self.endpoint          = {}
        self.endpoint['name']  = []
        self.endpoint['type']  = []
        self.endpoint['width'] = []
        self.endpoint['debug'] = []
        
        self.list_sub_ref = []
        self.list_sub_ins = []
        self.list_sub_pin = {}
        self.list_sub_net = {}
        
        self.mode_clock   = 'pos'
        self.signal_clock = 'clk'
        self.mode_rstn    = 'neg'
        self.signal_rstn  = 'rst_n' 
        
        self.dict_debug   = {}
    
    def __init__(self,sheet):
        self.initial()
        self.sheet = sheet
        invalid =  self.get_cell(0,0) and self.get_cell(0,0)[0]=='@' 
        if sheet.title[0] not in ['#','@'] and not invalid:
            self.rtl    = sheet.title[0]!='`'
            self.define = sheet.title[0]=='`' 
            self.do_basic_info(sheet.title)
            if self.rtl:
                self.do_process_rtl()
            else:
                self.do_process_define()    
    
    def get_cell_val(self,value):
        return str(value).strip() if value else ''
    
    def get_col(self,col):
        return [self.get_cell_val(self.sheet.cell(row+1,col+1).value) for row in range(self.sheet.max_row)]
        
    def get_row(self,row):
        return [self.get_cell_val(self.sheet.cell(row+1,col+1).value) for col in range(self.sheet.max_column)]
    
    def get_cell(self,row,col):
        return self.get_cell_val(self.sheet.cell(row+1,col+1).value)
    
    def do_basic_info(self,title):
        title_rest      = title[1:] if title[0] in ['#','@','`'] else title
        title_split     = title_rest.strip().split(default_split_char)
        self.file_name  = title_split[-1] + '.v'
        path            = '/'.join(title_split[0:-1])
        self.file_path  =  path + ('/' if path else '')  
    
    def do_process_define(self):
        for col in range(self.sheet.max_column):
            self.word_body[col]              = {}
            self.word_body[col]['submodule'] = ''
            self.word_body[col]['word']      = []
            if col==0:               
                for j in range(self.sheet.max_row):
                    cell_col0 = self.get_cell(j,0)
                    cell_col1 = self.get_cell(j,1)
                    if cell_col0:
                        if any([cell_col0.startswith(i) for i in ['//','`']]):
                            self.word_body[col]['word'].append(cell_col0+cell_col1)
                        else:
                            self.word_body[col]['word'].append('`define %s %s' %(cell_col0,(cell_col1 if cell_col1 else '')))
            elif (col!=1):
                for row in range(self.sheet.max_row):
                    if self.get_cell(row,col):
                        self.word_body[col]['word'].append(self.get_cell(row,col)) 
    
    def do_process_rtl(self):
        self.do_first_column()
        self.do_other_column()  
        
    def get_ljust_string(self,*string):
        next_continue = False
        line = ''
        for i in range(len(string)):
            if next_continue:
                next_continue = False
                continue
            curr          = string[i]
            next          = string[i+1] if i<(len(string)-1) else ''
            length        = next if isinstance(next,int) else 40
            next_continue = True if isinstance(next,int) else False
            line += curr.ljust(length,' ')
        return line
    
    def get_port_format(self,type,width,name):
        return '    '+self.get_ljust_string(type,10,width,name)
        
    def get_wire_format(self,type,width,name):
        return '    '+self.get_ljust_string(type,10,width,name)+';'
        
    def get_parameter_format(self,type,width,name,pval):
        return '    '+self.get_ljust_string(type,10,width,name+' = '+pval)+';'
    
    def get_instance_format(self,sub_pin,sub_net):
        return '    '+self.get_ljust_string('.'+sub_pin,40,'(    '+sub_net,40,')',1)
    
    def do_first_column(self):
        for i in range(self.sheet.max_row):
            each = self.get_cell(i,0)
            self.endpoint['name'].append('')
            self.endpoint['type'].append('')
            self.endpoint['width'].append('')
            self.endpoint['debug'].append(False)
            if i==0:
                if self.get_cell(1,0) and self.get_cell(1,0)[0]=='#':
                    each += ' '+self.get_cell(1,0) 
                self.top_name = each.lstrip('@').split()[0]
                self.top_para = each[each.index('#'):each.rindex(')')+1] if '#' in each else ''            
            elif each:
                if i==1 and each[0]=='#':
                    continue
                if each[0]=='`' or (len(each)>=2 and each[0:2]=='//'):
                    self.word_port.append(each)
                    self.word_wire.append(each)
                    continue
                if '#' in each:
                    name,type,width,pval,debug = self.get_name_type_width_pval(each)
                    self.endpoint['name'][i]   = name
                    self.endpoint['type'][i]   = type
                    self.endpoint['width'][i]  = width
                    self.endpoint['debug'][i]  = debug
                    if type in ['input','output','inout']:
                        self.word_port.append(self.get_port_format(type,width,name)) 
                    elif type in ['wire','reg']:
                        self.word_wire.append(self.get_wire_format(type,width,name))
                    else:
                        self.word_parameter[name] = self.get_parameter_format(type,width,name,pval)
                        #self.word_wire.append(self.get_parameter_format(type,width,name,pval))
        self.word_port = self.get_remove_empty_define(self.word_port)
        self.word_wire = self.get_remove_empty_define(self.word_wire) 
        key_remove = []
        for key in self.word_parameter:
            if any([key in self.endpoint['width'][i] and self.endpoint['type'][i] in ['input','output','inout'] for i in range(len(self.endpoint['name']))]):
                self.top_para = self.top_para.rstrip(')').rstrip('\n')+',\n'+self.word_parameter[key].rstrip(';')+'\n)' if self.top_para else '#(\n'+self.word_parameter[key].rstrip(';')+'\n)'
                key_remove.append(key)
        [self.word_parameter.pop(key) for key in key_remove] 
        self.word_title = 'module %s %s\n' % (self.top_name,self.top_para)             
        self.do_add_port_comma()
        
    def get_remove_empty_define(self,word):
        count = 0
        while count<(len(word)-1):
            if word[count].startswith('`if') and word[count+1].startswith('`endif'):
                del word[count]
                del word[count]
            else:
                count += 1
        return word
    
    get_last_position =  lambda self,one: len(one)-1-one[::-1].index(True) if any(one) else len(one)
    
    def do_add_port_comma(self):
        is_port = [any([each.strip().startswith(type) for type in ['input','output','inout']]) for i,each in enumerate(self.word_port)]
        is_start_define = [each.strip().startswith('`ifdef') for i,each in enumerate(self.word_port)]
        flag_define      = False
        is_inside_define = []
        for i,each in enumerate(self.word_port):
            flag_define  = True if each.strip().startswith('`ifdef') else False if each.strip().startswith('`endif') else flag_define
            is_inside_define.append(flag_define)
        is_inside_port   = [is_port[i] and is_inside_define[i] for i in range(len(is_port))]
        is_outside_port  = [is_port[i] and not is_inside_port[i] for i in range(len(is_port))]
        order_of_start_define = self.get_last_position(is_start_define)
        order_of_inside  = self.get_last_position(is_inside_port)
        order_of_outside = self.get_last_position(is_outside_port)
        odd    = (order_of_inside<len(is_port)) and (order_of_outside<len(is_port)) and (order_of_outside<order_of_inside)
        last_inside  = len(is_port) if (order_of_inside<len(is_port)) and (order_of_inside<order_of_outside) else order_of_inside
        last_outside = order_of_outside
        for i,each in enumerate(self.word_port):
            if odd and (i==order_of_start_define):
                self.word_port[i] = each+'\n,'
            if not ((i==last_inside) or (i==last_outside)) and any([each.strip().startswith(type) for type in ['input','output','inout']]):
                self.word_port[i] = each+','    
    
    def get_name_type_width_pval(self,value):
        debug = value[0]=='*'
        name  = value.lstrip('*').split()[0]
        type = {'#i':'input','#o':'output','#x':'inout','#w':'wire','#r':'reg','#p':'parameter'}[value[value.index('#'):value.index('#')+2]]
        width = value[value.find('#')+2:].strip()
        pval  = ''
        if '#p' in value:
            if len(width.split())==1:
                pval  = width if width else '0'
                width = ''
            else:
                pval  = ' '.join(width.split()[1:])
                width = width.split()[0]
        width = '[(%s)-1:0]' % width if width and '[' not in width else width
        return name,type,width,pval,debug
    
    def do_other_column(self):
        for col_cnt in range(self.sheet.max_column):
            self.word_body[col_cnt]              = {}
            self.word_body[col_cnt]['submodule'] = ''
            self.word_body[col_cnt]['word']      = []
            col_val = self.get_cell(0,col_cnt)
            if col_cnt:
                if col_val:
                    self.do_module(col_cnt)
                else:
                    self.do_content(col_cnt)
    
    def do_module(self,col_cnt):
        cell0 = self.get_cell(0,col_cnt)
        cell1 = self.get_cell(1,col_cnt)
        if cell1 and cell1.startswith('#'):
            module_info = cell0.split()[0] + cell1 + ( '' if len(cell0.split())==1 else cell0.split()[-1] )
        else:
            module_info = cell0
        ref_name = module_info.split()[0]
        ins_name = ref_name if len(module_info.split())==1 or module_info[-1]==')' else  module_info.split()[-1]
        if ins_name in self.list_sub_ins:
            ins_cnt = 0
            while ins_name+'_'+str(ins_cnt) in self.list_sub_ins:
                ins_cnt += 1
            ins_name = ins_name+'_'+str(ins_cnt)
        para = module_info[module_info.index('#'):module_info.rindex(')')+1] if '#' in module_info else ''
        self.list_sub_ref.append(ref_name)
        self.list_sub_ins.append(ins_name)
        self.list_sub_pin[ins_name] = []
        self.list_sub_net[ins_name] = []
        module_para = {}
        for i,each in enumerate(self.get_col(col_cnt)):
            if i and each:
                if (i==1 and each[0]=='#'):
                    continue
                if each[0]=='`' or (len(each)>=2 and each[0:2]=='//'):
                    self.word_body[col_cnt]['word'].append(each)
                    continue
                sub_pin = each.split()[0]    
                if '#' in each:
                    if '#p' in each and (not self.endpoint['name'][i] or (self.endpoint['type'][i]!='parameter')):
                        sub_net = each[each.index('#p')+2:].strip().split()[-1]
                    else:
                        sub_net = self.endpoint['name'][i]
                elif len(each)==1:
                    sub_net = ''
                else:
                    sub_net = ' '.join(each.split()[1:])
                    sub_net = sub_pin+sub_net[1:] if sub_net[0]=='.' else self.endpoint['name'][i]+sub_net[1:] if sub_net[0]=='-' else sub_net
                self.list_sub_pin[ins_name].append(sub_pin)
                self.list_sub_net[ins_name].append(sub_net)
                if '#p' in each:
                    module_para[sub_pin] = self.get_instance_format(sub_pin,sub_net)
                else:
                    self.word_body[col_cnt]['word'].append(self.get_instance_format(sub_pin,sub_net))
        if module_para:
            for key in module_para:
                para = para.rstrip(')').rstrip('\n')+',\n'+module_para[key].rstrip(',')+'\n)' if para else '#(\n'+module_para[key].rstrip(',')+'\n)'
        self.word_body[col_cnt]['submodule'] = '%s %s %s (' %(ref_name,para,ins_name)        
        self.do_add_instance_comma(col_cnt)
    
    def do_add_instance_comma(self,col_cnt):
        is_port = [each.strip().startswith('.') for i,each in enumerate(self.word_body[col_cnt]['word'])]
        is_start_define = [each.strip().startswith('`ifdef') for i,each in enumerate(self.word_body[col_cnt]['word'])]
        flag_define      = False
        is_inside_define = []
        for i,each in enumerate(self.word_body[col_cnt]['word']):
            flag_define  = True if each.strip().startswith('`ifdef') else False if each.strip().startswith('`endif') else flag_define
            is_inside_define.append(flag_define)
        is_inside_port   = [is_port[i] and is_inside_define[i] for i in range(len(is_port))]
        is_outside_port  = [is_port[i] and not is_inside_port[i] for i in range(len(is_port))]
        order_of_start_define = self.get_last_position(is_start_define)
        order_of_inside  = self.get_last_position(is_inside_port)
        order_of_outside = self.get_last_position(is_outside_port)
        odd    = (order_of_inside<len(is_port)) and (order_of_outside<len(is_port)) and (order_of_outside<order_of_inside)
        last_inside  = len(is_port) if (order_of_inside<len(is_port)) and (order_of_inside<order_of_outside) else order_of_inside
        last_outside = order_of_outside
        for i,each in enumerate(self.word_body[col_cnt]['word']):
            if odd and (i==order_of_start_define):
                self.word_body[col_cnt]['word'][i] = each+'\n,'
            if not ((i==last_inside) or (i==last_outside)) and each.strip().startswith('.'):
                self.word_body[col_cnt]['word'][i] = each+','         
        
    def do_content(self,col_cnt):
        empty_column = True
        for i,each in enumerate(self.get_col(col_cnt)):
            if i and each:
                empty_column = False
                if each[0]=='#' and self.endpoint['name'][i] and self.endpoint['type'][i]=='reg':
                    each_split = each.split()
                    def_val = each_split[0][1:] if len(each_split[0])>1 else '0'
                    other = ' '.join(each_split[1:])
                    other = self.endpoint['name'][i]+' '+other if other.startswith('<=') else other
                    self.word_body[col_cnt]['word'].append(self.get_ff_title(self.endpoint['name'][i],def_val)+other)
                elif each[0]=='#' and self.endpoint['name'][i] and self.endpoint['type'][i] in ['wire','output','inout']:
                    each_split = each[1:].split()
                    line = self.endpoint['name'][i]+' '+each[1:] if each_split[0][0]=='=' else each[1:]
                    line = 'assign ' + line
                    self.word_body[col_cnt]['word'].append(line)
                elif each[0]=='#' and not self.endpoint['name'][i]:
                    self.mode_clock,self.signal_clock = each[1:].split()[0:2]
                    if len(each[1:].split())>=4:
                        self.mode_rstn,self.signal_rstn = each[1:].split()[2:4]
                    else:
                        self.mode_rstn,self.signal_rstn = ['','']
                else:
                    self.word_body[col_cnt]['word'].append(each)
        if not empty_column:
            self.word_body[col_cnt]['word'].append('\n')
        
    def get_ff_title(self,signal,value):
        title  = 'always @ ( %sedge %s ' %(self.mode_clock,self.signal_clock)
        title += 'or %sedge %s )\n' %(self.mode_rstn,self.signal_rstn) if self.mode_rstn else ')\n'
        if self.mode_rstn:
            negtive_char = '~' if self.mode_rstn=='neg' else ''
            title += 'if (%s%s)\n    %s <= %s;\nelse\n    ' %(negtive_char,self.signal_rstn,signal,value)
        return title
        
    def get_debug_port(self):
        line = ''
        if self.dict_debug['port_name']:
            line += '`ifdef %s\n' % default_debug_symbol
            for i in range(len(self.dict_debug['port_name'])):
                line += (',\n'+self.get_port_format('output',self.dict_debug['port_width'][i],default_debug_symbol+'_'+self.dict_debug['port_name'][i]))
            line += '\n`endif\n'    
        return line
        
    def get_debug_line(self):
        line = ''
        if self.dict_debug['debug_name']:
            line += '`ifdef %s\n' % default_debug_symbol
            for i in range(len(self.dict_debug['debug_name'])):
                line += ( '    assign %s_%s = %s;\n' %(default_debug_symbol,self.dict_debug['debug_name'][i],self.dict_debug['debug_name'][i]) )
            line += '`endif\n'    
        return line 
    
    def get_debug_submodule(self,cnt):
        line = ''
        if self.dict_debug['submodule']['pin'][cnt]:
            line += '`ifdef %s\n' % default_debug_symbol
            for i in range(len(self.dict_debug['submodule']['pin'][cnt])):
                line += (',\n'+self.get_instance_format(default_debug_symbol+'_'+self.dict_debug['submodule']['pin'][cnt][i],default_debug_symbol+'_'+self.dict_debug['submodule']['net'][cnt][i]))
            line += '\n`endif\n'    
        return line           
        
    def get_all_word(self):
        line  = ''
        if self.rtl:
            line += self.word_title
            line += '(\n'
            line += ('\n'.join(self.word_port)+'\n')
            line += self.get_debug_port()
            line += ');\n\n'
            line += ('\n'.join(self.word_parameter.values())+'\n')
            line += ('\n'.join(self.word_wire)+'\n')
            line += '\n'
            line += self.get_debug_line()
            cnt = 0
            for i in range(len(self.word_body)):
                if self.word_body[i]['submodule']:
                    line += self.word_body[i]['submodule']+'\n'
                    line += '\n'.join(self.word_body[i]['word'])+'\n' 
                    line += self.get_debug_submodule(cnt)
                    line += ');\n\n'
                    cnt  += 1
                else:
                    line += '\n'.join(self.word_body[i]['word'])
            line += 'endmodule\n'
        else:
            for i in range(len(self.word_body)):
                line += '\n'.join(self.word_body[i]['word'])
            line += '\n'
        return line

###################################################################################
##Class: SheetAll
###################################################################################
class SheetAll:
    '''All valid sheets'''
    class_sheets   = {}
    dict_debug     = {}
    include_header = {}
    def __init__(self,argv):
        self.do_build_class_sheet(argv)
        self.do_initial_debug()
        self.do_insert_debug()
        self.do_inject_debug()
        self.do_output_vfiles(argv)        
    
    def do_build_class_sheet(self,argv):
        wb = openpyxl.load_workbook(argv.xlsx_file)
        for each in wb.sheetnames:
            if each[0] not in ['#','@']:
                self.class_sheets[each] = SheetRtl(wb[each])  
    
    def do_output_vfiles(self,argv):
        if argv.output_single:
            self.do_write_one_file(argv)
        else:
            argv.do_make_path(argv.dir_path)
            self.do_gen_f_file(argv)
            self.do_get_include_header(argv)
            self.do_gen_v_files(argv)
    
    def do_write_one_file(self,argv):
        with open(argv.root_path+argv.relative_dir_name+'.v','w') as f:
            for each in self.class_sheets:
                one = self.class_sheets[each]
                if one.define:
                    f.write(one.get_all_word()+'\n')
            for each in self.class_sheets:
                one = self.class_sheets[each]
                if one.rtl:
                    f.write(one.get_all_word()+'\n')
        
    def do_gen_f_file(self,argv):
        with open(argv.dir_path+argv.f_name,'w') as f:
            f.write(argv.f_header)
            for each in self.class_sheets:
                this_class = self.class_sheets[each]
                if this_class.rtl:
                    f.write(argv.f_prefix+this_class.file_path+this_class.file_name+'\n')
    
    def do_get_include_header(self,argv):
        self.include_header = ''
        for each in self.class_sheets:
            one = self.class_sheets[each]
            if one.define:
                self.include_header += '`include "%s%s%s"\n' %(argv.f_prefix,one.file_path,one.file_name)
    
    def do_gen_v_files(self,argv):
        for each in self.class_sheets:
            one = self.class_sheets[each]
            if one.rtl or one.define:
                argv.do_check_path(argv.dir_path+one.file_path)
                with open(argv.dir_path+one.file_path+one.file_name,'w') as v:
                    if self.class_sheets[each].rtl:
                        v.write(self.include_header)
                    v.write(one.get_all_word())
    
    def do_initial_debug(self):
        dict = {}
        for each in self.class_sheets:
            curr_class = self.class_sheets[each]
            dict[each]                      = {}
            dict[each]['module_name']       = curr_class.top_name
            dict[each]['port_name']         = []
            dict[each]['port_width']        = []
            dict[each]['submodule']         = {}
            dict[each]['submodule']['ref']  = curr_class.list_sub_ref
            dict[each]['submodule']['ins']  = curr_class.list_sub_ins
            dict[each]['submodule']['pin']  = {}
            dict[each]['submodule']['net']  = {}
            for i in range(len(dict[each]['submodule']['ref'])):
                dict[each]['submodule']['pin'][i]  = []
                dict[each]['submodule']['net'][i]  = []
            dict[each]['debug_name']        = []
            dict[each]['debug_width']       = []
            for i in range(len(curr_class.endpoint['debug'])):
                if curr_class.endpoint['debug'][i]:
                    dict[each]['debug_name'].append(curr_class.endpoint['name'][i])
                    dict[each]['debug_width'].append(curr_class.endpoint['width'][i])
        self.dict_debug = dict        
    
    def do_insert_debug(self):
        for each in self.dict_debug:
            if self.dict_debug[each]['debug_name']:
                for i in range(len(self.dict_debug[each]['debug_name'])):
                    self.do_insert_port(each,self.dict_debug[each]['debug_name'][i],self.dict_debug[each]['debug_width'][i])
                    
    def do_insert_port(self,ins_dict_name,ins_port_name,ins_port_width):
        self.dict_debug[ins_dict_name]['port_name'].append(ins_port_name)
        self.dict_debug[ins_dict_name]['port_width'].append(ins_port_width)
        for each in self.dict_debug:
            for i in range(len(self.dict_debug[each]['submodule']['ref'])):
                if self.dict_debug[ins_dict_name]['module_name']==self.dict_debug[each]['submodule']['ref'][i]:
                    self.dict_debug[each]['submodule']['pin'][i].append(ins_port_name)
                    net = ins_port_name
                    if net in self.dict_debug[each]['port_name']:
                        cnt = 0
                        while ins_port_name+'_'+str(cnt) in self.dict_debug[each]['port_name']:
                            cnt += 1
                        net = ins_port_name+'_'+str(cnt)
                    self.dict_debug[each]['submodule']['net'][i].append(net)
                    self.do_insert_port(each,net,ins_port_width)
    
    def do_inject_debug(self):
        for each in self.dict_debug:
            self.class_sheets[each].dict_debug = self.dict_debug[each]

###################################################################################
##Main
###################################################################################
#if __name__ == '__main__': 
def main():
    class_all = SheetAll(SysArgv())
    #zip output files: rtl.zip ./rtl
    #shutil.make_archive('rtl','zip','.','rtl')