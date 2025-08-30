##/////////////////////////////////////////////////////////////////////////////////////
##//
##//Copyright 2025  Matrix HDL
##//
##//Licensed under the Apache License, Version 2.0 (the "License");
##//you may not use this file except in compliance with the License.
##//You may obtain a copy of the License at
##//
##//    http://www.apache.org/licenses/LICENSE-2.0
##//
##//Unless required by applicable law or agreed to in writing, software
##//distributed under the License is distributed on an "AS IS" BASIS,
##//WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
##//See the License for the specific language governing permissions and
##//limitations under the License.
##//
##/////////////////////////////////////////////////////////////////////////////////////

###################################################################################
##import
###################################################################################
import openpyxl    ##openpyxl.load_workbook()
import os          ##os.path.exists()/os.makedirs()/os.getcwd()
import shutil      ##shutil.rmtree()
default_split_char      = '.'
###################################################################################
##class:SysArgv
###################################################################################
class SysArgv:
    ''' To analyze sys.argv and generate parameters. '''
    from sys import argv 
    xlsx_file = 'verilog.xlsx'  
    
    def __init__(self):
        del self.argv[0]
        given_para_title = []
        given_para_value = [self.get_extract_name(i) for i in given_para_title]
        file = [i for i in self.argv if i not in given_para_title+given_para_value]
        self.xlsx_file = file[0] if file else self.xlsx_file
        
    def get_extract_name(self,char):
        if char in self.argv:
            if self.argv.index(char)<((len(self.argv))-1) and not self.argv[self.argv.index(char)+1].startswith('-'):
                value = self.argv[self.argv.index(char)+1]
            else:
                value = True
        else:
            value = False
        return value
            
    def do_make_path(self,path):
        if os.path.exists(path):
            shutil.rmtree(path)
        os.makedirs(path)
        
    def do_check_path(self,path):
        if not os.path.exists(path):
            os.makedirs(path)     

###################################################################################
##class: ClassRtl
###################################################################################
class ClassRtl:
    '''RTL of one sheet'''
    def initial(self):
        self.file_name   = ''
        self.module_name = ''
        self.top_name    = ''
        self.top_para    = ''
        self.word_title        = ''
        self.word_port         = []
        self.word_wire         = []
        self.word_body         = {}        
        self.endpoint          = {}
        self.endpoint['name']  = []
        self.endpoint['type']  = []
        self.endpoint['width'] = []
        self.endpoint['row']   = []
        
        self.list_sub_ref = []
        self.list_sub_num = []
        self.list_sub_ins = []
        self.list_sub_pin = []
        self.list_sub_net = []
        self.list_sub_row = []
        
        self.mode_clock   = 'pos'
        self.signal_clock = 'clk'
        self.mode_rstn    = 'neg'
        self.signal_rstn  = 'rst_n' 
    
    def __init__(self,sheet):
        self.initial()
        self.sheet = sheet
        if sheet.title[0] not in ['#','@','`']:
            self.file_name = sheet.title
            self.module_name  = sheet.title.split(default_split_char)[-1]
            self.do_process_rtl()
    
    def get_cell_val(self,value):
        return str(value).strip() if value else ''
    
    def get_col(self,col):
        return [self.get_cell_val(self.sheet.cell(row+1,col+1).value) for row in range(self.sheet.max_row)]
        
    def get_row(self,row):
        return [self.get_cell_val(self.sheet.cell(row+1,col+1).value) for col in range(self.sheet.max_column)]
    
    def get_cell(self,row,col):
        return self.get_cell_val(self.sheet.cell(row+1,col+1).value)
    
    def do_process_rtl(self):
        self.do_first_column()
        self.do_other_column()  
    
    def do_first_column(self):
        port_char = ' '
        first_column = self.get_col(0)
        for i,each in enumerate(first_column):
            self.endpoint['name'].append('')
            self.endpoint['type'].append('')
            self.endpoint['width'].append('')
            self.endpoint['row'].append(0)
            if i==0:
                if first_column[1] and first_column[1][0]=='#':
                    each += ' '+first_column[1] 
                self.top_name = each.lstrip('@').split()[0]
                self.top_para = each[each.index('#'):each.rindex(')')+1] if '#' in each else ''  
                self.word_title = 'module %s %s\n' % (self.top_name,self.top_para)           
            elif each:
                if i==1 and each[0]=='#':
                    continue
                if each[0]=='`' or ( len(each)>=2 and each[0:2]=='//'):
                    self.word_port.append(each)
                    self.word_wire.append(each)
                    continue
                if '#' in each:
                    name,type,width,pval,debug = self.get_name_type_width_pval(each)
                    self.endpoint['name'][i]   = name
                    self.endpoint['type'][i]   = type
                    self.endpoint['width'][i]  = width
                    self.endpoint['row'][i]    = i
                    if type in ['input','output','inout']:
                        self.word_port.append(self.get_port_format(port_char,type,width,name)) 
                        port_char = ',' if port_char==' ' else port_char
                    elif type in ['wire','reg']:
                        self.word_wire.append(self.get_wire_format(type,width,name))
                    else:
                        self.word_wire.append(self.get_parameter_format(type,width,name,pval))              
    
    def get_name_type_width_pval(self,value):
        debug = value[0]=='*'
        name  = value.lstrip('*').split()[0]
        type  = 'input'     if '#i' in value else ''
        type += 'output'    if '#o' in value else '' 
        type += 'inout'     if '#x' in value else ''
        type += 'wire'      if '#w' in value else ''
        type += 'reg'       if '#r' in value else ''
        type += 'parameter' if '#p' in value else ''
        width = value[value.find('#')+2:].strip()
        pval  = ''
        if '#p' in value:
            if len(width.split())==1:
                pval  = width if width else '0'
                width = ''
            else:
                pval  = ' '.join(width.split()[1:])
                width = width.split()[0]
        width = '[(%s)-1:0]' % width if width and '[' not in width else width
        return name,type,width,pval,debug
    
    def do_other_column(self):
        for col_cnt,col_val in enumerate(self.get_row(0)):
            self.word_body[col_cnt]              = {}
            self.word_body[col_cnt]['submodule'] = ''
            self.word_body[col_cnt]['word']      = []
            if col_cnt:
                if col_val:
                    self.do_module(col_cnt)
                else:
                    self.do_content(col_cnt)
    
    def do_module(self,col_cnt):
        cell0 = self.get_row(0)[col_cnt]
        cell1 = self.get_row(1)[col_cnt]
        if cell1 and cell1.startswith('#'):
            module_info = cell0.split()[0] + ' ' + cell1 + ' ' + ( '' if len(cell0.split())==1 else cell0.split()[-1] )
        else:
            module_info = cell0
        ref_name = module_info.split()[0]
        ins_name = ref_name if len(module_info.split())==1 or module_info[-1]==')' else  module_info.split()[-1]
        if ins_name in self.list_sub_ins:
            ins_cnt = 0
            while ins_name+'_'+str(ins_cnt) in self.list_sub_ins:
                ins_cnt += 1
            ins_name = ins_name+'_'+str(ins_cnt)
        para = module_info[module_info.index('#'):module_info.rindex(')')+1] if '#' in module_info else ''
        self.list_sub_ref.append(ref_name)
        self.list_sub_num.append(col_cnt)
        self.list_sub_ins.append(ins_name)
        list_sub_pin = []
        list_sub_net = []
        list_sub_row = []
        pin_char = ' '
        self.word_body[col_cnt]['submodule'] = '%s %s %s (' %(ref_name,para,ins_name)
        for i,each in enumerate(self.get_col(col_cnt)):
            if i and each and not '#p' in each:
                if (i==1 and each[0]=='#') or (each.strip().startswith('//')):
                    continue
                if each[0]=='`':
                    self.word_body[col_cnt]['word'].append(each)
                    continue
                sub_pin = each.split()[0]    
                if '#' in each:
                    sub_net = self.endpoint['name'][i]
                elif len(each)==1:
                    sub_net = ''
                else:
                    sub_net = ' '.join(each.split()[1:])
                    if sub_net:
                        sub_net = sub_pin+sub_net[1:] if sub_net[0]=='.' else self.endpoint['name'][i]+sub_net[1:] if sub_net[0]=='-' else sub_net
                list_sub_pin.append(sub_pin)
                list_sub_net.append(sub_net)
                list_sub_row.append(i)
                self.word_body[col_cnt]['word'].append(self.get_instance_format(pin_char,sub_pin,sub_net))
                pin_char = ',' if pin_char==' ' else pin_char
        self.list_sub_pin.append(list_sub_pin)
        self.list_sub_net.append(list_sub_net)
        self.list_sub_row.append(list_sub_row)
        
    def do_content(self,col_cnt):
        empty_column = True
        for i,each in enumerate(self.get_col(col_cnt)):
            if i and each:
                empty_column = False
                if each[0]=='#' and self.endpoint['name'][i] and self.endpoint['type'][i]=='reg':
                    each_split = each.split()
                    def_val = each_split[0][1:] if len(each_split[0])>1 else '0'
                    other = ' '.join(each_split[1:])
                    other = self.endpoint['name'][i]+other if other.startswith('<=') else other
                    self.word_body[col_cnt]['word'].append(self.get_ff_title(self.endpoint['name'][i],def_val)+other+'\n')
                elif each[0]=='#' and self.endpoint['name'][i] and self.endpoint['type'][i] in ['wire','output','inout']:
                    each_split = each[1:].split()
                    line = self.endpoint['name'][i]+' '+each[1:] if each_split[0][0]=='=' else each[1:]
                    line = 'assign ' + line
                    self.word_body[col_cnt]['word'].append(line+'\n')
                elif each[0]=='#' and not self.endpoint['name'][i]:
                    self.mode_clock,self.signal_clock = each[1:].split()[0:2]
                    if len(each[1:].split())>=4:
                        self.mode_rstn,self.signal_rstn = each[1:].split()[2:4]
                    else:
                        self.mode_rstn,self.signal_rstn = ['','']
                else:
                    self.word_body[col_cnt]['word'].append(each)
        if not empty_column:
            self.word_body[col_cnt]['word'].append('\n')
        
    def get_ff_title(self,signal,value):
        title  = 'always @ ( %sedge %s ' %(self.mode_clock,self.signal_clock)
        title += 'or %sedge %s )\n' %(self.mode_rstn,self.signal_rstn) if self.mode_rstn else ')\n'
        if self.mode_rstn:
            negtive_char = '~' if self.mode_rstn=='neg' else ''
            title += 'if (%s%s)\n    %s <= %s;\nelse\n    ' %(negtive_char,self.signal_rstn,signal,value)
        return title
    
    def get_ljust_string(self,*string):
        next_continue = False
        line = ''
        for i in range(len(string)):
            if next_continue:
                next_continue = False
                continue
            curr          = string[i]
            next          = string[i+1] if i<(len(string)-1) else ''
            length        = next if isinstance(next,int) else 25
            next_continue = True if isinstance(next,int) else False
            line += curr.ljust(length,' ')
        return line
    
    def get_port_format(self,port_char,type,width,name):
        return '    '+self.get_ljust_string(port_char+type,10,width,name)
        
    def get_wire_format(self,type,width,name):
        return '    '+self.get_ljust_string(type,10,width,name+';')
        
    def get_parameter_format(self,type,width,name,pval):
        return '    '+self.get_ljust_string(type,10,width,name+' = '+pval+';')
    
    def get_instance_format(self,pin_char,sub_pin,sub_net):
        return '    '+self.get_ljust_string(pin_char+'.'+sub_pin,40,'(    '+sub_net,40,')',1)

###################################################################################
##Class: ClassHyperlink
###################################################################################
class ClassHyperlink:
    '''All valid sheets'''
    class_sheets              = {}
    ip                        = {}
    ip['column']              = {}
    ip['coordinate']          = {}
    ip['pin']                 = {}
    module                    = {}
    module['coordinate']      = {}
    module['pin']             = {}
    submodule                 = {}
    submodule['column']       = {}
    submodule['coordinate']   = {}
    module_miss               = {}
    module_miss['ip']         = []
    module_miss['module']     = []
    module_miss['submodule']  = []
    xlsx_file                 = ''                                  
    def __init__(self,argv):
        self.xlsx_file = os.path.basename(argv.xlsx_file)                 
        print('---Build every sheet...')
        self.do_build_class_sheet(argv)
        print('---Add hyperlink for each sheet...')
        self.do_add_hyperlink()
        #self.do_check_pins()
        print('---To adjust each sheet...')
        self.do_adjust_sheet()
        print('---To save xlsx file...')
        self.wb.save(argv.xlsx_file)
        self.wb.close()         
    
    def do_adjust_sheet(self):
        #from openpyxl.styles import Border,Side
        #border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        for each in self.wb.sheetnames:
            if not any([each.strip().startswith(i) for i in ['#','@','`']]) or each.strip()=='@module':
                self.do_set_sheet_width(self.wb[each])
                self.do_set_conditional_format(self.wb[each])            
            #if not any([each.strip().startswith(i) for i in ['#','@','`']]):
            #    for row in range(self.wb[each].max_row):
            #        for col in range(self.wb[each].max_column):
            #            self.wb[each].cell(row+1,col+1).border = border
    
    def do_set_conditional_format(self,ws):
        from openpyxl.styles import PatternFill,Font,Border,Side
        from openpyxl.styles.differential import DifferentialStyle        
        from openpyxl.formatting.rule import Rule
        #border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        font = Font(color='000000')
        field = ws.cell(1,1).coordinate+':'+ws.cell(ws.max_row+4,ws.max_column+4).coordinate
        #for i in range(1,ws.max_row+1):
        #    for j in range(1,ws.max_column+1):
        #        ws.cell(i,j).border = border
        fill_color = {'//':'F8CBAD','#x':'FFE699','#o':'B4C6E7','#i':'C6E0B4'}
        for char,color in fill_color.items():
            dxf  = DifferentialStyle(font=font ,fill=PatternFill(bgColor=color))#,border=border)
            rule = Rule(type='containsText', operator='containsText', formula=['NOT(ISERROR(SEARCH("%s",A1)))'%(char)], text=char, dxf=dxf)
            ws.conditional_formatting.add(field,rule)
        
    def do_set_sheet_width(self,ws):
        for i in range(1,ws.max_row+1+4):
            ws.row_dimensions[i].height = 15
        for i in range(1,ws.max_column+1+4):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 25 if i!=2 else 30        
    
    def do_check_pins(self):
        for sheet_name in self.class_sheets:
            sheet = self.class_sheets[sheet_name]
            for i,ref in enumerate(sheet.list_sub_ref):
                ins_port = sheet.list_sub_pin[i]
                get  = any([ref in self.ip['pin'].keys(),ref in self.module['pin'].keys()])
                ref_port = (self.ip['pin'][ref] if ref in self.ip['pin'].keys() else [])+(self.module['pin'][ref] if ref in self.module['pin'].keys() else [])
                ins_miss = [each for each in ins_port if each not in ref_port]
                ref_miss = [each for each in ref_port if each not in ins_port]
                if not get:
                    print('---Module:',sheet.top_name.ljust(20,' '),'---sub:',ref.ljust(20,' '),' have no reference')
                elif ins_miss:
                    print('---Module:',sheet.top_name.ljust(20,' '),'---sub:',ref.ljust(20,' '),ins_miss,' exist in ref, not in ins')
                elif ref_miss:
                    print('---Module:',sheet.top_name.ljust(20,' '),'---sub:',ref.ljust(20,' '),ref_miss,' exist in ins, not in ref')
    
    def do_build_class_sheet(self,argv):
        self.wb = openpyxl.load_workbook(argv.xlsx_file)
        for each in self.wb.sheetnames:
            if each=='@module':
                for col in range(self.wb[each].max_column):
                    cell = self.wb[each].cell(1,col+1).value
                    if cell:
                        name = cell.strip().lstrip('@')
                        link = "#'"+each+"'!"+self.wb[each].cell(1,col+1).coordinate
                        self.ip['coordinate'][name] = link
                        self.ip['column'][name]     = col
                        self.ip['pin'][name]        = []
                        for row in range(1,self.wb[each].max_row):
                            one = self.wb[each].cell(row+1,col+1).value
                            if row==1 and one and one.strip().startswith('#'):
                                continue
                            if one and not any([one.strip().startswith(i) for i in ['//','`']]) and not '#p' in one:
                                self.ip['pin'][name].append(one.strip().split()[0])
            if each[0] not in ['#','@','`']:
                self.class_sheets[each] = ClassRtl(self.wb[each]) 
                top_name = self.class_sheets[each].top_name
                self.module['coordinate'][top_name]    = '#'+each+'!'+self.wb[each].cell(1,1).coordinate
                self.module['pin'][top_name] = []
                for i,type in enumerate(self.class_sheets[each].endpoint['type']):
                    if type in ['input','output','inout']:
                        self.module['pin'][top_name].append(self.class_sheets[each].endpoint['name'][i])
                self.submodule['column'][top_name]     = {}
                self.submodule['coordinate'][top_name] = {}  
                for sub in self.class_sheets[each].list_sub_ref:
                    self.submodule['column'][top_name][sub]     = []
                    self.submodule['coordinate'][top_name][sub] = []
                for i,sub in enumerate(self.class_sheets[each].list_sub_ref):
                    column     = self.class_sheets[each].list_sub_num[i]  
                    ##coordinate = '#'+each+'!'+self.wb[each].cell(1,column+1).coordinate
                    coordinate = '#'+each+'!'+self.wb[each].cell(1,1).coordinate
                    self.submodule['column'][top_name][sub].append(column)
                    self.submodule['coordinate'][top_name][sub].append(coordinate)                   
        #wb.save(argv.xlsx_file)
        #wb.close() 
    
    def do_add_hyperlink(self):
        self.do_add_ip_link()    
        for sheet_name in self.class_sheets:
            self.do_add_top_link(sheet_name) 
            self.do_add_sub_link(sheet_name)
        
    def do_add_ip_link(self):
        sheet_name = '@module'
        if sheet_name not in self.wb:
            return
        for i in range(self.wb[sheet_name].max_column):
            cell = self.wb[sheet_name].cell(1,i+1).value
            if cell:
                name = cell.lstrip('@')
                top_name,coordinate = self.get_module_instance(name)
                if not top_name:
                    self.module_miss['ip'].append(name)
                self.do_remove_cell_link(self.wb[sheet_name].cell(1,i+1))
                self.do_remove_extra_link(sheet_name,i+1)
                if len(top_name):
                    self.do_add_cell_link(self.wb[sheet_name].cell(1,i+1),''.join(coordinate[0]))
                if len(top_name)>1:
                    cell = self.wb[sheet_name].cell(2,i+1).value
                    line_num = 2 if cell and cell.strip().startswith('#') else 1  
                    for num,each in enumerate(top_name):
                        cell = self.wb[sheet_name].cell(line_num+1,i+1).value
                        if cell and not cell.startswith('//'):
                            self.wb[sheet_name].insert_rows(line_num+1)
                        self.do_remove_cell_link(self.wb[sheet_name].cell(line_num+1,i+1))
                        self.wb[sheet_name].cell(line_num+1,i+1).value = '//'+each
                        self.do_add_cell_link(self.wb[sheet_name].cell(line_num+1,i+1),''.join(coordinate[num]))
                        line_num = line_num + 1                 
        
    def do_add_sub_link(self,sheet_name):
        for i,each in enumerate(self.class_sheets[sheet_name].list_sub_ref):
            col = self.class_sheets[sheet_name].list_sub_num[i]
            self.do_remove_cell_link(self.wb[sheet_name].cell(1,col+1))
            if each in self.module['coordinate']:
                self.do_add_cell_link(self.wb[sheet_name].cell(1,col+1),self.module['coordinate'][each])
            if each in self.ip['coordinate']:
                self.do_add_cell_link(self.wb[sheet_name].cell(1,col+1),self.ip['coordinate'][each])
            if not any([each in self.module['coordinate'],each in self.ip['coordinate']]):
                self.module_miss['submodule'].append(self.class_sheets[sheet_name].top_name+' '+each)
    
    def do_add_top_link(self,sheet_name):
        top_name,coordinate = self.get_module_instance(self.class_sheets[sheet_name].top_name)
        self.do_remove_cell_link(self.wb[sheet_name].cell(1,1))
        self.do_remove_extra_link(sheet_name,1)
        if not top_name:
            self.module_miss['module'].append(self.class_sheets[sheet_name].top_name)
        if len(top_name):
            self.do_add_cell_link(self.wb[sheet_name].cell(1,1),''.join(coordinate[0]))
        if len(top_name)>1:
            cell = self.wb[sheet_name].cell(2,1).value
            line_num = 2 if cell and cell.strip().startswith('#') else 1  
            for i,each in enumerate(top_name):
                cell = self.wb[sheet_name].cell(line_num+1,1).value
                if cell and not cell.startswith('//'):
                    self.wb[sheet_name].insert_rows(line_num+1)
                    cell = self.wb[sheet_name].cell(line_num+1,1).value
                self.do_remove_cell_link(self.wb[sheet_name].cell(line_num+1,1))
                self.wb[sheet_name].cell(line_num+1,1).value = '//'+each
                self.do_add_cell_link(self.wb[sheet_name].cell(line_num+1,1),''.join(coordinate[i]))
                line_num = line_num + 1    
    
    def do_remove_extra_link(self,sheet_name,column):
        sheet = self.wb[sheet_name]
        line_num = 2 if sheet.cell(2,column).value and sheet.cell(2,column).value.strip().startswith('#') else 1
        while sheet.cell(line_num+1,column).value and sheet.cell(line_num+1,column).value.strip().startswith('//') and sheet.cell(line_num+1,column).hyperlink:
            sheet.cell(line_num+1,column).value = None
            self.do_remove_cell_link(sheet.cell(line_num+1,column))
            line_num  += 1  
    
    def do_remove_cell_link(self,cell):
        if cell.hyperlink:
            cell.style = 'Normal'
            cell.hyperlink = None  
    
    def do_add_cell_link(self,cell,url):
        cell.style = 'Hyperlink'
        cell.hyperlink = self.xlsx_file+url
        cell.font  = openpyxl.styles.Font(name='Verdana',size=8,color='0563C1') #,underline='single'
    
    def get_module_instance(self,name):
        name = name.strip()
        top_name   = []
        coordinate = []
        for each in self.submodule['coordinate']:
            if name in self.submodule['coordinate'][each]:
                for one in self.submodule['coordinate'][each][name]:
                    top_name.append(each)
                    coordinate.append(one)
        return top_name,coordinate

###################################################################################
##Main
###################################################################################
#if __name__ == '__main__': 
def main():
    class_all = ClassHyperlink(SysArgv())