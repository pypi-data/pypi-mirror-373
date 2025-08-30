##/////////////////////////////////////////////////////////////////////////////////////
##//
##//Copyright 2025  Matrix HDL
##//
##//Licensed under the Apache License, Version 2.0 (the "License");
##//you may not use this file except in compliance with the License.
##//You may obtain a copy of the License at
##//
##//    http://www.apache.org/licenses/LICENSE-2.0
##//
##//Unless required by applicable law or agreed to in writing, software
##//distributed under the License is distributed on an "AS IS" BASIS,
##//WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
##//See the License for the specific language governing permissions and
##//limitations under the License.
##//
##/////////////////////////////////////////////////////////////////////////////////////

###################################################################################
##import
###################################################################################
import openpyxl    ##openpyxl.load_workbook()
import os          ##os.path.exists()/os.makedirs()/os.getcwd()
import shutil      ##shutil.rmtree()
###################################################################################
##class:SysArgv
###################################################################################
class SysArgv:
    ''' To analyze sys.argv and generate parameters. '''
    from sys import argv
    xlsx_file = 'verilog.xlsx' 
    rtl_files = []    

    def __init__(self):
        del self.argv[0]
        given_para_title = ['-o','-f','-v']
        given_para_value = [self.get_extract_name(i) for i in given_para_title]
        oname = given_para_value[0]
        fname = given_para_value[1]
        vname = given_para_value[2]
        files = [i for i in self.argv if i not in given_para_title+given_para_value]
        if vname:
            for root,folder_names,file_names in os.walk(vname):
                for file_name in file_names:
                    if file_name.endswith('.v'):
                        files.append(root+file_name)
        if fname:
            with open(fname) as f:
                other = f.read().strip().split()
            for i in other:
                if not any(i.startswith(each) for each in ['+','-','//']):
                    files.append(i)
        self.xlsx_file = oname if oname else self.xlsx_file
        self.rtl_files = files
        if not files:
            print("---No input .v files,exit...")
            exit()
        
    def get_extract_name(self,char):
        if char in self.argv:
            if self.argv.index(char)<((len(self.argv))-1) and not self.argv[self.argv.index(char)+1].startswith('-'):
                value = self.argv[self.argv.index(char)+1]
            else:
                value = True
        else:
            value = False
        return value
            
    def do_make_path(self,path):
        if os.path.exists(path):
            shutil.rmtree(path)
        os.makedirs(path)
        
    def do_check_path(self,path):
        if not os.path.exists(path):
            os.makedirs(path)    

###################################################################################
##class:VFiles
###################################################################################
class VFiles:
    '''The class of one .v file.'''
    module_name = []
    module_para = []
    port        = {}
    xlsx_file   = ''
    def __init__(self,argv):
        self.xlsx_file = argv.xlsx_file
        string = ''
        for file in argv.rtl_files:
            with open(file) as f:
                string += f.read()
        string = self.get_rm_comment(string)
        self.do_import_ports(string)
        self.do_write_xlsx()
        
    def do_write_xlsx(self):
        if (not os.path.exists(self.xlsx_file)):
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(self.xlsx_file)
        sheet_names = wb.sheetnames
        template_name = [i for i in sheet_names if i.strip().endswith('template')]
        if '@module' in sheet_names:
            del wb['@module']
            ws = wb.create_sheet('@module')        
        elif template_name:
            ws = wb.copy_worksheet(wb[template_name[0]])
            ws.title = '@module'
        else:
            ws = wb.create_sheet('@module')  
        for i in range(len(self.port)):
            ws.cell(column=i+1,row=1).value = self.module_name[i]
            ws.cell(column=i+1,row=1).font  = openpyxl.styles.Font(name='Verdana',size=8)
            ws.cell(column=i+1,row=2).value = self.module_para[i]
            ws.cell(column=i+1,row=2).font  = openpyxl.styles.Font(name='Verdana',size=8)
            for j in range(len(self.port[self.module_name[i]])):
                ws.cell(column=i+1,row=j+3).value = self.port[self.module_name[i]][j]
                ws.cell(column=i+1,row=j+3).font  = openpyxl.styles.Font(name='Verdana',size=8)
        self.do_set_sheet_width(ws) 
        self.do_set_conditional_format(ws)        
        wb.save(self.xlsx_file)
        wb.close()
    
    def do_set_conditional_format(self,ws):
        from openpyxl.styles import PatternFill,Font,Border,Side
        from openpyxl.styles.differential import DifferentialStyle        
        from openpyxl.formatting.rule import Rule
        border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        font = Font(color='000000')
        field = ws.cell(1,1).coordinate+':'+ws.cell(ws.max_row*2,ws.max_column*2).coordinate
        for i in range(1,ws.max_row+1):
            for j in range(1,ws.max_column+1):
                ws.cell(i,j).border = border
        fill_color = {'//':'F8CBAD','#x':'FFE699','#o':'B4C6E7','#i':'C6E0B4'}
        for char,color in fill_color.items():
            dxf  = DifferentialStyle(font=font ,fill=PatternFill(bgColor=color),border=border)
            rule = Rule(type='containsText', operator='containsText', formula=['NOT(ISERROR(SEARCH("%s",A1)))'%(char)], text=char, dxf=dxf)
            ws.conditional_formatting.add(field,rule)
        
    def do_set_sheet_width(self,ws):
        for i in range(1,ws.max_row+1):
            ws.row_dimensions[i].height = 15
        for i in range(1,ws.max_column+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 25
     
    def do_import_ports(self,string):    
        while string.find('module ')!=-1:
            string = string[string.find('module ')+len('module'):]
            name   = string.split()[0].strip().rstrip('(')
            self.module_name.append(name)
            self.port[name] = []
            para,string,result = self.get_port_list(string)
            self.module_para.append('')
            self.port[name].extend(self.get_parameter_list(para))
            port_has_type   = any(i in result for i in ['input ','output ','inout ','[',']'])
            for line in result.splitlines():
                if not line.strip():
                    self.port[name].append('')
                elif line.strip().startswith('`'):
                    self.port[name].append(line.strip())
                else:
                    signals = line.split(',')
                    for each in signals:
                        each = each.strip()
                        if each:
                            for port_type in ['input','output','inout']:
                                if any(each.startswith(i) for i in [port_type,port_type+'[']):
                                    type = port_type
                                    each = each[len(port_type):]
                                    width = each[each.find('['):each.find(']')+1] if '[' in each else '' 
                                    each  = each[each.find(']')+1:] if ']' in each else each
                                    char  = 'i' if type=='input' else 'o' if type=='output' else 'x'
                            self.port[name].append(each.strip().split()[-1]+(' #'+char+' '+width if port_has_type else ''))
            nearest_char,nearest_get,nearest_pos = self.get_nearest_chars(string,['endmodule','module'])
            string_this = string[:nearest_pos]
            string = string[nearest_pos+len('endmodule'):] if nearest_char=='endmodule' else string[nearest_pos:]
            if not port_has_type:
                port_def = self.get_port_definition(string_this)
                for port_type in ['input','output','inout']:
                    for each in port_def[port_type]:
                        char    = 'i' if port_type=='input' else 'o' if port_type=='output' else 'x'
                        width   = each[each.find('['):each.find(']')+1] if '[' in each else ''
                        signals = each[each.find(']')+1:] if ']' in each else each
                        for signal in signals.split(','):
                            self.port[name][self.port[name].index(signal.strip())] += ' #'+char+' '+width
    
    def get_parameter_list(self,para):
        result = []
        if para:
            para   = para[para.index('#(')+2:para.rindex(')')]
            lines  = para.strip().split(',')
            width  = ''
            for line in lines:
                array = line.strip().split('=')
                if array[0].strip().startswith('parameter '):
                    if '[' in array[0] and ']' in array[0]:
                        width = array[0][array[0].index('['):array[0].index(']')+1]+' '
                    else:
                        width = ''
                name  = array[0].strip().split()[-1]
                val   = ''.join(array[1].strip().split())
                result.append(name+' #p '+width+val)
        return result   
    
    def get_nearest_chars(self,string,chars):
        pos = [string.find(char) if string.find(char)!=-1 else len(string) for char in chars]
        idx = pos.index(min(pos))
        return chars[idx],pos[idx]!=len(string),pos[idx]
    
    def get_port_definition(self,string):
        lines = string.split(';')
        port_def = {}
        port_def['input']  = []
        port_def['output'] = []
        port_def['inout']  = []
        for line in lines:
            one = line.strip()
            for port_type in ['input','output','inout']:
                if any(one.startswith(i) for i in [port_type,port_type+'[']):
                    port_def[port_type].append(one[len(port_type):])              
        return port_def
    
    def get_rm_comment(self,string):
        nearest_char,nearest_get,nearest_pos = self.get_nearest_chars(string,['/*','//'])
        while nearest_get:
            if nearest_char=='/*':
                string = self.get_rm_slash_star(string)
            else:
                string = self.get_rm_two_slashes(string)
            nearest_char,nearest_get,nearest_pos = self.get_nearest_chars(string,['/*','//'])
        return string
    
    def get_rm_slash_star(self,string):
        start = string.find('/*')
        end = start + string[start:].find('*/')
        string = string[:start]+string[end+2:]
        return string
     
    def get_rm_two_slashes(self,string):
        start = string.find('//')
        end   = start + string[start:].find('\n')
        string = string[:start]+string[end:]
        return string
        
    def get_port_list(self,string):
        num_pound = len(string) if string.find('#')==-1 else string.find('#')
        num_paren = string.find('(')
        para      = ''
        if num_pound < num_paren:
            string,result = self.get_open_close_paren(string)
            para = '#( '+result+' )'
        string,result = self.get_open_close_paren(string)
        return para,string,result                
    
    def get_open_close_paren(self,string):
        result    = ''
        string    = string[string.find('(')+1:]
        cnt       = 1
        while cnt:
           open  = len(string) if string.find('(')==-1 else string.find('(')
           close = len(string) if string.find(')')==-1 else string.find(')')
           if open<close:
               result += string[:open+1]
               string  = string[open+1:]
               cnt += 1
           else:
               result += string[:close+1]
               string  = string[close+1:]
               cnt -= 1
        return string,result.rstrip(')')

###################################################################################
##Main
###################################################################################
#if __name__ == '__main__': 
def main():
    one_file = VFiles(SysArgv())   