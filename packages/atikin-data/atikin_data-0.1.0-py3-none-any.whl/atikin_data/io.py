import csv
import json
import os
from typing import List, Dict, Any, Optional, Union
from .frame import AtikinDataFrame


def read_csv(path: str, limit: Optional[int] = None, encoding: str = "utf-8") -> AtikinDataFrame:
    """
    Read CSV into AtikinDataFrame (streaming, memory-friendly for moderate files).
    """
    with open(path, newline="", encoding=encoding) as f:
        reader = csv.DictReader(f)
        data: List[Dict[str, Any]] = []
        for i, row in enumerate(reader):
            # normalize empty string to None
            normalized = {k: (v if v != "" else None) for k, v in row.items()}
            data.append(normalized)
            if limit and (i + 1) >= limit:
                break
    return AtikinDataFrame(data)


def read_json(path: str, limit: Optional[int] = None, encoding: str = "utf-8") -> AtikinDataFrame:
    """
    Read JSON — supports either a JSON array of objects or newline-delimited JSON (ndjson).
    """
    data: List[Dict[str, Any]] = []
    with open(path, "r", encoding=encoding) as f:
        # Detect if file starts with '[' -> JSON array
        first = f.read(1)
        f.seek(0)
        if first == "[":
            # load full array (ok for small/medium files)
            arr = json.load(f)
            for i, item in enumerate(arr):
                if isinstance(item, dict):
                    data.append(item)
                if limit and (i + 1) >= limit:
                    break
        else:
            # try ndjson (one JSON object per line)
            for i, line in enumerate(f):
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    if isinstance(obj, dict):
                        data.append(obj)
                except json.JSONDecodeError:
                    continue
                if limit and (i + 1) >= limit:
                    break
    return AtikinDataFrame(data)


def read_excel(path: str, sheet: Union[str, int] = 0, limit: Optional[int] = None) -> AtikinDataFrame:
    """
    Read Excel using openpyxl (if installed). If not installed, raises ImportError.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl is required to read Excel files. Install with `pip install 'atikin-data[excel]'`") from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # choose sheet
    if isinstance(sheet, int):
        sheets = wb.sheetnames
        ws = wb[sheets[sheet]]
    else:
        ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return AtikinDataFrame([])
    header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)]
    data: List[Dict[str, Any]] = []
    for i, row in enumerate(it):
        row_dict = {header[j]: row[j] if j < len(row) else None for j in range(len(header))}
        data.append(row_dict)
        if limit and (i + 1) >= limit:
            break
    return AtikinDataFrame(data)


def to_pandas_if_available(df: AtikinDataFrame):
    """
    If pandas is installed, convert AtikinDataFrame to pandas.DataFrame else return AtikinDataFrame.
    """
    try:
        import pandas as pd
    except Exception:
        return df
    return pd.DataFrame(df.to_dict())


def quick_summary(path: str, encoding: str = "utf-8") -> Dict[str, int]:
    """
    Fast summary: returns dict with rows, columns, missing count.
    For CSV we stream row-by-row (memory-friendly).
    For JSON/Excel we attempt reasonable approaches (Excel needs openpyxl).
    """
    _, ext = os.path.splitext(path)
    ext = ext.lower()

    if ext == ".csv":
        with open(path, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f)
            try:
                header = next(reader)
            except StopIteration:
                return {"rows": 0, "columns": 0, "missing": 0}
            cols = len(header)
            missing = [0] * cols
            rows = 0
            for row in reader:
                rows += 1
                for i in range(cols):
                    if i >= len(row) or (row[i].strip() == ""):
                        missing[i] += 1
        return {"rows": rows, "columns": cols, "missing": sum(missing)}

    elif ext in (".json", ".ndjson"):
        # simple approach: count objects and columns from first object
        rows = 0
        columns = 0
        missing = 0
        with open(path, "r", encoding=encoding) as f:
            first = f.read(1)
            f.seek(0)
            if first == "[":
                # load array (careful with large files)
                try:
                    arr = json.load(f)
                except Exception:
                    return {"rows": 0, "columns": 0, "missing": 0}
                rows = len(arr)
                if rows:
                    all_keys = set().union(*(obj.keys() for obj in arr if isinstance(obj, dict)))
                    columns = len(all_keys)
                    # count missing by brute force
                    for obj in arr:
                        for k in all_keys:
                            if k not in obj or obj[k] in (None, ""):
                                missing += 1
                return {"rows": rows, "columns": columns, "missing": missing}
            else:
                # ndjson streaming
                all_keys = set()
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        obj = json.loads(line)
                    except Exception:
                        continue
                    if isinstance(obj, dict):
                        rows += 1
                        all_keys.update(obj.keys())
                # second pass to count missing would be expensive; approximate missing as 0
                columns = len(all_keys)
                return {"rows": rows, "columns": columns, "missing": 0}

    elif ext in (".xls", ".xlsx"):
        try:
            import openpyxl
        except Exception:
            raise ImportError("openpyxl required for Excel quick_summary; install 'atikin-data[excel]'")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            return {"rows": 0, "columns": 0, "missing": 0}
        cols = len(header)
        rows = 0
        missing = 0
        for row in it:
            rows += 1
            for i in range(cols):
                if i >= len(row) or row[i] is None or (isinstance(row[i], str) and row[i].strip() == ""):
                    missing += 1
        return {"rows": rows, "columns": cols, "missing": missing}

    else:
        raise ValueError("Unsupported file extension for quick_summary. Supported: .csv, .json, .ndjson, .xls, .xlsx")
