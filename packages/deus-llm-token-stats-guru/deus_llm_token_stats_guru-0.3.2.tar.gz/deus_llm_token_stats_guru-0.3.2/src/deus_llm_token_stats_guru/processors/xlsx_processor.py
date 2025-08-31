"""XLSX (Excel) file processor."""

from pathlib import Path
from typing import Set, List
import logging

from .base import BaseFileProcessor
from ..types import CountResult
from .._exceptions import FileProcessingError

logger = logging.getLogger(__name__)


class XLSXProcessor(BaseFileProcessor):
    """Processor for Excel XLSX files."""
    
    @property
    def supported_extensions(self) -> Set[str]:
        """Return supported file extensions."""
        return {".xlsx", ".xls"}
    
    @property
    def processor_name(self) -> str:
        """Return processor name."""
        return "Excel"
    
    def extract_text(self, file_path: Path) -> List[str]:
        """Extract text content from XLSX file.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            List of text strings from all worksheets
            
        Raises:
            FileProcessingError: If XLSX cannot be read
        """
        try:
            # Try openpyxl first (most comprehensive)
            try:
                import openpyxl
                return self._extract_with_openpyxl(file_path)
            except ImportError:
                logger.debug("openpyxl not available, trying pandas")
                pass
            
            # Fallback to pandas
            try:
                import pandas as pd
                return self._extract_with_pandas(file_path)
            except ImportError:
                logger.debug("pandas not available, trying xlrd")
                pass
            
            # Final fallback to xlrd for older .xls files
            try:
                import xlrd
                return self._extract_with_xlrd(file_path)
            except ImportError:
                pass
            
            raise FileProcessingError(
                f"No XLSX/XLS processing libraries available. Install: pip install openpyxl xlrd"
            )
            
        except Exception as e:
            logger.warning(f"⚠️  Failed to extract text from {file_path}: {e}")
            return []
    
    def _extract_with_openpyxl(self, file_path: Path) -> List[str]:
        """Extract text using openpyxl library."""
        import openpyxl
        
        text_content = []
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = []
            
            # Add sheet name as context
            sheet_text.append(f"Sheet: {sheet_name}")
            
            # Extract cell values
            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        cell_str = str(cell.value).strip()
                        # Filter out dates, numbers-only, and empty values
                        if (cell_str and 
                            len(cell_str) > 1 and
                            not cell_str.replace('.', '').replace('-', '').replace('/', '').isdigit() and
                            cell_str.lower() not in {'nan', 'null', 'none'}):
                            row_values.append(cell_str)
                
                if row_values:
                    sheet_text.append(" ".join(row_values))
            
            if len(sheet_text) > 1:  # More than just sheet name
                text_content.extend(sheet_text)
        
        return text_content
    
    def _extract_with_pandas(self, file_path: Path) -> List[str]:
        """Extract text using pandas library."""
        import pandas as pd
        
        text_content = []
        
        try:
            # Read all sheets
            excel_data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            
            for sheet_name, df in excel_data.items():
                sheet_text = [f"Sheet: {sheet_name}"]
                
                # Add column headers
                if not df.columns.empty:
                    header_text = " ".join(str(col) for col in df.columns if str(col) != "nan")
                    if header_text.strip():
                        sheet_text.append(header_text)
                
                # Add cell values
                for column in df.columns:
                    try:
                        column_values = df[column].astype(str).fillna("").tolist()
                        meaningful_values = [
                            val for val in column_values 
                            if val and val.lower() not in {'nan', 'null', 'none'} and len(val) > 1
                        ]
                        if meaningful_values:
                            sheet_text.append(" ".join(meaningful_values))
                    except Exception as e:
                        logger.debug(f"⚠️  Skipping problematic column {column}: {e}")
                        continue
                
                if len(sheet_text) > 1:
                    text_content.extend(sheet_text)
        
        except Exception as e:
            logger.debug(f"⚠️  Pandas extraction failed: {e}")
            # Try reading just the first sheet
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                for column in df.columns:
                    column_values = df[column].astype(str).fillna("").tolist()
                    meaningful_values = [
                        val for val in column_values 
                        if val and val.lower() not in {'nan', 'null', 'none'}
                    ]
                    if meaningful_values:
                        text_content.append(" ".join(meaningful_values))
            except Exception:
                pass
        
        return text_content
    
    def _extract_with_xlrd(self, file_path: Path) -> List[str]:
        """Extract text using xlrd library (for older .xls files)."""
        import xlrd
        
        text_content = []
        workbook = xlrd.open_workbook(file_path)
        
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            sheet_text = [f"Sheet: {sheet.name}"]
            
            for row_idx in range(sheet.nrows):
                row_values = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.value:
                        cell_str = str(cell.value).strip()
                        if (cell_str and 
                            len(cell_str) > 1 and
                            cell_str.lower() not in {'nan', 'null', 'none'}):
                            row_values.append(cell_str)
                
                if row_values:
                    sheet_text.append(" ".join(row_values))
            
            if len(sheet_text) > 1:
                text_content.extend(sheet_text)
        
        return text_content
    
    def process_file(self, file_path: Path) -> CountResult:
        """Process XLSX file with enhanced metadata.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            CountResult with Excel-specific statistics
        """
        logger.info(f"📊 Processing Excel file: {file_path.name}")
        
        # Extract text content
        text_content = self.extract_text(file_path)
        total_tokens = sum(self.count_tokens_in_text(text) for text in text_content)
        
        # Get basic file statistics
        file_size = file_path.stat().st_size
        
        # Estimate rows and sheets from text content
        row_count = len([line for line in text_content if not line.startswith("Sheet:")])
        sheet_count = len([line for line in text_content if line.startswith("Sheet:")])
        
        result: CountResult = {
            "file_path": str(file_path),
            "total_tokens": total_tokens,
            "row_count": row_count,
            "column_count": sheet_count,  # Use sheet count as "columns" for Excel
            "encoding_model": self.encoding_model,
            "file_size_bytes": file_size,
        }
        
        logger.info(f"✅ Processed {file_path.name}: {total_tokens:,} tokens, ~{row_count} rows, {sheet_count} sheets")
        return result