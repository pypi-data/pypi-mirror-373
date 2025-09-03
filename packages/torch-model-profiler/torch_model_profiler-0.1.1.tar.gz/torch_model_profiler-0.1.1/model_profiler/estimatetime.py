import openpyxl
from openpyxl.styles import PatternFill

def estimate_inference_time(layer_stats, 
                            compute_tops=100,    # 硬體算力 (TOPS)
                            mem_bw_gbs=200,      # DRAM 帶寬 (GB/s)
                            sram_size_mb=16):    # SRAM 大小 (MB)
    """
    根據硬體規格，估算模型的推論時間 (秒)
    - layer_stats: 來自 profile_flops_and_memory_layername 的輸出
    - compute_tops: 硬體峰值算力 (以 TOPS 為單位)
    - mem_bw_gbs: 記憶體頻寬 (以 GB/s 為單位)
    - sram_size_mb: SRAM 容量 (MB)，超過的資料需走 DRAM
    """

    total_flops = sum(item[4] for item in layer_stats)  # flops
    total_mem   = sum(item[5] for item in layer_stats)  # memory

    # 換算單位
    compute_cap = compute_tops * 1e12  # TOPS → FLOPs/s
    mem_bw      = mem_bw_gbs * 1e9     # GB/s → Bytes/s
    sram_bytes  = sram_size_mb * 1024 * 1024

    # Compute-bound latency
    compute_time = total_flops / compute_cap

    # Memory-bound latency
    if total_mem <= sram_bytes:
        memory_time = total_mem / (mem_bw * 10)  # SRAM 假設比 DRAM 快 10 倍
    else:
        memory_time = total_mem / mem_bw

    latency = max(compute_time, memory_time)

    print(f"=== Inference Estimation ===")
    print(f"Total FLOPs: {total_flops/1e9:.6f} GFLOPs")
    print(f"Total Memory: {total_mem/1e6:.6f} MB")
    print(f"Compute Time: {compute_time*1e3:.6f} ms")
    print(f"Memory Time:  {memory_time*1e3:.6f} ms")
    print(f"Estimated Latency: {latency*1e3:.6f} ms")

    return latency



def export_profile_to_excel_withinferencetime(layer_stats, 
                                              filename="profile_report.xlsx",
                                              compute_tops=100, 
                                              mem_bw_gbs=200, 
                                              sram_size_mb=16):
    """
    將 profile_flops_and_memory_layername 的結果輸出成 Excel 檔案。
    - Memory-bound: 整列紅色背景
    - Compute-bound: 整列綠色背景
    - Balanced: 整列灰色背景
    - 額外新增 Statistics 分頁
    - 新增每層的 compute time / memory time / latency (ms)
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profile Report"

    # 標題列
    headers = ["Layer(Name)", "Input Shape", "Output Shape",
               "FLOPs (M)", "Memory (KB)", "FLOP/Byte", "Params (K)",
               "Tag", "Compute Time (ms)", "Memory Time (ms)", "Latency (ms)"]
    ws.append(headers)

    # 填色定義
    fill_mem = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_cmp = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_bal = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    # 硬體參數
    compute_cap = compute_tops * 1e12   # TOPS → FLOPs/s
    mem_bw      = mem_bw_gbs * 1e9      # GB/s → Bytes/s
    sram_bytes  = sram_size_mb * 1024 * 1024

    # 統計累加器
    summary = {
        "Memory-bound ❗": {"count": 0, "flops": 0, "params": 0},
        "Compute-bound ⚡": {"count": 0, "flops": 0, "params": 0},
        "Balanced": {"count": 0, "flops": 0, "params": 0},
    }
    total_flops, total_mem, total_params, total_latency = 0, 0, 0, 0

    # 寫入每一層
    for _, lname, in_shape, out_shape, f, m, r, params, tag, color in layer_stats:
        # === 推論時間計算 ===
        compute_time = f / compute_cap if compute_cap > 0 else 0
        if m <= sram_bytes:
            memory_time = m / (mem_bw * 10)  # 假設 SRAM 比 DRAM 快 10 倍
        else:
            memory_time = m / mem_bw
        latency = max(compute_time, memory_time)

        row = [
            lname, in_shape, out_shape,
            f / 1e6, m / 1024, round(r, 2),
            params / 1e3, tag,
            compute_time * 1e3,  # ms
            memory_time * 1e3,
            latency * 1e3
        ]
        ws.append(row)

        # 整列填色
        if "Memory-bound" in tag:
            fill = fill_mem
        elif "Compute-bound" in tag:
            fill = fill_cmp
        else:
            fill = fill_bal
        for cell in ws[ws.max_row]:
            cell.fill = fill

        # 更新統計
        if tag in summary:
            summary[tag]["count"] += 1
            summary[tag]["flops"] += f
            summary[tag]["params"] += params
        total_flops += f
        total_mem += m
        total_params += params
        total_latency += latency

    # 自動調整欄寬
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # 新增統計頁籤
    ws2 = wb.create_sheet("Statistics")
    ws2.append(["Tag", "Count", "FLOPs (M)", "Params (K)"])
    for tag, vals in summary.items():
        ws2.append([tag, vals["count"], vals["flops"] / 1e6, vals["params"] / 1e3])
    ws2.append([])
    ws2.append(["Total FLOPs (G)", total_flops / 1e9])
    ws2.append(["Total Memory (MB)", total_mem / 1e6])
    ws2.append(["Total Params (M)", total_params / 1e6])
    ws2.append(["Estimated Latency (ms)", total_latency * 1e3])

    wb.save(filename)
    print(f"✅ 已輸出 Excel 檔案: {filename}")
