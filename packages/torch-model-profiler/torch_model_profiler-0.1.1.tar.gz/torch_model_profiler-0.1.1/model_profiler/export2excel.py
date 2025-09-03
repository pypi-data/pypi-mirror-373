import openpyxl
from openpyxl.styles import PatternFill


def export_profile_to_excel(layer_stats, filename="profile_report.xlsx"):
    """
    將 profile_flops_and_memory_layername 的結果輸出成 Excel 檔案。
    - Memory-bound: 整列紅色背景
    - Compute-bound: 整列綠色背景
    - Balanced: 整列灰色背景
    - 額外新增 Statistics 分頁
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profile Report"

    # 標題列
    headers = ["Layer(Name)", "Input Shape", "Output Shape",
               "FLOPs (M)", "Memory (KB)", "FLOP/Byte", "Params (K)", "Tag"]
    ws.append(headers)

    # 填色定義
    fill_mem = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_cmp = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_bal = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    # 統計累加器
    summary = {
        "Memory-bound ❗": {"count": 0, "flops": 0, "params": 0},
        "Compute-bound ⚡": {"count": 0, "flops": 0, "params": 0},
        "Balanced": {"count": 0, "flops": 0, "params": 0},
    }
    total_flops, total_mem, total_params = 0, 0, 0

    # 寫入每一層
    for _, lname, in_shape, out_shape, f, m, r, params, tag, color in layer_stats:
        row = [
            lname, in_shape, out_shape,
            f / 1e6, m / 1024, round(r, 2),
            params / 1e3, tag
        ]
        ws.append(row)

        # 整列填色
        if "Memory-bound" in tag:
            fill = fill_mem
        elif "Compute-bound" in tag:
            fill = fill_cmp
        else:
            fill = fill_bal

        for cell in ws[ws.max_row]:
            cell.fill = fill

        # 更新統計
        if tag in summary:
            summary[tag]["count"] += 1
            summary[tag]["flops"] += f
            summary[tag]["params"] += params
        total_flops += f
        total_mem += m
        total_params += params

    # 自動調整欄寬
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # 新增統計頁籤
    ws2 = wb.create_sheet("Statistics")
    ws2.append(["Tag", "Count", "FLOPs (M)", "Params (K)"])
    for tag, vals in summary.items():
        ws2.append([
            tag, vals["count"],
            vals["flops"] / 1e6,
            vals["params"] / 1e3
        ])
    ws2.append([])
    ws2.append(["Total", "", total_flops / 1e9, total_params / 1e6])
    ws2.append(["Total Memory (MB)", total_mem / 1e6])

    wb.save(filename)
    print(f"✅ export Excel file done: {filename}")

