import ssl
import os
import re
import certifi
from pathlib import Path
import json as Json
import requests
from requests.exceptions import RequestException
#pip install lxml
from lxml import etree
#pip install proxyscrape
import proxyscrape
#pip install openpyxl
import openpyxl
#pip install pymysql
import pymysql

# 全局变量和函数
UserAgentMac="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.6 Safari/605.1.15"
UserAgentIPhone="Mozilla/5.0 (iPhone; CPU iPhone OS 15_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.6 Mobile/15E148 Safari/604.1"
UserAgent=UserAgentMac
Http=requests.Session()
IsProxy=False
IsProxyLocal=False #本地抓包调试

# 获取代理
def GetProxy():
    # 无代理
    if(not IsProxy):return None
    # 本地代理
    if(IsProxyLocal):
        Http.verify = scriptDir+ 'ssl.pem'
        proxy={'http': 'http://127.0.0.1:8888', 'https': 'http://127.0.0.1:8888'}
        return proxy
    # 在线代理
    collector = proxyscrape.create_collector('default', "http")
    collector.get_proxies()
    proxyArr = proxy = collector.get_proxy({"country":"cn"})
    proxy={'http': 'http://'+proxyArr[0]+':'+str(proxyArr[1]), 'https': 'http://'+proxyArr[0]+':'+str(proxyArr[1])}
    return proxy

# get请求
def Get(url,params=None,headers=None,encoding="utf-8"):
    try:
        if(headers is None): headers={"User-Agent":UserAgent}
        res = Http.get(url,headers=headers, params=params, proxies=GetProxy())
    except RequestException as e:
        print("###请求网址："+url+":::")
        print(e)
        return None
    res.encoding=encoding
    return res.text

# 获取json字段，返回字段名和类型
# countPre预留字节数
def GetJsonFields(jsonObj,encoding="utf-8",countPre=10):
    fields={}
    if jsonObj is None: return fields # None则返回空
    if isinstance(jsonObj,list):
        if len(jsonObj)==0: return fields # 空列表则返回空
        else:isList=True
    else: isList=False
    if isList: obj= jsonObj[0] # 读取第一个元素
    for key in obj:
        field={"type":"varchar","size":countPre}
        fields[key]=field
        if isList:
            for it in jsonObj:
                text=it[key]
                if(text is None):continue
                countSrc= field["size"]
                count=len(text.encode(encoding))+countPre # 中文3字节，英文1字节，预留字节
                field["size"]=count if countSrc==None else (count if count>countSrc else countSrc)
        else:
            text=str(jsonObj[key])
            count=len(text.encode(encoding))+countPre # 中文3字节，英文1字节，预留字节
            field["size"]=count
    return fields

# 获取mysql表字段，返回字段名和类型
# countPre预留字节数
def GetMysqlFields(dbconfigOrConnect,tableName):
    funName="getMysqlFields: "
    fields={}
    if tableName is None: 
        print(funName+"表名不能为空")
        return fields
    conn=GetMysqlConnect(dbconfigOrConnect)
    if conn is None: 
        print(funName+"数据库连接失败")
        return fields
    cursor = conn.cursor()
    try:
        sql = f"SHOW COLUMNS FROM {tableName}"
        cursor.execute(sql)
        results = cursor.fetchall()
        for row in results:
            field={"type":row[1],"size":None}
            typeMatch=re.match(r'(\w+)(\((\d+)\))?',row[1])
            if typeMatch:
                field["type"]=typeMatch.group(1)
                if typeMatch.group(3):
                    field["size"]=int(typeMatch.group(3))
            fields[row[0]]=field
    except Exception as e:
        print(e)
    finally:
        cursor.close()
        CloseMysqlConnect(conn,dbconfigOrConnect)
        return fields

# 更新mysql表字段长度
def UpdateMysqlFieldsSize(dbconfigOrConnect,tableName,fields):
    funName="UpdateMysqlFieldsSize: "
    fieldsMysql=GetMysqlFields(dbconfigOrConnect,tableName)
    if(fieldsMysql=={}):
        # 创建表
        CreateMysqlTable(dbconfigOrConnect,tableName,fields)
        return fields
    for key in fields:
        if key not in fieldsMysql: continue
        typeMysql=fieldsMysql[key]["type"].lower()
        sizeMysql=fieldsMysql[key]["size"]
        typeJson=fields[key]["type"].lower()
        sizeJson=fields[key]["size"]
        if typeMysql in ["varchar","char","text","tinytext","mediumtext","longtext"]:
            if sizeJson is not None and (sizeMysql is None or sizeJson>sizeMysql):
                # 扩展字段长度
                sql = f"ALTER TABLE {tableName} MODIFY COLUMN {key} {typeMysql}({sizeJson})"
                print(sql)
                conn=GetMysqlConnect(dbconfigOrConnect)
                if conn is None: 
                    print(funName+"数据库连接失败")
                    return
                cursor = conn.cursor()
                try:
                    cursor.execute(sql)
                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    print(funName+str(e))
                finally:
                    cursor.close()
                    CloseMysqlConnect(conn,dbconfigOrConnect)

# 创建mysql表
def CreateMysqlTable(dbconfigOrConnect,tableName,fields):
    funName="CreateMysqlTable: "
    if tableName is None: 
        print(funName+"表名不能为空")
        return
    if fields is None or len(fields)==0:
        print(funName+"字段不能为空")
        return
    conn=GetMysqlConnect(dbconfigOrConnect)
    if conn is None: 
        print(funName+"数据库连接失败")
        return
    cursor = conn.cursor()
    try:
        fieldStrs=["id int NOT NULL AUTO_INCREMENT PRIMARY KEY"]
        for key in fields:
            if(key.lower()=="id"):continue
            type=fields[key]["type"].lower()
            size=fields[key]["size"]
            if type in ["varchar","char"]:
                size=size if size is not None and size>0 else 255
                fieldStrs.append(f"{key} {type}({size})")
            elif type in ["int","bigint","smallint","tinyint","mediumint"]:
                fieldStrs.append(f"{key} {type}")
            elif type in ["float","double","decimal"]:
                fieldStrs.append(f"{key} {type}(10,2)")
            elif type in ["date","datetime","timestamp"]:
                fieldStrs.append(f"{key} {type}")
            elif type in ["text","tinytext","mediumtext","longtext"]:
                fieldStrs.append(f"{key} {type}")
            else:
                fieldStrs.append(f"{key} varchar(255)")
        fieldStr=",".join(fieldStrs)
        sql = f"CREATE TABLE IF NOT EXISTS {tableName} ({fieldStr}) ENGINE=InnoDB DEFAULT CHARSET=utf8"
        print(sql)
        cursor.execute(sql)
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(e)
    finally:
        cursor.close()
        CloseMysqlConnect(conn,dbconfigOrConnect)

# 获取mysql连接
def GetMysqlConnect(dbconfigOrConnect):
    if dbconfigOrConnect is None : return None
    if isinstance(dbconfigOrConnect,dict):
        dbconfigOrConnect["charset"]=dbconfigOrConnect.get("charset", 'utf8')
        conn = pymysql.connect(**dbconfigOrConnect)
    else:
        conn=dbconfigOrConnect
    return conn
# 关闭mysql连接:只有dbconfigOrConnect为dict时才关闭连接，否则跳过
def CloseMysqlConnect(conn,dbconfigOrConnect=None):
    if conn is None: return
    if isinstance(dbconfigOrConnect,dict):
        conn.close()

# 保存json到mysql数据库对应表中
def JsonToMysql(jsonObj, dbconfigOrConnect,tableName, isAppend=False):
    funName="JsonToMysql: "
    fieldsJson=GetJsonFields(jsonObj)
    conn = GetMysqlConnect(dbconfigOrConnect)
    fieldsMysql=UpdateMysqlFieldsSize(conn,tableName,fieldsJson)
    conn.commit()
    # 更新数据
    cursor = conn.cursor()
    data=[jsonObj] if isinstance(jsonObj,dict) else jsonObj
    try:
        if not isAppend:
            cursor.execute(f"DELETE FROM {tableName} where 1=1") # 清空数据
        i=0
        for item in data:
            keys = item.keys()
            values = tuple(item.values())
            sql = f"INSERT INTO {tableName} ({','.join(keys)}) VALUES {values}"
            cursor.execute(sql)  # 参数化查询更安全
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(funName+str(e))
    finally:
        cursor.close()
        CloseMysqlConnect(conn,dbconfigOrConnect)

# 从mysql表中分页读取记录并返回json对象
def MysqlToJson(dbconfigOrConnect,tableName,where=None,page=1,pageSize=100):
    funName="MysqlToJson: "
    if tableName is None: 
        print(funName+"表名不能为空")
        return []
    conn=GetMysqlConnect(dbconfigOrConnect)
    if conn is None: 
        print(funName+"数据库连接失败")
        return []
    cursor = conn.cursor(pymysql.cursors.DictCursor)
    try:
        offset=(page-1)*pageSize
        sql = f"SELECT * FROM {tableName} " + (f" WHERE {where} " if where is not None else "") + f" LIMIT {offset},{pageSize}"
        cursor.execute(sql)
        results = cursor.fetchall()
        return results
    except Exception as e:
        print(funName+str(e))
        return []
    finally:
        cursor.close()
        CloseMysqlConnect(conn,dbconfigOrConnect)

# 保存json到Excel文件中
def JsonToExcel(jsonObj,dirPath,name,isAppend=False):
    funName="jsonToExcel: "
    if not os.path.exists(dirPath):
        print(funName+"目录不存在")
        return
    # 创建Excel工作簿和工作表
    excelFile=dirPath+name+".xlsx"
    if isAppend and os.path.exists(excelFile):
        wb=openpyxl.load_workbook(excelFile)
        ws = wb.active
        startRow=ws.max_row
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        startRow=2
        ws.title = name
    # 写入数据
    if isinstance(jsonObj,dict):
        col=1
        for key in jsonObj:
            ws.cell(row=1,column=col,value=key)
            ws.cell(row=startRow,column=col,value=jsonObj[key])
            col+=1
    else:
        if(len(jsonObj)==0):
            print(funName+"没有数据")
            return
        col=1
        for key in jsonObj[0]:
            ws.cell(row=1,column=col,value=key)
            col+=1
        row=startRow
        for it in jsonObj:
            col=1
            for key in it:
                ws.cell(row=row,column=col,value=it[key])
                col+=1
            row+=1
    # 保存Excel文件
    wb.save(excelFile)
    print(funName+"保存完毕,"+excelFile)

# 保存json对文件
# page>=1表示分页，会创建name文件夹;0表示不分页,不会创建name文件夹,会替换同名文件;
def JsonToFile(jsonObj,dirPath,name,page=0,encoding="utf-8"):
    # 格式化json字符串
    rst=Json.dumps(jsonObj,indent=2,ensure_ascii=False)
    # 创建目录
    dir=dirPath if page==0 else dirPath+name+"/"
    path = Path(dir).absolute()
    path.mkdir(parents=True, exist_ok=True)
    # 写文件
    filePath=dir+str(page)+".json"
    open(filePath,"w",encoding=encoding).write(rst)
# 判断json文件是否存在
# page从1开始,0表示不分页
def ExistJsonFile(dirPath,name,page=0):
    dir=dirPath if page==0 else dirPath+name+"/"
    filePath=dir+str(page)+".json"
    return os.path.exists(filePath)

# Html保存到文件
# page>=1表示分页，会创建name文件夹;0表示不分页,不会创建name文件夹,会替换同名文件;
def HtmlToFile(data,dirPath,name,page=0,encoding="utf-8"):
    dir=dirPath if page==0 else dirPath+name+"/"
    path = Path(dir).absolute()
    path.mkdir(parents=True, exist_ok=True)
    filePath=dir+str(page)+".html"
    # 写文件
    open(filePath,"w",encoding=encoding).write(data)
# 判断html文件是否存在
# page从1开始,0表示不分页
def ExistHtmlFile(dirPath,name,page=0):
    dir=dirPath if page==0 else dirPath+name+"/"
    filePath=dir+str(page)+".html"
    return os.path.exists(filePath)
# 读取html文件
# page从1开始,0表示不分页
def GetHtmlFile(dirPath,name,page=0,encoding="utf-8"):
    dir=dirPath if page==0 else dirPath+name+"/"
    filePath=dir+str(page)+".html"
    rst=open(filePath,"r",encoding=encoding).read()
    return rst

# 将目录下的json文件合并或分别保存到Excel文件
def JsonDirToExcel(dirPath,name,encoding="utf-8",isMerge=False):
    funName="jsonFileToExcelFile: "
    dir=dirPath+name+"/"
    path = Path(dir).absolute()
    if not os.path.exists(path):
        print(funName+"目录不存在"+dir)
        return
    files=os.listdir(path)
    allData=[]
    for file in files:
        if not file.endswith(".json"):continue
        filePath=dir+file
        jsonStr=open(filePath,"r",encoding=encoding).read()
        jsonObj=Json.loads(jsonStr)
        if isMerge:
            allData+=jsonObj
        else:
            # 分别保存
            JsonToExcel(jsonObj,dir,file.replace(".json",""))
    # 合并保存
    if isMerge:
        if(len(allData)==0):
            print(funName+"没有数据")
            return
        JsonToExcel(allData,dirPath,name)

# 将目录下的json文件合并保存到mysql
def JsonDirToMysql(dirPath, dbconfigOrConnect,name,encoding="utf-8"):
    funName="jsonDirToMysql: "
    dir=dirPath+name+"/"
    path = Path(dir).absolute()
    if not os.path.exists(path):
        print(funName+"目录不存在"+dir)
        return
    files=os.listdir(path)
    allData=[]
    for file in files:
        if not file.endswith(".json"):continue
        filePath=dir+file
        jsonStr=open(filePath,"r",encoding=encoding).read()
        jsonObj=Json.loads(jsonStr)
        allData+=jsonObj
    if(len(allData)==0):
        print(funName+"没有数据")
        return
    JsonToMysql(allData,dbconfigOrConnect,name)

# 从mysql读取数据并保存到json文件
def MysqlToJsonDir(dbconfigOrConnect,dirPath,name,pageSize=100,encoding="utf-8"):
    funName="MysqlToJsonDir: "
    if not os.path.exists(dirPath):
        print(funName+"目录不存在")
        return
    page=1
    isLoop=True
    while isLoop:
        list=MysqlToJson(dbconfigOrConnect,name,None,page,pageSize)
        if(list is None or len(list)==0):
            isLoop=False
            continue
        JsonToFile(list,dirPath,name,page)
        print(funName+"第"+str(page)+"页, 保存完毕")
        page+=1

# 从Excel文件读取数据并保存到json文件
def ExcelToJsonDir(dirPath,name,pageSize=100,encoding="utf-8"):
    funName="ExcelToJsonDir: "
    excelFile=dirPath+name+".xlsx"
    if not os.path.exists(excelFile):
        print(funName+"文件不存在"+excelFile)
        return
    wb=openpyxl.load_workbook(excelFile)
    ws = wb.active
    maxRow=ws.max_row
    if maxRow<2:
        print(funName+"没有数据")
        return
    headers=[]
    for col in range(1,ws.max_column+1):
        headers.append(ws.cell(row=1,column=col).value)
    page=1
    isLoop=True
    while isLoop:
        list=[]
        startRow=(page-1)*pageSize+2
        endRow=startRow+pageSize-1
        if startRow>maxRow:
            isLoop=False
            continue
        if endRow>maxRow:endRow=maxRow
        for row in range(startRow,endRow+1):
            item={}
            for col in range(1,ws.max_column+1):
                key=headers[col-1]
                value=ws.cell(row=row,column=col).value
                item[key]=str(value).strip() if value is not None else ""
            list.append(item)
        JsonToFile(list,dirPath,name,page)
        print(funName+"第"+str(page)+"页, 保存完毕")
        page+=1
