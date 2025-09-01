from deepdiff import DeepDiff
from dateutil.relativedelta import relativedelta
from urllib.parse import urlparse
import datetime
import openpyxl


class DRTFunctions:

    def read_xlsx(self, file_path, sheet_names=None, is_head=True) -> list:
        """
        读取xlsx文焕
        :param file_path:       文件路径
        :param sheet_names:     sheet名称，可以为None, list, tuple类型
        :param is_head:         是否有列名
        :return: 
        """
        myxls = openpyxl.load_workbook(file_path)
        if sheet_names is None:
            sheet_name_list = myxls.sheetnames
        elif isinstance(sheet_names, (list, tuple)):
            for sn in sheet_names:
                if sn not in myxls.sheetnames:
                    raise ValueError('Sheet name "{}" does not exist!'.format(sn))
            else:
                sheet_name_list = sheet_names
        else:
            raise ValueError('Parameter <sheet_name> should be a list, tuple, or None')
        xlsx_content = list()
        for sheet_name_item in sheet_name_list:
            sheet_content = list()
            active_sheet = myxls[sheet_name_item]
            if is_head:
                headers = [cell.value for cell in active_sheet[1] if cell.value]
                for row in active_sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    sheet_content.append(row_dict)
            else:
                for rows in active_sheet.iter_rows(values_only=True):
                    last_not_none_index = None
                    for n, r in enumerate(rows[::-1]):
                        if r is not None:
                            last_not_none_index = n
                            break
                    row_list = last_not_none_index and rows[:-last_not_none_index] or rows
                    sheet_content.append(row_list)
            xlsx_content.append({"sheet_name": sheet_name_item, "sheet_data": sheet_content})
        return xlsx_content

    def datetime_strftime(self, strftime=None, **kwargs):
        '''
        日期时间处理
        支持：years, months, days, weeks, hours, minutes, seconds, microseconds
             year,  month,  day,  week,  hour,  minute,  second,  microsecond
        :param strftime: ("%Y-%m-%d %H:%M:%S", )
        :param kwargs:   (month=1, day=1)<今年1月1日>   (months=-1, days=1)<上一个月往后1天>
        :return: 年月日时分秒
        '''
        new_date = datetime.datetime.now() + relativedelta(**kwargs)
        if strftime:
            new_date = new_date.strftime(strftime)
        return new_date

    def datatime_timestamp(self, **kwargs):
        '''
        时间戳处理
        支持：years, months, days, weeks, hours, minutes, seconds, microseconds
             year,  month,  day,  week,  hour,  minute,  second,  microsecond
        :param kwargs: (month=1, day=1)<今年1月1日>   (months=-1, days=1)<上一个月往后1天>
        :return: 时间戳（精确到s）
        '''
        new_date = datetime.datetime.now() + relativedelta(**kwargs)
        return int(new_date.timestamp())

    # url元组
    def get_url(self, url):
        return urlparse(url)


class DRTDiff:
    def diff(self, data1, data2, **kwargs):
        diff = DeepDiff(data1, data2, **kwargs)
        if diff:
            raise AssertionError(diff.pretty())
