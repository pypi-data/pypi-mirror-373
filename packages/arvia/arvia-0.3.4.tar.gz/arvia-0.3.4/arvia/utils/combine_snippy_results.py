from arvia.utils.aeruginosa_snippy import (
    filter_snippy_result,
    filter_synonymous_variants,
    AERUGINOSA_GENES,
)
from arvia.utils.aeruginosa_polymorphisms import PAERUGINOSA_POLYMORPHISMS
from arvia.utils.aeruginosa_truncations import check_truncations
import pandas as pd
import glob
from pathlib import Path
import numpy as np
import openpyxl


def concat_files(file_list, selected_bcs=[], header="infer"):
    l = []
    for i in file_list:
        folder = Path(i).parent
        bc = Path(folder).name
        if selected_bcs and bc not in selected_bcs:
            continue
        try:
            temp_df = pd.read_csv(i, sep=None, header=header, engine="python")
            temp_df["bc"] = bc
            l.append(temp_df)

        except Exception as e:
            print(e)
            print("Error: File is not tab or comma delimited, could be empty")

    data = pd.concat(l)
    return data


def apply_paeruginosa_polymorphisms(row, d=PAERUGINOSA_POLYMORPHISMS):
    effect = row["EFFECT"].split(" ")
    prot_effect = [i for i in effect if i.startswith("p.")]
    assert len(prot_effect) <= 1
    if prot_effect:
        prot_effect = prot_effect[0].replace("p.", "")
        if d.get(row["LOCUS_TAG"]) and prot_effect in d.get(row["LOCUS_TAG"]):
            row["polymorphic_substitution"] = "possible polymorphism"

    return row


def get_mutation_props(value):
    try:
        mut_allele_count, ref_allele_count = [int(i.split(":")[1]) for i in value.split(" ")]
        mutation_prop = 100 * mut_allele_count / (mut_allele_count + ref_allele_count)
        return mutation_prop, mut_allele_count
    except Exception as e:
        return None, None


def color_cells(value):
    if value == "-":
        return "text-align: left"
    else:
        mutation_prop, mut_allele_count = get_mutation_props(value)
        if mutation_prop:
            color = None
            if mutation_prop >= 75 and mut_allele_count >= 5:
                color = "#069a2e"
            else:
                color = "#ffaf19"
            return f"background-color : {color}; color: black" + "; text-align: left"
        elif "PMF:" in value:
            color = "#ffaf19"
            return f"background-color : {color}; color: black" + "; text-align: left"
        else:
            return "text-align: left"


def color_cells_v2(value):
    background_color = ""
    s = {"*", "fs", "del", "dup", "ins", "?", "trunc", "possible_missing_feature"}
    search = [True for i in s if i in value]
    common = "; vertical-align: middle"
    if len(search) >= 1:
        background_color = "#ffd17a"
        return f"background-color : {background_color}; color: black" + "; text-align: left" + common
    elif value.startswith("NZC"):
        d = {i.split("=")[0]: float(i.split("=")[1][:-1]) for i in value.split(" ")}
        if d["NZC"] < 100:
            background_color = "#ffd17a"
            return f"background-color : {background_color}; color: black" + "; text-align: left" + common
        elif d["NZC"] >= 100 and d["Depth"] < 10:
            background_color = " #ffea69"
            return f"background-color : {background_color}; color: black" + "; text-align: left" + common
    elif value == "No assembly given":
        background_color = " #B8B8B8"
        return f"background-color : {background_color}; color: black" + "; text-align: left" + common
    elif value == "-":
        return f"color: black" + "; text-align: center" + common
    else:
        return "text-align: left" + common


def highlight_header(s):
    return [
        "background-color: #2a6099; color: white; font-weight: bold; text-align: center; vertical-align: middle; border: 1.2px solid black;"
        for v in s
    ]


def prepare_mutations_for_wide_pivot(data: pd.DataFrame):
    if len(data) >= 1:
        # Normalize positions based on strands (orders mutations)
        data["normalized_pos"] = data["POS"] * (data["STRAND"] + "1").astype(int)
        data = data.sort_values(["CHROM", "LOCUS_TAG", "normalized_pos"], ascending=True).reset_index()

        # Create gene id
        data["gene_id"] = data.apply(
            lambda row: f"{row['LOCUS_TAG']}__{row['GENE']}"
            if row["GENE"]
            else f"{row['LOCUS_TAG']}__{row['PRODUCT']}",
            axis=1,
        )

        # Extract mutation parts (nucl, prot...)
        def check_mut_format(row):
            # Example of row["mut"] = [('missense_variant', '1472C>T', 'Ala491Val')]
            if row["mut"]:
                # Check if the supposed protein mutation is just a number
                # This happens in splice region variants like: 
                # -> splice_region_variant&stop_retained_variant c.2420_*5delGACAGGTinsAACGGGG p.808
                #                                                                                 |> this part
                # If it is a number then return the full effect, else keep as is
                try:
                    _ = int(row["mut"][0][2])
                    return ("", "", row["EFFECT"])
                except Exception as e:                
                    return row["mut"][0]
            else:
                # If pattern did not match then row["mut"] will be an empty list
                # Then we return the full effect
                return ("", "", row["EFFECT"])

        data["mut"] = data["EFFECT"].str.findall("(.*) c\.(.*) p\.(.*)")
        data["mut"] = data.apply(check_mut_format, axis=1)
        data[["mutation_type", "mutation_nucl", "mutation_prot"]] = pd.DataFrame(
            data["mut"].to_list(), columns=["mutation_type", "mutation_nucl", "mutation_prot"]
        )

        # Check mutation heterogeneity and depth
        data["mut_qc"] = (
            data["EVIDENCE"]
            .apply(get_mutation_props)
            .apply(
                lambda x: f"(Fails QC: {round(x[0],2)}%, {x[1]}x)"
                if x[0] is not None and (x[0] < 75 or x[1] < 5)
                else ""
            )
        )
        data["mutation_prot"] = (data["mutation_prot"] + " " + data["mut_qc"]).str.strip()

        # Add polymorphisms info
        if "polymorphic_substitution" in data.columns:
            data["mutation_prot"] = data.apply(
                lambda row: f"{row['mutation_prot']} (POLY)"
                if row["polymorphic_substitution"] is not np.nan and row["polymorphic_substitution"] != "-"
                else row["mutation_prot"],
                axis=1,
            )
    else:  # no mutations
        data[
            [
                "normalized_pos",
                "gene_id",
                "mut",
                "mutation_type",
                "mutation_nucl",
                "mutation_prot",
                "mut_qc",
                "mutation_prot",
            ]
        ] = ""

    return data


def get_default_snippy_combination(
    input_files: list, output_file: str = None, selected_bcs: list = [], paeruginosa: bool = False
):

    # -- Concat files --
    df = concat_files(input_files, selected_bcs=selected_bcs)
    if selected_bcs:
        bcs = selected_bcs
        assert len([i for i in selected_bcs if i not in list(df["bc"].unique())]) == 0, "Lacking one input"
    else:
        bcs = list(df["bc"].unique())

    # -- Filter synonymous --
    df = filter_synonymous_variants(df)

    # -- Run paeruginosa filters --
    if paeruginosa:
        df = filter_snippy_result(df).apply(
            apply_paeruginosa_polymorphisms, axis=1
        )  # ATTENTION ONLY FOR PSEUDOMONAS!!!

    # -- Format and copy for later --
    df = df.fillna("-")
    df_copy = df.copy().reset_index()

    # -- Pivot wide --
    index_cols = ["CHROM", "LOCUS_TAG", "POS", "GENE", "PRODUCT", "EFFECT"]
    if "polymorphic_substitution" in df.columns:
        index_cols += ["polymorphic_substitution"]
    pivoted_df = df.pivot(
        index=index_cols,
        columns="bc",
        values="EVIDENCE",
    )

    # -- Paint --
    bcs_with_no_muts = [i for i in bcs if i not in pivoted_df.columns]
    for i in bcs_with_no_muts:
        pivoted_df[i] = "-"

    columns_to_paint = bcs
    pivoted_df = pivoted_df.fillna("-")
    pivoted_df = pivoted_df.style.applymap(color_cells, subset=pd.IndexSlice[:, columns_to_paint])
    pivoted_df = pivoted_df.apply_index(highlight_header, axis="columns", level=[0])
    pivoted_df = pivoted_df.apply_index(lambda s: [
            "text-align: left; vertical-align: middle"
            for v in s
        ], axis="index")

    # Writer object
    if output_file:
        sheet_name = "Snippy module comparison"

        # ---- Save a xlsx ----
        writer = pd.ExcelWriter(output_file, engine="xlsxwriter")

        # Convert the styled dataframe to an XlsxWriter Excel object in specific sheet
        pivoted_df.to_excel(writer, sheet_name=sheet_name)

        # Select sheet and apply formatting
        workbook = writer.book
        sheet = writer.sheets[sheet_name]

        first_sample_col = 7 if "polymorphic_substitution" in df.columns else 6

        # sheet.autofit()  # autofit row widths
        # sheet.set_row(1, 45)  # height of row
        # sheet.set_row(2, 2, [], {"hidden": True})  # row is hidden
        sheet.set_column(1, 1, 10)  # width of column 1
        sheet.set_column(4, 5, 30)  # width of columns 4-5
        # sheet.set_column(6, 6, 23)  # width of column 6
        sheet.set_column(first_sample_col, first_sample_col + len(pivoted_df.columns), 20)  # width of sample columns
        sheet.freeze_panes(1, first_sample_col)  # freeze first row and first 7 column

        # Save
        workbook.close()

        # ---- Load again and paint remaining things (i dont know how paint multiindex header in pandas) ----
        wb = openpyxl.load_workbook(output_file)
        sheet = wb[sheet_name]

        for cell in [f"{i}1" for i in list("ABCDEFG")]:
            sheet[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical="center")
            sheet[cell].font = openpyxl.styles.Font(name='Calibri', bold=True, color="FFFFFF")
            sheet[cell].fill = openpyxl.styles.PatternFill("solid", fgColor="2a6099")

        # Save again
        wb.save(output_file)

    return df_copy, bcs


def paeruginosa_combine_all_mutational_res(
    default_df: pd.DataFrame,
    oprd_fs: list,
    oprd_refs_fs: list,
    gene_coverage_fs: list,
    truncation_fs: list,
    bcs: list,
    output_file: str,
    bcs_without_assembly: list = [],
    filter_poly: bool = False,
):
    if not oprd_fs:
        raise Exception("Expected oprD VC with closest ref. files")

    temp = prepare_mutations_for_wide_pivot(default_df.copy())

    if filter_poly:
        temp = temp[temp["polymorphic_substitution"] != "possible polymorphism"]

    # ---- Prepare oprd closest reference muts ----
    oprd_df = concat_files(oprd_fs)
    oprd_df = filter_synonymous_variants(oprd_df)
    oprd_df = prepare_mutations_for_wide_pivot(oprd_df)
    oprd_df["CHROM"] = oprd_df["CHROM"].str.replace("Pseudomonas_aeruginosa_", "")
    oprd_df = oprd_df[["bc", "CHROM", "gene_id", "mutation_prot"]]

    # -- Pivot all muts --
    temp["gene_id"] = temp["gene_id"] + "__(Snippy)"
    temp_pivot = (
        temp.drop_duplicates()
        .groupby(["bc", "gene_id"], sort=False, dropna=False)["mutation_prot"]
        .apply(lambda x: "; ".join(list(x)))
        .reset_index(name="muts")
        .pivot(index="bc", columns=["gene_id"], values="muts")
    )
    bcs_with_no_muts = [i for i in bcs if i not in temp_pivot.reset_index()["bc"].to_list()]
    assert len(bcs_with_no_muts) == 0, f"Unexpected samples without mutations: {bcs_with_no_muts}"

    # -- Pivot closest oprD muts --
    temp_oprd = (
        oprd_df.drop_duplicates()
        .groupby(["bc", "gene_id"], sort=False, dropna=False)["mutation_prot"]
        .apply(lambda x: "; ".join(list(x)))
        .reset_index(name="muts")[
            [
                "bc",
                "muts",
            ]
        ]
    )
    # add reference ids on the side because sometimes no mutations are detected in oprd_fs and ref is lost
    oprd_refs = concat_files(oprd_refs_fs, header=None).reset_index(drop=True)
    oprd_refs.columns = ["id", "ref", "CHROM", "identity", "coverage", "bc"]
    temp_oprd = (
        pd.merge(oprd_refs[["bc", "CHROM"]], temp_oprd, on="bc", how="left")
        .rename(
            columns={
                "CHROM": "PA0958-alt__oprD__closest reference used",
                "muts": "PA0958-alt__oprD__closest reference",
            }
        )
        .fillna("-")
    )

    # -- Truncations based on assembly --
    truncations_df = concat_files(truncation_fs)
    truncations_df["locus_gene"] = (
        truncations_df["locus_tag"].astype(str)
        + "__"
        + truncations_df["gene"].astype(str)
        + "__"
        + "(Assembly BLAST)"
    )
    temp_truncation = truncations_df.pivot(index="bc", columns="locus_gene", values="comment")

    # ---- Gene coverage ----
    gene_coverage = concat_files(gene_coverage_fs)
    gene_coverage = gene_coverage[["bc", "locus_tag", "non_zero_coverage", "mean_depth"]]
    gene_coverage["comment"] = gene_coverage.apply(
        lambda row: f"NZC={round(100*row['non_zero_coverage'],2)}% Depth={round(row['mean_depth'],2)}x", axis=1
    )
    # gene_coverage["locus_tag"] = gene_coverage["locus_tag"].str.split(".", expand=True)[0]
    gene_coverage = pd.merge(
        pd.DataFrame(AERUGINOSA_GENES.items(), columns=["locus_tag", "gene"]), gene_coverage, on="locus_tag"
    )
    gene_coverage["locus_gene"] = (
        gene_coverage["locus_tag"].astype(str) + "__" + gene_coverage["gene"].astype(str) + "__" + "(Gene coverage)"
    )
    temp_gene_cov = gene_coverage.pivot(index="bc", columns="locus_gene", values="comment")

    # -- Merge --
    xxx = pd.merge(temp_pivot, temp_oprd, on="bc", how="left")
    xxx = pd.merge(xxx, temp_truncation, on="bc", how="left")
    xxx = pd.merge(xxx, temp_gene_cov, on="bc", how="left")
    xxx = xxx[["bc"] + [i for i in sorted(xxx.columns) if i != "bc"]]
    xxx[xxx.isna()] = "-"

    # md = pd.read_csv(
    #     "/media/usuario/15d7f455-39ea-40e7-add2-de57c58767eb/analyses/IMIREL/revision_paper_anacandela_imirel/correlation.tsv",
    #     sep="\t",
    # )
    # md = md.rename(columns={"Isolate ID in Database": "bc"})[["CODE", "bc"]]
    # md["bc"] = md["bc"].str.strip()
    # xxx = pd.merge(xxx, md, on="bc", how="left")

    # md = pd.read_excel("/home/usuario/Descargas/ARGA.xlsx")
    # md = md.rename(
    #     columns={
    #         "ID fastq": "bc",
    #         "Profundidad (mediana)": "depth",
    #         "Integridad (%)": "completeness",
    #         "Contaminación (%)": "contamination",
    #     }
    # )[["bc", "ST", "depth", "completeness", "contamination"]]
    # md["bc"] = md["bc"].astype(str)
    # xxx["bc"] = xxx["bc"].astype(str)
    # xxx = pd.merge(xxx, md, on="bc", how="left")

    index_cols = ["bc"]  # , "CODE", "ST", "depth", "completeness", "contamination"
    xxx = xxx[index_cols + [i for i in xxx.columns if i not in index_cols]]

    # xxx.style.applymap(
    #     color_cells_v2,
    #     subset=pd.IndexSlice[
    #         :,
    #         [c for c in xxx.columns if c not in index_cols],
    #     ],
    # ).to_excel("/home/usuario/Proyectos/Results/test7.xlsx")

    yyy = pd.melt(xxx, id_vars=index_cols, var_name="section")
    yyy["locus_tag"] = yyy["section"].str.split("__", expand=True)[0]

    # Check if all genes appear in all sections of the table
    # (For example, if snippy didnt find a mutation its column will not appear)
    for ref_locus, ref_gene_name in AERUGINOSA_GENES.items():
        for section in ["Snippy"]: # "Assembly BLAST", "Gene coverage", 
            for bc in set(yyy["bc"].to_list()):
                long_section = f"{ref_locus}__{ref_gene_name}__({section})"
                conditions = (
                    (yyy["bc"]==bc)
                    & (yyy["section"]==long_section)
                    & (yyy["locus_tag"]==ref_locus)
                )
                if len(yyy[conditions])==0:
                    yyy = pd.concat(
                        [
                            yyy,
                            pd.DataFrame({"bc": [bc], "section": [long_section], "value": ["-"], "locus_tag": [ref_locus]})
                        ]
                    )
    yyy = yyy.reset_index(drop=True).sort_values(["bc", "locus_tag", "section"], ascending=True)

    # Fill values of samples without blast result with something that indicates the sample did not have an assembly
    if bcs_without_assembly:
        yyy.loc[
            (yyy["section"].str.contains("Assembly BLAST")) 
            & (yyy["bc"].isin(bcs_without_assembly)) 
            & (yyy["value"]=="-"), 
            "value"
        ] = "No assembly given"
    
    # Save this table, which is the long version of the final table and contains everthing
    zzz = yyy.copy()
    zzz["section"] = zzz["section"].str.replace("(","").str.replace(")","")
    zzz[["_","gene","section"]] = zzz["section"].str.split("__", expand=True) # the string in this field at this point is like "PA0424__mexR__Gene coverage"
    zzz = zzz.drop(columns = ["_"])
    zzz = zzz[["bc","locus_tag", "gene","section","value"]]

    # Sort with specific order
    expected_categories = ['Snippy', 'Assembly BLAST', 'Gene coverage', 'closest reference', 'closest reference used']
    temp = [i for i in zzz.section.unique() if i not in expected_categories]
    assert len(temp)==0, f"At least a value from was not in expected_categories: {temp}"
    zzz["section"] = pd.Categorical(zzz["section"], categories=expected_categories)
    zzz = zzz.sort_values(["bc", "locus_tag", "section"], ascending=True)

    # Save
    zzz.to_csv(output_file, sep="\t", index=None)

    return zzz


def create_merged_xlsx_result(
    combined_long_f: Path,
    output_file: Path,
    input_files_dict: dict,
    amrfinderplus_fs: list = [],
    mlst_fs: list = [],
    ):

    # output_file = "/home/usuario/Proyectos/Results/tests/arvia/ARVIA.xlsx"
    # combined_long_f = "/home/usuario/Proyectos/Results/tests/arvia/arvia/temp/results_merged/full_long.tsv"
    # amrfinderplus_fs = glob.glob("/home/usuario/Proyectos/Results/tests/arvia/arvia/results_per_sample/*/*_amrfinderplus.tsv")
    # mlst_fs = glob.glob("/home/usuario/Proyectos/Results/tests/arvia/arvia/results_per_sample/*/*_mlst.tsv")

    # # input_files_dict = {
    # #     "ARGA000190": {
    # #         "pipeline": "x4",
    # #     },
    # #     "ARGA00024": {
    # #         "pipeline": "x3",
    # #     },
    # #     "ARGA00025": {
    # #         "pipeline": "x2",
    # #     },
    # #     "ARGA00461": {
    # #         "pipeline": "x1",
    # #     },
    # # }
    # input_files_dict = {'ARGA00080': {'reads': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00080/ARGA00080_R1.fastq.gz', '/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00080/ARGA00080_R2.fastq.gz'], 'assembly': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/assembly/05_final_assembly/ARGA00080/ARGA00080_assembly.fasta'], 'reads_type': 'paired_end', 'pipeline': 'full_pipeline'}, 'ARGA00043': {'reads': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00043/ARGA00043_R1.fastq.gz', '/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00043/ARGA00043_R2.fastq.gz'], 'assembly': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/assembly/05_final_assembly/ARGA00043/ARGA00043_assembly.fasta'], 'reads_type': 'paired_end', 'pipeline': 'full_pipeline'}, 'ARGA00043_sr': {'reads': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00043/ARGA00043_R1.fastq.gz', '/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00043/ARGA00043_R2.fastq.gz'], 'assembly': [], 'reads_type': 'paired_end', 'pipeline': 'only_reads'}, 'ARGA00086': {'reads': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00086/ARGA00086_R1.fastq.gz', '/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00086/ARGA00086_R2.fastq.gz'], 'assembly': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/assembly/05_final_assembly/ARGA00086/ARGA00086_assembly.fasta'], 'reads_type': 'paired_end', 'pipeline': 'full_pipeline'}, 'ARGA00050': {'reads': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00050/ARGA00050_R1.fastq.gz', '/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00050/ARGA00050_R2.fastq.gz'], 'assembly': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/assembly/05_final_assembly/ARGA00050/ARGA00050_assembly.fasta'], 'reads_type': 'paired_end', 'pipeline': 'full_pipeline'}, 'ARGA00034': {'reads': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00034/ARGA00034_R1.fastq.gz', '/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/short_reads_qc/06_clean_reads/ARGA00034/ARGA00034_R2.fastq.gz'], 'assembly': ['/home/usuario/Proyectos/Results/ARGA/ARGA_ALL_v2/assembly/05_final_assembly/ARGA00034/ARGA00034_assembly.fasta'], 'reads_type': 'paired_end', 'pipeline': 'full_pipeline'}}

    assembly_analysis_section_name = "General assembly analysis"

    def generate_expected_empty_dataframe(bcs, s1: str, s2: str, value: str = np.nan):
        return pd.DataFrame(
            {
                "bc": bcs,
                "S1": [s1 for i in range(0, len(bcs))],
                "S2": [s2 for i in range(0, len(bcs))],
                "value": [value for i in range(0, len(bcs))]
            }
        )

    # ---- Pipelines ----
    pipelines_df = pd.DataFrame(
        [
            {"bc": k, "pipeline": v["pipeline"]} for k,v in input_files_dict.items()
        ]
    )
    pipelines_df.columns = ["bc","value"]
    pipelines_df["S1"] = ""
    pipelines_df["S2"] = "Pipeline"

    # ---- Pipelines ----
    reads_type_df = pd.DataFrame(
        [
            {"bc": k, "reads_type": v["reads_type"]} for k,v in input_files_dict.items()
        ]
    )
    reads_type_df.columns = ["bc","value"]
    reads_type_df["S1"] = ""
    reads_type_df["S2"] = "Reads type"
    reads_type_df["value"] = reads_type_df["value"].fillna("-")


    # ---- Mutations ----
    combined_long_df = pd.read_csv(combined_long_f, sep="\t")    
    combined_long_df["S1"] = combined_long_df["locus_tag"]
    combined_long_df["S2"] = combined_long_df.apply(lambda r: f"{r['locus_tag']} \n{r['gene']} \n({r['section']})", axis=1)

    # ---- Amrfinder ----
    # Read and process
    if amrfinderplus_fs:
        amrfinder_df = concat_files(amrfinderplus_fs)

        # - Separate pdc from the rest of genes -
        # type: AMR, STRESS, or VIRULENCE.
        # subtype: ANTIGEN, BIOCIDE, HEAT, METAL, PORIN, STX_TYPE
        # class: antibiotic class
        element_is_pdc = amrfinder_df["mod_element_symbol"].str.contains("blaPDC")
        element_is_amr = amrfinder_df["type"]=="AMR"
        element_is_stress = amrfinder_df["type"]=="STRESS"
        element_is_vir = amrfinder_df["type"]=="VIRULENCE"
        amr_genes = amrfinder_df[~element_is_pdc & (element_is_amr | element_is_stress)].groupby("bc", group_keys=True)["mod_element_symbol"].apply(lambda x: "; ".join(sorted(x))).reset_index(name="value")
        amr_pdc = amrfinder_df[element_is_pdc & element_is_amr].groupby("bc", group_keys=True)["mod_element_symbol"].apply(lambda x: "; ".join(sorted(x))).reset_index(name="value")
        # vir_genes = amrfinder_df[element_is_vir].groupby("bc", group_keys=True)["mod_element_symbol"].apply(lambda x: "; ".join(sorted(x))).reset_index(name="value")
        # stress_genes = amrfinder_df[element_is_stress].groupby("bc", group_keys=True)["mod_element_symbol"].apply(lambda x: "; ".join(sorted(x))).reset_index(name="value")

        # Prepare for merge
        amr_genes["S1"] = assembly_analysis_section_name
        amr_genes["S2"] = "Other AMR Genes"
        amr_pdc["S1"] = assembly_analysis_section_name
        amr_pdc["S2"] = "PDC"
        # vir_genes["S1"] = assembly_analysis_section_name
        # vir_genes["S2"] = "Virulence genes"
        # stress_genes["S1"] = assembly_analysis_section_name
        # stress_genes["S2"] = "Stress genes"

    else:
        amr_genes = generate_expected_empty_dataframe(
            bcs=list(combined_long_df["bc"].unique()), 
            s1=assembly_analysis_section_name,
            s2="Other AMR Genes",
        )
        amr_pdc = generate_expected_empty_dataframe(
            bcs=list(combined_long_df["bc"].unique()), 
            s1=assembly_analysis_section_name,
            s2="PDC",
        )

    # ---- MLST ----
    if mlst_fs:
        mlst_df = concat_files(mlst_fs)
        mlst_df = mlst_df.rename(columns={"final_mlst": "value"})
        mlst_df = mlst_df[["bc", "mlst_model", "value"]]

        # Prepare for merge
        mlst_model_df = mlst_df[["bc","mlst_model"]].rename(columns={"mlst_model":"value"})
        mlst_df = mlst_df[["bc","value"]]
        mlst_model_df["S1"] = assembly_analysis_section_name
        mlst_model_df["S2"] = "MLST Model"
        mlst_df["S1"] = assembly_analysis_section_name
        mlst_df["S2"] = "MLST"
    else:
        mlst_df = generate_expected_empty_dataframe(
            bcs=list(combined_long_df["bc"].unique()), 
            s1=assembly_analysis_section_name,
            s2="MLST",
        )

    # ---- Merge ----
    # This df has to habe columns
    # S1: main section
    # S2: subsection
    # bc: sample id
    # value: value for analysis in subsection
    df = pd.concat(
        [
            pipelines_df,
            reads_type_df,
            mlst_model_df,
            mlst_df,
            amr_pdc,
            amr_genes,
            # vir_genes,
            # stress_genes,
            combined_long_df,
        ]
    )

    # ---- Final XLSX ----
    # Pivot
    index_cols = ["bc"]
    idx = pd.IndexSlice
    no_assembly_given_cols = ["MLST","MLST Model","PDC","Other AMR Genes"]
    temp_pivot = df.pivot(index=index_cols, columns=["S1", "S2"], values="value")

    # On those bcs that did not have an assembly, they will have np.nan in mlst, pdc etc
    # we want to fill those with "No assembly given"
    # It can also happen that an assembly was given but the analysis returned a np.nan
    # in that case we fillna with "-"
    only_reads_bc_pos = temp_pivot[("", "Pipeline")]=="only_reads"
    temp_pivot.loc[only_reads_bc_pos, idx[:,no_assembly_given_cols]] = temp_pivot.loc[only_reads_bc_pos, idx[:,no_assembly_given_cols]].fillna("No assembly given")
    temp_pivot.loc[~only_reads_bc_pos, idx[:,no_assembly_given_cols]] = temp_pivot.loc[~only_reads_bc_pos, idx[:,no_assembly_given_cols]].fillna("-")

    # Format pivot with colors
    temp = (
        temp_pivot
        .style.applymap(
            color_cells_v2, #subset=[i for i in temp_pivot.columns if i[1] not in no_assembly_given_cols ]
        )
        .apply_index(highlight_header, axis="columns", level=[0, 1])
        .apply_index(highlight_header, axis="index")
    )

    # Writer object
    writer = pd.ExcelWriter(output_file, engine="xlsxwriter")
    sheet_name = "ARVIA"

    # Convert the styled dataframe to an XlsxWriter Excel object in specific sheet
    temp.to_excel(writer, sheet_name=sheet_name)

    # Select sheet and apply formatting
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    worksheet.autofit()  # autofit row widths
    worksheet.set_row(1, 45)  # height of row
    worksheet.set_column_pixels(first_col=6, last_col=len(temp.columns)+1, width=190) # set max width from column N to last
    worksheet.set_column_pixels(first_col=3, last_col=4, width=85) # set max width for mlst and mlst_model
    worksheet.set_column(0, 0, 15)  # width of first column
    worksheet.freeze_panes(2, 1)  # freeze first 2 rows and first column

    # Save
    workbook.close()

    # ---- Reload excel ----
    # Load again and delete multiindex row (bc), should be third line
    # If it is not deleted it could be hidden, but when the user orders the excel
    # other row will be hidden. Thus, we just try to remove it
    wb = openpyxl.load_workbook(output_file)
    sheet = wb[sheet_name]
    assert sheet.cell(row=3, column=1).value == "bc", f"Unexpected: Row 3, column 1 was not 'bc'. Stopping just in case, check {output_file=}"
    sheet.delete_rows(3, 1)

    # Format A2 to add hyperlink
    sheet['A2'].hyperlink = "https://github.com/pablo-aja-macaya/ARVIA"
    sheet['A2'].value = "ARVIA Github"
    sheet['A2'].style = "Hyperlink"
    sheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical="center")
    sheet['A2'].font = openpyxl.styles.Font(name='Calibri', bold=True, underline="single")
    
    # Format A1
    sheet['A1'].value = ""

    # Save again
    wb.save(output_file)
    
    # ---- Save to .tsv also ----
    # Drop first row in multilevel column index
    temp_pivot.columns = temp_pivot.columns.droplevel(0)
    
    # Remove line breaks
    temp_pivot.columns = [i.replace("\n", "") for i in temp_pivot.columns]
    
    # Save wide table to tsv
    temp_pivot.to_csv(f"{Path(output_file).parent}/{Path(output_file).stem}.tsv", sep="\t")
    
    # Save long table to tsv
    melted_df = pd.melt(temp_pivot.reset_index(), id_vars='bc', var_name="section", value_name="value")
    melted_df.to_csv(f"{Path(output_file).parent}/{Path(output_file).stem}.long.tsv", sep="\t", index=None)
    
    # l = []
    # for row in melted_df.to_dict("records"):
    #     l += [
    #         {
    #             "bc": row["bc"],
    #             "section": row["section"],
    #             "value": i,
    #             "presence": "yes"
    #         } for i in row["value"].split(", ")
    #     ]

    # # comparative_pivot = pd.DataFrame(l).drop_duplicates().pivot(index=["section", "value"], columns="bc", values="presence").fillna("-")
    # comparative_pivot = melted_df.pivot(index="section", columns="bc", values="value").fillna("-")
    # comparative_pivot.to_excel("/home/usuario/Proyectos/Results/tests/arvia/test.xlsx")
    # print("\n", comparative_pivot.sort_values("section"))
    # raise Exception

    return True

