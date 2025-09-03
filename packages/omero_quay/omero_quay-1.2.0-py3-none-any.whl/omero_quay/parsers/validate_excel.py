from __future__ import annotations

import itertools
import logging
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from ..core.config import get_conf
from ..core.connect import omero_conn
from .errors import (
    ExcelValidationError,
    MissingExcelSheetError,
    NoExistingOMEROUser,
    UserNotInOMEROGroup,
)

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)


def validate_user_existence(user_login):
    """
    Validate user existence in OMERO database. Raise error if not in database.
    """
    conf = get_conf()
    with omero_conn(conf) as conn:
        try:
            _ = next(
                iter(
                    conn.getObjects("Experimenter", attributes={"omeName": user_login})
                )
            )
        except StopIteration as err:
            raise NoExistingOMEROUser(user_login) from err


def check_if_user_in_group(user_login, group_name):
    """
    Check if OMERO user is in OMERO group. Raise error if not.
    """
    conf = get_conf()
    with omero_conn(conf) as conn:
        group = conn.getObject("ExperimenterGroup", attributes={"name": group_name})
    if not group:
        # New group
        return True

    members = [e._omeName._val for e in group.linkedExperimenterList()]
    if user_login not in members:
        raise UserNotInOMEROGroup(user_login, group_name)

    return True


def check_everything(xlsx_path):
    validate_xlsx_structure(xlsx_path)
    validate_assays_parent_studies(xlsx_path)
    check_study_owner_is_parent_owner_or_contributor(xlsx_path)
    check_assay_owner_is_parent_study_owner_or_contributor(xlsx_path)
    check_if_assay_path_is_subpath_of_other_assay(xlsx_path)
    return True


def validate_xlsx_structure(xlsx_path):
    """
    Validates the structure of the XLSX file.

    Conditions:
        - Global:
            - sheet `"Investigation", "Study", "Assay"` are present
        - Investigation
            - Must have a name
            - No space in investigation name
            - manager must be a valid user
            - owners: at least one valid user
            - contributors: zero or more valid users
            - collaborators: zero or more valid users
            - no shared users between owners, contributors and collaborators (**not** true for manager)
        - Study
            - Must have a valid owner
            - Must have a name
            - Must have valid parent investigation (name present in the "investigation" sheet)
        - Assay
            - Must have a valid owner
            - All the paths must be different

    """
    # open an Excel file and return a workbook
    wb = load_workbook(xlsx_path, read_only=True)
    for sheet in ["Investigation", "Study", "Assay"]:
        if sheet not in wb.sheetnames:
            raise MissingExcelSheetError(sheet, Path(xlsx_path).name)

    investigation_df = pd.read_excel(xlsx_path, sheet_name="Investigation")
    investigation_df.dropna()
    if "name" not in investigation_df.columns:
        msg = "No name column in investigation sheet"
        raise ExcelValidationError(msg)
    for i, investigation in investigation_df.iterrows():
        # TODO if not investigation["name"]
        if pd.isna(investigation["name"]) or not (
            isinstance(investigation["name"], str)
        ):
            msg = f"The investigation at row {i} does not have a name"
            raise ExcelValidationError(msg)
        if bool(re.search(r"\s", investigation["name"])) is True:
            msg = f"{investigation['name']} has spaces and is not a valid name for an investigation"
            raise ExcelValidationError(msg)

        investigation_owners = investigation["owners"]
        investigation_contributors = investigation["contributors"]
        investigation_collaborators = investigation["collaborators"]

        if pd.isna(investigation["contributors"]):
            investigation_contributors = ""
        if pd.isna(investigation["collaborators"]):
            investigation_collaborators = ""

        if (investigation_contributors == "") and (investigation_collaborators == ""):
            continue

        owners = set(investigation_owners.split(","))
        contributors = set(investigation_contributors.split(","))
        collaborators = set(investigation_collaborators.split(","))

        if owners.intersection(contributors):
            msg = f"""A owner can't be a contributor and vice-versa for investigation {i}.
                Problem with users {owners.intersection(contributors)}"""
            raise ExcelValidationError(msg)
        if contributors.intersection(collaborators):
            msg = f"""A contributor can't be a collaborator and vice-versa for investigation {i}.
                Problem with users {contributors.intersection(collaborators)}"""
            raise ExcelValidationError(msg)
        if collaborators.intersection(owners):
            msg = f"""A collaborator can't be a owner and vice-versa for investigation {i}.
                Problem with users {collaborators.intersection(owners)}"""
            raise ExcelValidationError(msg)

    study_df = pd.read_excel(xlsx_path, sheet_name="Study")
    study_df.dropna()
    if "name" not in study_df.columns:
        msg = "No name column in study sheet"
        raise ExcelValidationError(msg)
    for i, study in study_df.iterrows():
        if pd.isna(study["name"]) or not (isinstance(study["name"], str)):
            msg = f"Study {i} does not have name"
            raise ExcelValidationError(msg)

    assay_df = pd.read_excel(xlsx_path, sheet_name="Assay")
    assay_df.dropna()
    assays_paths_list = assay_df["path"]
    if "owner" not in assay_df.columns:
        msg = "No owner column in assay sheet"
        raise ExcelValidationError(msg)
    if assays_paths_list.is_unique is False:
        msg = f"At least two paths for assays are the same: {assays_paths_list}"
        raise ExcelValidationError(msg)

    for i, study in study_df.iterrows():
        if study["parent"] not in investigation_df["name"].tolist():
            msg = f"Study {i}, {study['name']} does not have investigation matching in investigation sheet - DEBUG: Study parent investigation: {study['parent']} ; Investigations names: {investigation_df['name']}"
            raise ExcelValidationError(msg)


def validate_logins(xlsx_path):
    log.info("No login validation for now on file %s", xlsx_path)
    return True


def validate_assays_parent_studies(xlsx_path):
    """
    Check if assays are in studies in study sheet of XLSX file.

    Conditions:
        - Assay:
            - parent must be a valid study (name in the Study sheet)
    """
    assays_df = pd.read_excel(xlsx_path, sheet_name="Assay")
    studies_df = pd.read_excel(xlsx_path, sheet_name="Study")

    assays_parents = set(assays_df["parent"])
    studies_names = set(studies_df["name"])
    if not assays_parents.issubset(studies_names):
        msg = f"Assay does not have study matching in investigation sheet. Check assays parents {assays_parents} and studies names {studies_names}."
        raise ExcelValidationError(msg)


def check_study_owner_is_parent_owner_or_contributor(xlsx_path):  # OK
    """
    Check if study owner is parent investigation owner or contributor.
    """
    studies_df = (
        pd.read_excel(xlsx_path, sheet_name="Study")
        .replace(pd.NA, "", inplace=False)
        .replace(np.nan, "", inplace=False)
    )
    investigations_df = (
        pd.read_excel(xlsx_path, sheet_name="Investigation")
        .replace(pd.NA, "", inplace=False)
        .replace(np.nan, "", inplace=False)
    )
    for i, study in studies_df.iterrows():
        study_parent = study["parent"]
        study_owner = study["owner"]
        for _, investigation in investigations_df.iterrows():
            investigation_name = investigation["name"]
            investigations_owners = set(investigation["owners"].split(","))
            investigations_contributors = set(investigation["contributors"].split(","))

            if (investigation_name == study_parent) and (
                study_owner
                not in investigations_owners.union(investigations_contributors)
            ):
                msg = f"""Study owner is not an owner nor a contributor of its parent investigation.
                           for study {i}, name {study["name"]}"""
                raise ExcelValidationError(msg)


def check_assay_owner_is_parent_study_owner_or_contributor(xlsx_path):
    """
    Check if assay owner is parent study owner or contributor.
    """
    assays_df = (
        pd.read_excel(xlsx_path, sheet_name="Assay")
        .replace(pd.NA, "")
        .replace(np.nan, "")
    )
    studies_df = (
        pd.read_excel(xlsx_path, sheet_name="Study")
        .replace(pd.NA, "")
        .replace(np.nan, "")
    )
    investigations_df = (
        pd.read_excel(xlsx_path, sheet_name="Investigation")
        .replace(pd.NA, "")
        .replace(np.nan, "")
    )
    for i, assay in assays_df.iterrows():
        assay_owner = assay["owner"]
        assay_parent_study = assay["parent"]

        for _, study in studies_df.iterrows():
            study_name = study["name"]
            study_parent = study["parent"]
            for _, investigation in investigations_df.iterrows():
                investigation_name = investigation["name"]
                investigations_owners = set(investigation["owners"].split(","))
                investigations_contributors = set(
                    investigation["contributors"].split(",")
                )

                if (
                    (investigation_name == study_parent)
                    and (study_name == assay_parent_study)
                    and assay_owner
                    not in investigations_owners.union(investigations_contributors)
                ):
                    msg = f"""Assay owner is not an owner nor a contributor of its parent investigation.
                           for assay {i}, name {assay["name"]}"""
                    raise ExcelValidationError(msg)


def check_if_assay_path_is_subpath_of_other_assay(xlsx_path):
    """
    Check path validity of assays:

    Conditions:
        - Assay:
            - Path must exist as a subdir of the investigation in the buffer machine
            - No assay path cannot be a subpath of another assay (see `pathlib.Path`)
            - This is invalid:

                | path         | name   |
                | ------------ | ------ |
                | Dir1         | assay1 |
                | Dir1/SubDir1 | assay2 |
    """
    assay_df = pd.read_excel(xlsx_path, sheet_name="Assay")
    assays_paths_list = assay_df["path"]

    for path, other in itertools.permutations(assays_paths_list, 2):
        if Path(path).is_relative_to(other):
            msg = f"An assay path: {path} is subpath of another one: {other}"
            raise ExcelValidationError(msg)
